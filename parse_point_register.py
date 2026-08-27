# -*- coding: utf-8 -*-
"""
지자기 점조서(★.계속~지자기 점조서) 파싱 — 관측망 30점의 **원본 계보**
========================================================================

`data/geomag_network_30.csv`(30점 좌표표)의 출처를 확정하기 위해 지리원 점조서
원본을 직접 읽는다. 점조서는 측점마다 시트 하나이며 다음을 담는다:

    점의부호(1-8-3 등) · 도엽번호·명 · 소재지 · 관할 · 매설일·매설자 ·
    경위도(도분초) · 표고 · 관측 이력(1984~)

⚠️ **점의부호 끝자리가 «설치 차수»다.** 표석을 이설·재설하면 그 자리가 올라간다:

    임계 1-8-3 (2004.10.27 매설) → 삼척 1-8-4 (2022.10.12 매설)
    경주 1-11-3               → 영천 1-11-4
    양산 1-12-3               → 언양 1-12-4

즉 **같은 계보(1-8-*)의 최신 차수가 현재 관측점**이고, 앞 차수는 폐지된 옛
위치다. 시트가 34개인데 관측망이 30점인 이유가 이것이다(계보 4개가 두 시트씩).

⚠️ **경위도 라벨이 뒤바뀌어 있다** — Φ 행에 경도, λ 행에 위도가 들어 있다.
   값 범위(경도 124~132 · 위도 33~39)로 판별한다.

    python parse_point_register.py            # 요약
    python parse_point_register.py --csv      # docs/output 에 대조표 저장
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
REGISTER_XLSX = Path(
    r"D:\LX_yoons\2026_research\2026_지자기 연구\20260811_지리원 야장 자료"
    r"\지구물리측량\지자기측량\2025지자기"
    r"\★.계속~지자기 점조서-문서2025.12.15수정됨.ver.xlsx")

SKIP_SHEETS = {"기본1(양식)", "Sheet2", "지자기 성과고시 내역2021까지"}
CODE_RE = re.compile(r"^1-\d{1,2}(?:-\d{1,2})?$")
SHEET_RE = re.compile(r"^(N[IJ]\s?52-[\d\-]+)\s*,?\s*(.*)$")


def _s(v):
    return "" if v is None else str(v).replace("\n", " ").strip()


def _f(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def parse_sheet(ws):
    """시트 하나 → 점조서 레코드."""
    grid = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 70),
                                         max_col=min(ws.max_column, 14)), 1):
        for c, cell in enumerate(row, 1):
            if cell.value is not None:
                grid[(r, c)] = cell.value

    d = {"시트": ws.title}

    # ── 점의부호 · 도엽 ──────────────────────────────────────
    for (r, c), v in grid.items():
        t = _s(v)
        if CODE_RE.match(t) and not d.get("점의부호"):
            d["점의부호"] = t
        m = SHEET_RE.match(t)
        if m and not d.get("도엽번호"):
            d["도엽번호"], d["도엽명"] = m.group(1), m.group(2).strip()

    # ── 라벨 기준 값 (라벨 셀 오른쪽에서 첫 비어있지 않은 값) ──
    LABELS = ("매설자", "매설일", "점의소재지", "경로", "관할", "표고", "경위도")

    def right_of(label, within=4):
        """라벨 오른쪽 첫 값. **다른 라벨을 값으로 집으면 빈 값**으로 본다."""
        for (r, c), v in grid.items():
            if _s(v).replace(" ", "").startswith(label.replace(" ", "")):
                for cc in range(c + 1, c + 1 + within):
                    t = _s(grid.get((r, cc)))
                    if not t:
                        continue
                    if any(t.replace(" ", "").startswith(L) for L in LABELS):
                        return "", r          # 값이 비어 다음 라벨이 잡힌 것
                    return t, r
        return "", None

    d["소재지"], _ = right_of("점의소재지")
    d["관할"], _ = right_of("관할")
    d["매설일"], r_ins = right_of("매설일")
    d["매설자"], _ = right_of("매설자")
    d["경로"], _ = right_of("경로")

    # ── 경위도 — Φ/λ 라벨이 뒤바뀌어 있어 값 범위로 판별 ──────
    r_ll = None
    for (r, c), v in grid.items():
        if _s(v).replace(" ", "") == "경위도":
            r_ll = r
            break
    if r_ll:
        for rr in (r_ll, r_ll + 1):
            dms = [_f(grid.get((rr, c))) for c in (3, 4, 5, 6)]
            if dms[0] is None or dms[1] is None:
                continue
            dd = dms[0] + (dms[1] or 0) / 60 + ((dms[2] or 0) + (dms[3] or 0)) / 3600
            if 124 <= dms[0] <= 132:
                d["경도"], d["경도_dms"] = dd, dms
            elif 33 <= dms[0] <= 39:
                d["위도"], d["위도_dms"] = dd, dms
        for rr in (r_ll, r_ll + 1):
            h = _f(grid.get((rr, 8)))
            if h is not None and 0 <= h < 2000:
                d["표고"] = h

    # ── 관측 이력 ────────────────────────────────────────────
    obs_rows = sorted(r for (r, c), v in grid.items()
                      if c == 1 and _s(v).replace(" ", "") == "관측일자")
    dates, recs = [], []
    if obs_rows:
        # ⚠️ 시트마다 «관측일자» 표가 두 벌이다 — 두 번째는 빈 양식이라
        #    거기까지 읽으면 최종관측이 1984 로 되돌아간다(실제로 그랬다).
        stop = obs_rows[1] if len(obs_rows) > 1 else min(ws.max_row, 70) + 1
        for rr in range(obs_rows[0] + 2, stop):
            t = _s(grid.get((rr, 1)))
            if not re.match(r"^(19|20)\d{2}", t):
                continue
            dates.append(t)
            # B/C/D=편각 도·분·초 · E/F/G=복각 · H=전자력 · K=장비
            # ⚠️ 부호는 «도» 칸에만 붙는다 — 분·초는 크기이므로 절대값을
            #    쓰고 도의 부호를 곱해야 한다(한국 편각은 서편각 = 음수).
            rec = {"일자": t, "편각": None, "복각": None,
                   "전자력": None, "장비": _s(grid.get((rr, 11)))}
            for key, cols in (("편각", (2, 3, 4)), ("복각", (5, 6, 7))):
                dg, mi, se = (_f(grid.get((rr, c))) for c in cols)
                if dg is None:
                    continue
                sign = -1.0 if dg < 0 else 1.0
                rec[key] = sign * (abs(dg) + abs(mi or 0) / 60
                                   + abs(se or 0) / 3600)
            f = _f(grid.get((rr, 8)))
            # 한반도 총자력 범위 밖은 기재 오류로 본다(서산 2007 = 60,812).
            rec["전자력"] = f if (f is not None and 40000 <= f <= 60000) else None
            if any(rec[k] is not None for k in ("편각", "복각", "전자력")):
                recs.append(rec)
    d["관측이력"] = dates
    d["관측횟수"] = len(dates)
    d["최종관측"] = dates[-1] if dates else ""
    d["관측기록"] = recs
    if recs:
        last = recs[-1]
        d["편각"] = last["편각"]
        d["복각"] = last["복각"]
        d["전자력"] = last["전자력"]
        d["관측장비"] = last["장비"]
        d["최종관측"] = last["일자"]

    # ── 비고성 문구(이설·망실·재설) 수집 ────────────────────
    notes = []
    for (r, c), v in grid.items():
        t = _s(v)
        if any(k in t for k in ("망실", "이설", "재설", "폐기", "복구", "훼손")):
            notes.append(t[:120])
    d["비고"] = " / ".join(sorted(set(notes))[:4])
    return d


def lineage(code):
    """점의부호에서 계보(1-8) 와 차수(3) 를 분리."""
    if not code:
        return "", None
    parts = code.split("-")
    if len(parts) >= 3:
        return "-".join(parts[:2]), int(parts[2])
    return "-".join(parts[:2]), 1


def load_register(path=REGISTER_XLSX):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    recs = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        recs.append(parse_sheet(wb[name]))
    wb.close()
    for r in recs:
        r["계보"], r["차수"] = lineage(r.get("점의부호", ""))
    return recs


_YEAR_RE = re.compile(r"(19|20)\d{2}")


def _bury_year(rec):
    """매설일에서 연도. 없으면 최종 관측 연도로 대신한다."""
    for key in ("매설일", "최종관측"):
        m = _YEAR_RE.search(rec.get(key, "") or "")
        if m:
            return int(m.group(0))
    return 0


def current_points(recs):
    """
    계보마다 **현재 관측점** 하나를 고른다.

    ⚠️ 점의부호 끝자리(차수)만으로 고르면 안 된다 — 실제로 뒤집힌 계보가 있다:
    1-23 은 「양산 1-23-3(2005 매설)」과 「언양 1-23-1(2024.10.16 매설)」인데
    차수가 낮은 쪽이 더 새로 매설됐다. **매설일을 1순위**로, 같으면 차수를 쓴다.
    """
    best = {}
    for r in recs:
        k = r["계보"]
        key = (_bury_year(r), r["차수"] or 0)
        if k not in best or key > best[k][0]:
            best[k] = (key, r)
    return sorted((v[1] for v in best.values()),
                  key=lambda r: (int(r["계보"].split("-")[1]) if "-" in r["계보"] else 0))


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    if not REGISTER_XLSX.exists():
        print(f"[중단] 점조서 없음: {REGISTER_XLSX}")
        return 1
    recs = load_register()
    cur = current_points(recs)
    sup = [r for r in recs if r not in cur]

    print(f"점조서 시트 {len(recs)}개 · 계보 {len(cur)}개(= 현재 관측점) · "
          f"폐지 차수 {len(sup)}개")
    print(f"\n{'계보':<7}{'점의부호':<10}{'도엽명':<8}{'매설일':<13}"
          f"{'위도':>11}{'경도':>12}{'관측':>5}  최종관측")
    for r in cur:
        print(f"{r['계보']:<7}{r.get('점의부호',''):<10}{r.get('도엽명',''):<8}"
              f"{r.get('매설일','')[:12]:<13}{r.get('위도',float('nan')):>11.6f}"
              f"{r.get('경도',float('nan')):>12.6f}{r['관측횟수']:>5}  {r['최종관측']}")

    print(f"\n■ 폐지된 옛 차수 {len(sup)}개 (이설·재설로 대체)")
    for r in sup:
        nxt = next((c for c in cur if c["계보"] == r["계보"]), None)
        print(f"   {r.get('점의부호','')} {r.get('도엽명','')} "
              f"(매설 {r.get('매설일','')[:12]}) → "
              f"{nxt.get('점의부호','') if nxt else '?'} "
              f"{nxt.get('도엽명','') if nxt else '?'}"
              f" (매설 {nxt.get('매설일','')[:12] if nxt else '?'})")

    note = [r for r in recs if r["비고"]]
    if note:
        print(f"\n■ 비고(망실·이설·재설) {len(note)}건")
        for r in note:
            print(f"   {r.get('도엽명') or r['시트']}: {r['비고'][:110]}")

    if "--csv" in sys.argv:
        import csv
        import datetime as dt
        out = ROOT / "docs" / "output"
        out.mkdir(parents=True, exist_ok=True)
        f = out / f"{dt.datetime.now():%Y%m%d_%H%M%S}_점조서_현재관측점.csv"
        with open(f, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow(["계보", "점의부호", "도엽번호", "도엽명", "소재지", "관할",
                        "매설일", "매설자", "위도", "경도", "표고",
                        "관측횟수", "최종관측"])
            for r in cur:
                w.writerow([r["계보"], r.get("점의부호", ""), r.get("도엽번호", ""),
                            r.get("도엽명", ""), r.get("소재지", ""), r.get("관할", ""),
                            r.get("매설일", ""), r.get("매설자", ""),
                            r.get("위도", ""), r.get("경도", ""), r.get("표고", ""),
                            r["관측횟수"], r["최종관측"]])
        print(f"\n[저장] {f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
