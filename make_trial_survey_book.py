# -*- coding: utf-8 -*-
"""
지자기 시험 탐사 표준 야장 — `docs/output/*_시험탐사_표준야장.xlsx`
====================================================================

선점 검토 50지점(등급 A 5 · B 도엽대표 29 · 기존 관측망 선점대상 16)의
**시험 탐사용 표준 야장**.

    python make_trial_survey_book.py

## 이 야장이 존재하는 이유

⚠️ **야장에 칸이 없으면 그 값은 어떤 분석으로도 복원되지 않는다.** 이 프로젝트가
실제로 겪은 일이다 — 야장 68개를 전수로 훑었더니 **F 측정일시 실기재가 0건**이라
세션 단위 외부장 보정이 «원리적으로» 불가능해졌다(CLAUDE.md 12-C B절).

## Codex 독립 검토 반영 (2026-09-07, Full Review)

`2026_nscenter` 의 codex-review 프로토콜로 초안을 Red Team 검토에 넘겨
**Critical 6 · Major 6 · Minor 2 / MAJOR REVISION REQUIRED** 를 받았고 전부 반영했다.
요청·응답 원문은 `2026_nscenter/docs/codex/20260907_111551_Geomag_TrialSurvey_*.md`.

지적의 뿌리는 하나였다 — **「기록하라」는 문장은 있는데 관측값·시각·위치·기기·
방향을 하나의 레코드로 묶는 구조가 없다.** 그래서 이 야장의 중심은 서식이 아니라
**관측행 스키마**다:

    Site — Visit — Date — Session — Seq — Component — Location — Instrument

| 지적 | 반영 |
|---|---|
| **C1** 1시간 간격 2회 반복으로는 일변화와 공간변화를 분리할 수 없다 | 「⑥ 기준자력계」 시트 신설(연속 기록 필수). 측선은 **중심점 재관측(A–B–A)·역순 재측선**으로 바꿨다. 연속기록이 없으면 결과를 「시간변화 미분리」로 표시하고 **선점 확정 근거로 쓰지 않는다** |
| **C2** D·I·F 를 묶는 고유키와 개별 타임스탬프가 없다 | 「⑦ 절대측정 관측행」이 위 8-키를 강제한다. F 는 **판독마다** 초 단위 시각 |
| **C3** 근거 없는 임계가 법정요건과 섞여 최종 판정을 정한다 | **법정·기록 게이트(⑨)** 와 **정량 지표(⑩)** 를 분리. 임계는 **파일럿 전 미확정**으로 명시하고 판정 수식에 넣지 않았다 |
| **C4** 방위각 결정·방향·gon 이 구조로 강제되지 않는다 | 원시각 **gon 고정**(단위 칸 편집 불가), 도 환산은 읽기전용 계산. FROM/TO·마크 ID·결정방법·근거좌표·이전값 대비 차·재결정 사유 |
| **C5** 기존점 「직접 대조 가능」은 확인된 실패 이력과 충돌한다 | 성분별 **비교가능 상태 플래그**로 쪼갰다. 좌표 충돌점은 **HOLD** |
| **C6** §13② 를 도상 확인만으로 확정한다 | 도상거리·현장 확인거리를 나누고 **필수 게이트**로 연결 |
| **M1** 2.8 km 예측구배와 수 m 현장구배는 비교 대상이 아니다 | 예측값은 **층화·탐색 변수**로만. 자료공백 5점은 `NOT AVAILABLE` 유지, 감점 금지 |
| **M2** 측선 규모가 크지만 2차원 탐지 충분성은 보장 못 한다 | **파일럿 먼저** → 최소 공통 설계 + 이상 시 추가 격자(단계적) |
| **M3** 원시 판독값이 보존되지 않는다 | 판독 3회를 **각각** 남기고 중앙값은 계산식 |
| **M4** 검증 방식이 새 형식 오류를 만든다 | 시각은 실제 datetime(초)+시간대, UTC 자동. 범위 이탈은 **차단이 아니라 경고** |
| **M5** 기기 상호대조가 시간변화·위치차와 안 나뉜다 | **교차배치(swap)** 절차. 2 nT 는 경고값으로 격하 |
| **M6** 센서 위치·높이 재현성 필드가 없다 | 실제 높이·E/N 오프셋·Location ID·재설치 확인 |
| **m1** 중복 측점 처리 불명확 | **Point ID** 유일 부여 |
| **m2** 결측·미실시·불합격이 구분 안 된다 | 상태코드 5종 |

⚠️ **Codex 가 Human 확인을 요구한 6건은 이 야장이 답하지 않는다** — §19·§20 의
공식 적용단위, 기존점 표석·표지 동일성, 기준자력계 투입 여부, 임계값 승인,
§13② 시설 분류, 상시관측소 원자료 이용조건. 「⑫ 확인 필요」 시트에 그대로 남겼다.

## 자료 흐름

`trial_survey_points.load_points()` 하나만 읽는다 — 그것이 `export_selection_50`
을 호출하므로 **지도·선점검토 엑셀과 언제나 같은 50점**이다.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import trial_survey_points as TP

ROOT = Path(__file__).parent
OUT_DIR = ROOT / "docs" / "output"

# 좌표 동일성이 해소되지 않아 시험 탐사 판정을 보류하는 점 (Codex C5).
#   미원 — '10~'19 성과표와 2019 야장 좌표가 353 m 어긋난다. 같은 표석인지
#          확인되기 전에는 신규 관측을 과거 성과와 견줄 수 없다.
# ⚠️ 남양(현재 좌표 미상)은 애초에 선점 대상 16점에 없어 여기 들어가지 않는다.
HOLD_SITES = {
    "미원": "좌표 353 m 불일치 — 표석 동일성 확인 전 과거 성과 대조 보류",
}

# 상태코드 (Codex m2) — 빈칸이 0·불합격으로 집계되는 것을 막는다.
STATUS = ["NOT AVAILABLE", "NOT OBSERVED", "INVALID", "NOT APPLICABLE", "PENDING"]

FONT = "맑은 고딕"
F_TITLE = Font(name=FONT, size=13, bold=True, color="FFFFFF")
F_SEC = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
F_LBL = Font(name=FONT, size=10, bold=True)
F_VAL = Font(name=FONT, size=10)
F_SMALL = Font(name=FONT, size=9, color="555555")
F_WARN = Font(name=FONT, size=9, bold=True, color="B2182B")
F_HDR = Font(name=FONT, size=9, bold=True, color="FFFFFF")
F_KEY = Font(name=FONT, size=9, bold=True, color="FFFFFF")

FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_KEY = PatternFill("solid", fgColor="2E5C8A")     # 연결키 열
FILL_SEC = PatternFill("solid", fgColor="4472C4")
FILL_LBL = PatternFill("solid", fgColor="F2F2F2")
FILL_FIELD = PatternFill("solid", fgColor="FFF8E1")   # 현장 기입란
FILL_AUTO = PatternFill("solid", fgColor="E8F0E8")    # 사전 채움·계산
FILL_WARN = PatternFill("solid", fgColor="FDE9E7")
FILL_LOCK = PatternFill("solid", fgColor="EDEDED")    # 편집 금지

AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_TL = Alignment(horizontal="left", vertical="top", wrap_text=True)
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BAND = {"A": "E2F0E4", "B": "E4EDF7", "기존점": "FDF2E0"}

DTFMT = "yyyy-mm-dd hh:mm:ss"


# ── 셀 도우미 ────────────────────────────────────────────────
def title(ws, span, text, row=1):
    ws.merge_cells(f"A{row}:{span}{row}")
    c = ws[f"A{row}"]
    c.value, c.font, c.fill, c.alignment = text, F_TITLE, FILL_TITLE, AL_L
    ws.row_dimensions[row].height = 26


def note(ws, row, span, text, warn=False):
    ws.merge_cells(f"A{row}:{span}{row}")
    c = ws[f"A{row}"]
    c.value, c.alignment = text, AL_TL
    c.font = F_WARN if warn else F_SMALL
    if warn:
        c.fill = FILL_WARN
    segs = str(text).split("\n")
    ws.row_dimensions[row].height = 14 * sum(max(1, -(-len(s) // 78)) for s in segs) + 5


def table(ws, row, headers, widths, keys=()):
    """머리글. `keys` 에 든 열은 **연결키**로 진한 색을 준다(Codex C2)."""
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F_KEY if h in keys else F_HDR
        c.fill = FILL_KEY if h in keys else FILL_TITLE
        c.alignment, c.border = AL_C, BORDER
    ws.row_dimensions[row].height = 34
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def blank_rows(ws, start, n, ncol, fill=FILL_FIELD, fmts=None):
    """현장 기입용 빈 행."""
    for r in range(start, start + n):
        for j in range(1, ncol + 1):
            c = ws.cell(row=r, column=j)
            c.border, c.fill, c.font, c.alignment = BORDER, fill, F_VAL, AL_C
            if fmts and j in fmts:
                c.number_format = fmts[j]


def dv(ws, ref, items):
    d = DataValidation(type="list", formula1='"' + ",".join(items) + '"',
                       allow_blank=True, showDropDown=False)
    ws.add_data_validation(d)
    d.add(ref)
    return d


def dv_warn(ws, ref, lo, hi, msg):
    """⚠️ 저장을 **차단하지 않는다**(Codex M4) — 실제 이상값을 잃지 않기 위해서다."""
    d = DataValidation(type="decimal", operator="between", formula1=lo, formula2=hi,
                       allow_blank=True, showErrorMessage=True, errorStyle="warning",
                       errorTitle="예상 범위 밖", error=msg)
    ws.add_data_validation(d)
    d.add(ref)
    return d


# ══════════════════════════════════════════════════════════════
def sheet_intro(wb):
    ws = wb.create_sheet("① 기입규약")
    for col, w in zip("ABCDEFGH", [22, 20, 20, 20, 20, 20, 20, 20]):
        ws.column_dimensions[col].width = w
    title(ws, "H", "지자기 시험 탐사 표준 야장 — 기입규약")
    r = 2
    note(ws, r, "H",
         "이 야장은 서식이 아니라 «재발 방지 장치»다. 과거 야장 68개를 전수로 훑었더니 "
         "F 측정일시 실기재가 0건이어서 세션 단위 외부장 보정이 원리적으로 불가능해졌다. "
         "칸이 없으면 그 값은 어떤 분석으로도 복원되지 않는다."); r += 1
    note(ws, r, "H",
         "⚠ 이 양식은 Codex 독립 검토(2026-09-07, Critical 6 · Major 6)를 반영한 판본이다. "
         "정량 임계값은 «파일럿 전 미확정»이며 판정 수식에 넣지 않았다 — 임계로 50점을 "
         "바로 가르지 말 것.", warn=True); r += 2

    def block(head, rows):
        nonlocal r
        ws.merge_cells(f"A{r}:H{r}")
        c = ws[f"A{r}"]
        c.value, c.font, c.fill, c.alignment = head, F_SEC, FILL_SEC, AL_L
        r += 1
        for k, v in rows:
            ws.cell(row=r, column=1, value=k).font = F_LBL
            ws.cell(row=r, column=1).fill = FILL_LBL
            ws.cell(row=r, column=1).alignment = AL_C
            ws.cell(row=r, column=1).border = BORDER
            ws.merge_cells(f"B{r}:H{r}")
            c2 = ws.cell(row=r, column=2, value=v)
            c2.font, c2.alignment, c2.border = F_VAL, AL_L, BORDER
            ws.row_dimensions[r].height = 14 * max(1, -(-len(v) // 92)) + 6
            r += 1
        r += 1

    block("단위 — 논쟁 없이 고정한다", [
        ("각도(현장 원시)", "곤 gon (1회전 400 · 대척 200 gon). 현장 판독은 반드시 gon 으로 적는다. "
                          "도(°) 환산값은 읽기전용 계산 칸에만 둔다 — 단위를 고르게 하면 "
                          "입력 후 단위를 바꿔도 값이 안 바뀌어 의미만 뒤집힌다."),
        ("총자력 F", "nT, 소수 1자리 (§17 0.1 nT 판독)"),
        ("좌표", "십진도. 위도·경도를 «각각 다른 열»에 숫자로 적고 좌표계·취득방법·정확도를 함께 남긴다."),
        ("거리·높이", "m, 소수 2자리"),
    ])
    block("시각 — 세 자료의 시간대가 서로 다르다", [
        ("기입", "날짜 + 시:분:초 를 한 칸에 (예 2026-10-14 13:24:05). 분까지만 적으면 세션 결합이 깨진다."),
        ("시간대", "야장은 KST. 상시관측소 CYG·Kp 는 UT, KASA 캐시는 KST — 결합 지점마다 확인할 것. "
                 "UTC 환산은 계산 칸이 자동으로 만든다(KST − 9h)."),
        ("자정 경계", "날짜를 함께 적으므로 자정을 넘겨도 세션이 끊기지 않는다."),
    ])
    block("방위각 — 정·역 혼동이 29쌍 있었다", [
        ("규약", "진북 기준 · 북 = 0° · 시계방향 · **FROM = 기준점, TO = 방위표지**. 반대 방향을 적으면 그 방문의 편각이 통째로 틀어진다."),
        ("마크 식별", "표지마다 고정 Mark ID 를 부여하고 시준한 «지점»(표지 상단/중심 등)까지 적는다."),
        ("참방위각", "결정방법(천문·자이로·RTK 장기선·좌표계산)과 근거좌표·좌표 정확도를 남긴다. "
                   "이전 확정값과의 차이와 재결정 사유를 반드시 적는다 — 재방문 잔여 RMS 33.7′ 의 원인이 이것이다."),
    ])
    block("상태코드 — 빈칸을 0 으로 세지 않기 위해", [
        (s, d) for s, d in zip(STATUS, [
            "자료가 존재하지 않음 (예: 항공자력 미측선 구역)",
            "관측을 실시하지 않음",
            "관측했으나 무효 (장비 오류·기입 오류 등, 사유 필수)",
            "그 점에는 해당하지 않는 항목",
            "미정 — 확인 대기",
        ])
    ])
    block("색 규약", [
        ("연한 노랑", "현장 기입란"),
        ("연한 초록", "사전 채움 또는 계산값 — 현장에서 고치지 않는다"),
        ("회색", "편집 금지"),
        ("연한 적색", "주의·경고"),
    ])
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_points(wb, pts):
    ws = wb.create_sheet("② 대상 50점")
    n = {"A": 0, "B": 0, "기존점": 0}
    for p in pts:
        n[p["구분"]] += 1
    title(ws, "S", f"시험 탐사 대상 {len(pts)}점 — 등급 A {n['A']} · "
                   f"B 도엽대표 {n['B']} · 기존점 {n['기존점']}")
    note(ws, 2, "S",
         "선점 검토 결과 그대로다(export_selection_50 단일 출처). Site ID 는 이 야장 전체의 "
         "연결키이며 다른 시트는 이 값을 그대로 쓴다.")
    note(ws, 3, "S",
         "⚠ 「예측 |∇ΔT|」는 KIGAM 1.5분(약 2.8 km) 격자의 중앙차분값이다. 현장 수 m 스케일 "
         "측정과 «물리적으로 다른 양»이므로 합격기준으로 쓰지 않는다 — 사전 층화·탐색 변수일 뿐이다. "
         "자료공백 5점(오호·문산·강화·포천·화천)은 NOT AVAILABLE 로 두고 감점하거나 보간하지 말 것.",
         warn=True)

    hdr = ["Site ID", "연번", "구분", "원 번호", "지점명", "관할본부", "도엽번호", "도엽명",
           "위도(십진)", "경도(십진)", "표고(m)", "소재지",
           "예측 |∇ΔT|\n(nT/km)\n※판정 미사용", "예측 상태", "최근접\n관측소", "관측소\n거리(km)",
           "표지1 방위각\n(기재값)", "표지1\n거리(m)", "시험탐사 상태"]
    table(ws, 4, hdr,
          [11, 5, 7, 10, 16, 12, 13, 10, 11, 11, 9, 32, 12, 13, 9, 9, 14, 9, 34],
          keys={"Site ID"})
    for i, p in enumerate(pts, 1):
        r = 4 + i
        gap = p["예측구배"] is None
        hold = HOLD_SITES.get(p["지점명"])
        sid = f"S{i:02d}"
        p["_sid"] = sid
        vals = [sid, i, p["구분"], p["번호"], p["지점명"], p["관할본부"],
                p["도엽번호"], p["도엽명"], p["위도"], p["경도"], p["표고"],
                p["소재지"],
                None if gap else round(p["예측구배"], 1),
                "NOT AVAILABLE" if gap else "OK",
                p["최근접관측소"], round(p["관측소거리"]),
                p["표지1_방위각"] or "NOT APPLICABLE", p["표지1_거리"],
                f"HOLD — {hold}" if hold else "시험 탐사 대상"]
        fill = PatternFill("solid", fgColor=BAND[p["구분"]])
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border, c.fill = F_VAL, BORDER, fill
            c.alignment = AL_L if j in (12, 19) else AL_C
        for j, f in ((9, "0.000000"), (10, "0.000000"), (11, "0.0"),
                     (13, "0.0"), (18, "0.0")):
            ws.cell(row=r, column=j).number_format = f
        if gap:
            ws.cell(row=r, column=14).font = F_WARN
        if hold:
            ws.cell(row=r, column=19).font = F_WARN
            ws.cell(row=r, column=19).fill = FILL_WARN
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_site(wb, pts):
    """③ 점 정보 — 좌표 실측·검증, §13② 이격, 기존점 비교가능 플래그."""
    ws = wb.create_sheet("③ 점정보·좌표·이격")
    title(ws, "X", "점 정보 · 좌표 검증 · §13② 이격거리")
    note(ws, 2, "X",
         "좌표는 «십진도 숫자»로 위도·경도를 각각 적고 좌표계·취득방법·정확도·출처를 함께 남긴다. "
         "소수 자릿수만 맞추면 허위 정밀도가 생기고 위·경도 뒤바뀜도 못 막는다.")
    note(ws, 3, "X",
         "⚠ §13② 는 «도상 확인만으로 확정하지 않는다». 도상거리와 현장 확인거리를 각각 적고, "
         "판정은 ⑨ 게이트에서 필수 조건으로 쓴다. 시설 분류가 불명확하면 임의 판정하지 말고 PENDING.",
         warn=True)
    hdr = ["Site ID", "지점명", "구분",
           "사전 위도", "사전 경도",
           "현장 실측 위도", "현장 실측 경도", "좌표계", "측지기준", "취득방법",
           "수평정확도(m)", "좌표 출처", "좌표 검증상태",
           "직류철도(km)\n도상", "직류철도(km)\n현장", "교류·일반철도(km)\n도상",
           "교류·일반철도(km)\n현장", "고압철탑(km)\n도상", "고압철탑(km)\n현장",
           "송전탑(km)\n도상", "송전탑(km)\n현장", "최근접 시설 ID·유형",
           "자료 기준일", "§13② 판정"]
    table(ws, 4, hdr,
          [11, 15, 7, 11, 11, 12, 12, 12, 12, 13, 11, 16, 14,
           11, 11, 13, 13, 11, 11, 11, 11, 22, 11, 12],
          keys={"Site ID"})
    for i, p in enumerate(pts, 1):
        r = 4 + i
        pre = [p["_sid"], p["지점명"], p["구분"], p["위도"], p["경도"]]
        for j, v in enumerate(pre, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border, c.fill, c.alignment = F_VAL, BORDER, FILL_AUTO, AL_C
        for j, f in ((4, "0.000000"), (5, "0.000000")):
            ws.cell(row=r, column=j).number_format = f
        for j in range(6, len(hdr) + 1):
            c = ws.cell(row=r, column=j)
            c.border, c.fill, c.font, c.alignment = BORDER, FILL_FIELD, F_VAL, AL_C
            if j in (6, 7):
                c.number_format = "0.000000"
            elif 14 <= j <= 21:
                c.number_format = "0.00"
            elif j == 11:
                c.number_format = "0.00"
    last = 4 + len(pts)
    dv(ws, f"H5:H{last}", ["WGS84", "KGD2002", "Bessel1841", "PENDING"])
    dv(ws, f"J5:J{last}", ["RTK-GNSS", "네트워크RTK", "정적GNSS", "휴대GNSS",
                           "기존 성과 인용", "PENDING"])
    dv(ws, f"M5:M{last}", ["검증완료 — 사전값과 일치", "검증완료 — 사전값 정정",
                           "불일치 — HOLD", "PENDING"])
    dv(ws, f"X5:X{last}", ["적합", "부적합", "PENDING"])
    dv_warn(ws, f"F5:F{last}", 33, 39.5, "한반도 위도 범위(33~39.5°)를 벗어났다. "
                                         "위·경도를 바꿔 적지 않았는지 확인할 것.")
    dv_warn(ws, f"G5:G{last}", 124, 132, "한반도 경도 범위(124~132°)를 벗어났다. "
                                         "위·경도를 바꿔 적지 않았는지 확인할 것.")
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_existing(wb, pts):
    """④ 기존점 대조 가능성 — 성분별로 쪼갠다 (Codex C5)."""
    ws = wb.create_sheet("④ 기존점 대조가능성")
    ex = [p for p in pts if p["구분"] == "기존점"]
    title(ws, "N", f"기존 관측점 {len(ex)}점 — 과거 성과와의 «성분별» 대조 가능성")
    note(ws, 2, "N",
         "⚠ 「기존 성과가 있다」와 「직접 비교할 수 있다」는 다른 문제다. 과거 기록에는 마크 "
         "방위각 변동(재방문 잔여 RMS 33.7′)·F 측정시각 전무·방위 방향 혼동이 섞여 있다. "
         "따라서 적어도 편각 D 와 시간보정이 필요한 F 는 무조건 직접 비교할 수 없다.", warn=True)
    note(ws, 3, "N",
         "성분마다 «비교 가능 / 조건부 / 불가»를 따로 판정하고, 근거를 적는다. "
         "표석·표지 동일성이 확인되지 않으면 그 점 전체를 HOLD 로 둔다.")
    hdr = ["Site ID", "지점명", "지점코드", "종전 성과 연도",
           "종전 편각(°)", "종전 복각(°)", "종전 총자력(nT)", "종전 관측장비",
           "동일 표석 확인", "동일 표지 확인", "좌표 검증",
           "D 비교가능", "I 비교가능", "F 비교가능"]
    table(ws, 4, hdr, [11, 12, 11, 13, 12, 12, 13, 13, 13, 13, 13, 12, 12, 12],
          keys={"Site ID"})
    for i, p in enumerate(ex, 1):
        r = 4 + i
        s = p["종전성과"] or {}
        hold = HOLD_SITES.get(p["지점명"])
        pre = [p["_sid"], p["지점명"], p["번호"], s.get("최종관측", ""),
               s.get("편각"), s.get("복각"), s.get("총자력"), s.get("관측장비", "")]
        for j, v in enumerate(pre, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border, c.fill, c.alignment = F_VAL, BORDER, FILL_AUTO, AL_C
        for j, f in ((5, "0.0000"), (6, "0.0000"), (7, "#,##0.0")):
            ws.cell(row=r, column=j).number_format = f
        for j in range(9, len(hdr) + 1):
            c = ws.cell(row=r, column=j)
            c.border, c.font, c.alignment = BORDER, F_VAL, AL_C
            c.fill = FILL_WARN if hold else FILL_FIELD
            if hold:
                c.value = "HOLD"
    last = 4 + len(ex)
    dv(ws, f"I5:K{last}", ["확인", "불일치", "PENDING", "HOLD"])
    dv(ws, f"L5:N{last}", ["비교가능", "조건부", "불가", "PENDING", "HOLD"])
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_mark(wb, pts):
    """⑤ 방위표지·참방위각 (Codex C4)."""
    ws = wb.create_sheet("⑤ 방위표지·참방위각")
    title(ws, "T", "방위표지 · 참방위각 결정 — 편각 30′ 잔차의 주범")
    note(ws, 2, "T",
         "⚠ 규약: 진북 기준 · 북 = 0° · 시계방향 · FROM = 기준점 · TO = 방위표지. "
         "현장 원시각은 «gon» 으로 적는다(1회전 400 · 대척 200). 도(°) 환산은 계산 칸이다. "
         "200 gon 역방향과 180° 역방향이 섞이면 대규모 오류가 난다.", warn=True)
    note(ws, 3, "T",
         "재방문 16구간 전부에서 마크 참방위각이 60″ 넘게 변했다. 마크를 바꾸는 것 자체는 "
         "오류가 아니지만, 바뀐 마크의 참방위각이 틀리면 그 방문의 편각이 통째로 틀어진다. "
         "그래서 «이전 확정값 대비 차이»와 «재결정 사유»가 필수다.")
    hdr = ["Site ID", "Mark ID", "지점명", "표지 구분",
           "표지 위도", "표지 경도", "시준 지점(표지의 어디)",
           "FROM", "TO",
           "정 시준 원시각(gon)", "반 시준 원시각(gon)", "폐합차(gon)",
           "참방위각(gon)", "참방위각(°)\n계산", "역방위(°)\n계산",
           "결정방법", "근거좌표 정확도(m)",
           "이전 확정 방위각(°)", "이전값 대비 차(″)", "재결정 사유"]
    table(ws, 4, hdr,
          [11, 11, 14, 10, 12, 12, 20, 12, 12, 15, 15, 11, 13, 13, 13, 16, 14, 15, 14, 30],
          keys={"Site ID", "Mark ID"})
    r = 5
    for p in pts:
        for k in (1, 2):
            pre = [p["_sid"], f"{p['_sid']}-M{k}", p["지점명"], f"방위표지{k}",
                   p.get(f"표지{k}_위도"), p.get(f"표지{k}_경도")]
            for j, v in enumerate(pre, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.font, c.border, c.fill, c.alignment = F_VAL, BORDER, FILL_AUTO, AL_C
                if j in (5, 6) and v is not None:
                    c.number_format = "0.000000"
            ws.cell(row=r, column=8, value="기준점").fill = FILL_LOCK
            ws.cell(row=r, column=9, value="방위표지").fill = FILL_LOCK
            for j in (8, 9):
                c = ws.cell(row=r, column=j)
                c.font, c.border, c.alignment = F_VAL, BORDER, AL_C
            for j in (7, 10, 11, 13, 16, 17, 18, 20):
                c = ws.cell(row=r, column=j)
                c.border, c.fill, c.font, c.alignment = BORDER, FILL_FIELD, F_VAL, AL_C
                if j in (10, 11, 13):
                    c.number_format = "0.0000"
                elif j in (17, 18):
                    c.number_format = "0.00"
            # 계산 칸 — 폐합차·도 환산·역방위·이전값 차
            ws.cell(row=r, column=12, value=f"=IF(COUNT(J{r}:K{r})=2,"
                                            f"ABS(ABS(K{r}-J{r})-200),\"\")")
            ws.cell(row=r, column=14, value=f"=IF(N(M{r})=0,\"\",M{r}*0.9)")
            ws.cell(row=r, column=15, value=f"=IF(N(M{r})=0,\"\",MOD(M{r}*0.9+180,360))")
            ws.cell(row=r, column=19, value=f"=IF(OR(N(M{r})=0,N(R{r})=0),\"\","
                                            f"(M{r}*0.9-R{r})*3600)")
            for j in (12, 14, 15, 19):
                c = ws.cell(row=r, column=j)
                c.border, c.fill, c.font, c.alignment = BORDER, FILL_AUTO, F_VAL, AL_C
                c.number_format = "0.0000" if j in (12, 14, 15) else "0.0"
            r += 1
    last = r - 1
    dv(ws, f"P5:P{last}", ["천문관측", "자이로", "RTK 장기선", "좌표계산", "기존 성과 인용", "PENDING"])
    dv_warn(ws, f"L5:L{last}", 0, 0.02, "마크 폐합차가 0.02 gon(약 65″)을 넘는다. 시준을 재확인할 것.")
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_base(wb):
    """⑥ 기준자력계 — C1 의 답. 이게 없으면 시공간 분리가 성립하지 않는다."""
    ws = wb.create_sheet("⑥ 기준자력계")
    title(ws, "M", "현지 기준자력계(연속 기록) — 일변화와 공간변화를 나누는 유일한 수단")
    note(ws, 2, "M",
         "⚠ 같은 측선을 시간을 두고 두 번 재는 것만으로는 «공간장»과 «시간변화»를 나눌 수 없다. "
         "같은 순서로 돌면 위치와 시각이 함께 움직이기 때문이다. 상시관측소는 중앙 101 km · "
         "최대 229 km 떨어져 있고, 그 거리에서 «외부장이 전역 균일»하다는 근사가 깨짐을 이 "
         "프로젝트가 이미 실증했다.", warn=True)
    note(ws, 3, "M",
         "→ 현지 고정점에 기준자력계를 두고 «초 단위로 동기화»해 연속 기록한다. "
         "기준자력계를 운용하지 못한 점은 ⑩ 판정에서 「시간변화 미분리」로 표시하고 "
         "**선점 확정 근거로 쓰지 않는다**. (투입 여부는 예산·인력 판단 — ⑫ 참조)")
    hdr = ["Site ID", "Visit ID", "기준자력계 위치 ID",
           "위도", "경도", "센서 높이(m)", "기기 제조사·모델", "기기번호",
           "기록 간격(s)", "기록 시작(KST)", "기록 종료(KST)",
           "이동 자력계와 시계 동기 방법", "기록파일명"]
    table(ws, 4, hdr, [11, 10, 16, 11, 11, 11, 18, 12, 11, 19, 19, 24, 26],
          keys={"Site ID", "Visit ID"})
    blank_rows(ws, 5, 60, len(hdr),
               fmts={4: "0.000000", 5: "0.000000", 6: "0.00", 9: "0",
                     10: DTFMT, 11: DTFMT})
    dv(ws, "L5:L64", ["GNSS 시각 동기", "NTP 동기", "수동 대조(오차 기록)", "미동기 — INVALID"])
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_profile(wb):
    """⑦ 자기환경 측선 관측행 (Codex M2·M3·M6·m1)."""
    ws = wb.create_sheet("⑦ 자기환경 측선")
    title(ws, "R", "자기환경 시험 — 측점 «관측행»")
    note(ws, 2, "R",
         "측선 설계는 **파일럿 먼저**다. 소수 지점에서 조밀하게 재어 필요한 간격·범위를 정한 뒤 "
         "전 지점에 최소 공통 설계를 적용하고, 이상 징후가 있는 곳만 격자를 추가한다 — "
         "처음부터 전 지점 516 판독을 강제하면 일정에 눌려 좌표·시각 기록이 먼저 생략된다.")
    note(ws, 3, "R",
         "⚠ 순회 순서는 «중심점 재관측(A–B–A)»과 «역순 재측선»을 포함해야 한다. 같은 순서로 "
         "두 번 도는 것은 시간변화를 분리하지 못한다. 판독은 «3회를 각각» 남기고 중앙값은 "
         "계산 칸이다 — 중앙값만 남기면 오기·순간이상·제외값을 재검증할 수 없다.", warn=True)
    hdr = ["Site ID", "Visit ID", "Point ID", "측선", "순회 회차", "순회 순서(정/역)",
           "E 오프셋(m)", "N 오프셋(m)",
           "실제 센서 높이(m)", "높이 기준", "기기번호",
           "관측 시각(KST)", "UTC 환산",
           "F 판독1(nT)", "F 판독2(nT)", "F 판독3(nT)", "중앙값(nT)", "상태·비고"]
    table(ws, 4, hdr,
          [11, 10, 12, 12, 10, 13, 11, 11, 13, 12, 11, 19, 19, 12, 12, 12, 12, 26],
          keys={"Site ID", "Visit ID", "Point ID"})
    N = 400
    blank_rows(ws, 5, N, len(hdr),
               fmts={7: "0.00", 8: "0.00", 9: "0.00", 12: DTFMT,
                     14: "0.0", 15: "0.0", 16: "0.0"})
    for r in range(5, 5 + N):
        c = ws.cell(row=r, column=13, value=f'=IF(L{r}="","",L{r}-TIME(9,0,0))')
        c.number_format, c.fill = DTFMT, FILL_AUTO
        c.border, c.font, c.alignment = BORDER, F_VAL, AL_C
        c2 = ws.cell(row=r, column=17,
                     value=f'=IF(COUNT(N{r}:P{r})=0,"",MEDIAN(N{r}:P{r}))')
        c2.number_format, c2.fill = "0.0", FILL_AUTO
        c2.border, c2.font, c2.alignment = BORDER, F_VAL, AL_C
    dv(ws, f"F5:F{4+N}", ["정", "역"])
    dv(ws, f"J5:J{4+N}", ["지면", "표석 상면", "PENDING"])
    dv_warn(ws, f"N5:P{4+N}", 30000, 60000,
            "한반도 총자력 범위(30,000~60,000 nT) 밖이다. 실제 이상일 수 있으니 "
            "지우지 말고 「상태·비고」에 사유를 적을 것.")
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_session(wb):
    """⑧ 절대측정 관측행 — Codex C2 의 핵심."""
    ws = wb.create_sheet("⑧ 절대측정 관측행")
    title(ws, "V", "절대측정 — 세션 아래 «관측행». 이 시트가 이 야장의 중심이다")
    note(ws, 2, "V",
         "연결키: Site — Visit — Date — Session — Seq — Component — Location — Instrument. "
         "이 여덟 중 하나만 비어도 D·I·F 를 다시 묶을 수 없다. 과거에 F 시각이 사라진 것이 "
         "정확히 이 결합의 실패였다.")
    note(ws, 3, "V",
         "⚠ F 는 «판독마다» 초 단위 시각을 적는다(§19①6 총자기장·시간 동시 측정). "
         "D·I 는 시작·종료 시각을 각각 적는다(§20 20분 이내 판정 근거). "
         "원시 판독값을 남기고 최종값만 적지 않는다.", warn=True)
    hdr = ["Site ID", "Visit ID", "관측일자", "Session", "Seq",
           "Component", "Location ID", "Instrument",
           "시작 시각(KST)", "종료 시각(KST)", "소요(분)\n계산", "UTC 시작\n계산",
           "원시 판독(gon)", "원시 판독 단위", "산출값",
           "D(°)", "I(°)", "F(nT)",
           "센서 높이(m)", "관측자", "상태", "비고"]
    table(ws, 4, hdr,
          [11, 10, 12, 9, 6, 12, 13, 12, 19, 19, 10, 19, 14, 13, 12, 11, 11, 12,
           11, 11, 13, 28],
          keys={"Site ID", "Visit ID", "관측일자", "Session", "Seq",
                "Component", "Location ID", "Instrument"})
    N = 500
    blank_rows(ws, 5, N, len(hdr),
               fmts={3: "yyyy-mm-dd", 9: DTFMT, 10: DTFMT, 13: "0.0000",
                     16: "0.0000", 17: "0.0000", 18: "0.0", 19: "0.00"})
    for r in range(5, 5 + N):
        c = ws.cell(row=r, column=11,
                    value=f'=IF(OR(I{r}="",J{r}=""),"",(J{r}-I{r})*1440)')
        c.number_format = "0.0"
        c2 = ws.cell(row=r, column=12, value=f'=IF(I{r}="","",I{r}-TIME(9,0,0))')
        c2.number_format = DTFMT
        for cc in (c, c2):
            cc.fill, cc.border, cc.font, cc.alignment = FILL_AUTO, BORDER, F_VAL, AL_C
        u = ws.cell(row=r, column=14, value="gon")
        u.fill, u.border, u.font, u.alignment = FILL_LOCK, BORDER, F_VAL, AL_C
    dv(ws, f"F5:F{4+N}", ["편각 D", "복각 I", "총자력 F"])
    dv(ws, f"G5:G{4+N}", ["DI점", "원격점", "기준자력계"])
    dv(ws, f"U5:U{4+N}", ["정상"] + STATUS)
    dv_warn(ws, f"P5:P{4+N}", -20, 0, "한국은 서편각이라 D 가 음수여야 한다. 부호를 확인할 것 "
                                      "(지우지 말고 사유를 비고에).")
    dv_warn(ws, f"Q5:Q{4+N}", 30, 70, "복각 I 예상 범위(30~70°) 밖이다. 셀이 날짜 서식으로 "
                                      "바뀌지 않았는지 확인할 것.")
    dv_warn(ws, f"R5:R{4+N}", 30000, 60000, "총자력 예상 범위 밖이다.")
    dv_warn(ws, f"K5:K{4+N}", 0, 20, "§20 관측시간 20분을 넘겼다.")
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_dualF(wb):
    """⑨ 두 지점 F — A–B–A (Codex C2·M5·M6)."""
    ws = wb.create_sheet("⑨ 두지점 F·기기대조")
    title(ws, "Q", "두 지점 F (센서 위치차) · 기기 상호대조 — A–B–A 로 시간변화를 뺀다")
    note(ws, 2, "Q",
         "⚠ 두 지점을 한 번씩 순차로 재면 «위치차 + 그 사이 시간변화»가 섞인다. "
         "A(DI점) → B(원격점) → A(DI점) 순서로 재고, A 두 값의 선형내삽으로 B 시각의 A 값을 "
         "추정해 ΔF 를 낸다. 두 기기를 «서로 바꿔» 한 번 더 재면 기기 바이어스도 분리된다.",
         warn=True)
    note(ws, 3, "Q",
         "과거 야장의 「기본값」은 두 번째 지점 측정값이 아니라 «측정값의 정수부»였다 — "
         "즉 센서 위치차 보정에 쓸 두 지점 F 가 애초에 존재하지 않았다(68개 전수 확인).")
    hdr = ["Site ID", "Visit ID", "Set ID", "순서", "Location ID",
           "실제 거리 A↔B(m)", "방위(기준점→원격점, °)",
           "기기번호", "센서 높이(m)", "관측 시각(KST)",
           "F 판독1", "F 판독2", "F 판독3", "중앙값\n계산",
           "교차배치(swap) 여부", "ΔF(위치차, nT)", "비고"]
    table(ws, 4, hdr,
          [11, 10, 10, 10, 13, 14, 18, 11, 11, 19, 11, 11, 11, 12, 16, 14, 28],
          keys={"Site ID", "Visit ID", "Set ID"})
    N = 200
    blank_rows(ws, 5, N, len(hdr),
               fmts={6: "0.00", 7: "0.0", 9: "0.00", 10: DTFMT,
                     11: "0.0", 12: "0.0", 13: "0.0", 16: "0.0"})
    for r in range(5, 5 + N):
        c = ws.cell(row=r, column=14,
                    value=f'=IF(COUNT(K{r}:M{r})=0,"",MEDIAN(K{r}:M{r}))')
        c.number_format, c.fill = "0.0", FILL_AUTO
        c.border, c.font, c.alignment = BORDER, F_VAL, AL_C
    dv(ws, f"D5:D{4+N}", ["A1 (DI점)", "B (원격점)", "A2 (DI점 재측)"])
    dv(ws, f"E5:E{4+N}", ["DI점", "원격점"])
    dv(ws, f"O5:O{4+N}", ["원배치", "교차배치(swap)"])
    dv_warn(ws, f"K5:M{4+N}", 30000, 60000, "총자력 예상 범위 밖이다.")
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_gate(wb, pts):
    """⑩ 법정·기록 게이트 — 임계와 분리한다 (Codex C3)."""
    ws = wb.create_sheet("⑩ 게이트 점검")
    title(ws, "N", "법정 요건 · 기록 완전성 게이트 — 통과/실패로만 가른다")
    note(ws, 2, "N",
         "⚠ 근거 없는 정량 임계로 선점을 확정하지 않기 위해, 판정을 두 층으로 나눈다. "
         "이 시트는 «반드시 통과해야 하는 것»만 다룬다 — 법정 조문과 기록 완전성이다. "
         "정량 지표는 ⑪에서 «참고»로만 본다.", warn=True)
    hdr = ["Site ID", "지점명",
           "§13② 이격", "§17 0.1 nT 판독", "§19① 1일 6세션 이상",
           "§19①6 F·시각 동시", "§20 정수차 30′", "§20 20분 이내",
           "§21 상시 기준점 대조",
           "기준자력계 연속기록", "두 지점 F(A–B–A)", "참방위각 결정",
           "좌표 검증", "게이트 종합"]
    table(ws, 4, hdr, [11, 15, 11, 13, 15, 14, 12, 12, 14, 14, 14, 12, 11, 13],
          keys={"Site ID"})
    for i, p in enumerate(pts, 1):
        r = 4 + i
        for j, v in enumerate([p["_sid"], p["지점명"]], 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border, c.fill, c.alignment = F_VAL, BORDER, FILL_AUTO, AL_C
        for j in range(3, len(hdr)):
            c = ws.cell(row=r, column=j)
            c.border, c.fill, c.font, c.alignment = BORDER, FILL_FIELD, F_VAL, AL_C
        c = ws.cell(row=r, column=len(hdr),
                    value=f'=IF(COUNTIF(C{r}:L{r},"실패")>0,"실패",'
                          f'IF(COUNTIF(C{r}:L{r},"통과")=10,"통과","미완"))')
        c.border, c.fill, c.font, c.alignment = BORDER, FILL_AUTO, F_VAL, AL_C
    last = 4 + len(pts)
    dv(ws, f"C5:L{last}", ["통과", "실패", "PENDING", "NOT APPLICABLE"])
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_result(wb, pts):
    """⑪ 정량 지표 — 임계는 미확정 (Codex C3·M1)."""
    ws = wb.create_sheet("⑪ 지표·판정")
    title(ws, "M", "정량 지표 · 시험 탐사 판정")
    note(ws, 2, "M",
         "⚠ 아래 지표의 «임계값은 아직 정해지지 않았다». 파일럿에서 반복일·기기·관측자 분산을 "
         "추정하고, 후속 D·I·F 허용오차에서 역산한 오차예산과 결합해 정한 뒤 «50점 판정 전에 "
         "동결»한다. 그전에는 이 칸의 수치로 선점을 확정·부적합 처리하지 않는다.", warn=True)
    note(ws, 3, "M",
         "「예측 |∇ΔT|」와 「현장 국소구배」는 공간 지지범위가 달라 직접 비교 대상이 아니다 — "
         "일치·불일치를 자료 오류나 부적합으로 읽지 말 것. 탐색적 연관성 분석에만 쓴다.")
    hdr = ["Site ID", "지점명", "구분",
           "현장 국소구배 중앙(nT/m)", "측정범위 max−min(nT)", "반복 재현성 RMS(nT)",
           "기기 상호대조차(nT)", "세션 정수차(′)",
           "시간변화 분리 여부", "예측 |∇ΔT|(참고)",
           "게이트 종합", "시험 탐사 판정", "판정 근거·특이사항"]
    table(ws, 4, hdr, [11, 15, 7, 17, 16, 16, 15, 13, 15, 14, 12, 14, 40],
          keys={"Site ID"})
    for i, p in enumerate(pts, 1):
        r = 4 + i
        gap = p["예측구배"] is None
        for j, v in enumerate([p["_sid"], p["지점명"], p["구분"]], 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border, c.fill, c.alignment = F_VAL, BORDER, FILL_AUTO, AL_C
        for j in range(4, 10):
            c = ws.cell(row=r, column=j)
            c.border, c.fill, c.font, c.alignment = BORDER, FILL_FIELD, F_VAL, AL_C
            c.number_format = "0.00" if j in (4,) else "0.0"
        c = ws.cell(row=r, column=10,
                    value="NOT AVAILABLE" if gap else round(p["예측구배"], 1))
        c.border, c.fill, c.font, c.alignment = BORDER, FILL_AUTO, F_VAL, AL_C
        if gap:
            c.font = F_WARN
        c = ws.cell(row=r, column=11, value=f"='⑩ 게이트 점검'!N{r}")
        c.border, c.fill, c.font, c.alignment = BORDER, FILL_AUTO, F_VAL, AL_C
        for j in (12, 13):
            c = ws.cell(row=r, column=j)
            c.border, c.fill, c.font, c.alignment = BORDER, FILL_FIELD, F_VAL, AL_C
            c.alignment = AL_L if j == 13 else AL_C
    last = 4 + len(pts)
    dv(ws, f"I5:I{last}", ["분리됨 — 기준자력계 연속기록", "미분리 — 확정 근거 불가"])
    dv(ws, f"L5:L{last}", ["선점 확정", "조건부", "부적합", "재측정", "HOLD", "PENDING"])
    return ws


# ══════════════════════════════════════════════════════════════
def sheet_human(wb):
    """⑫ Human 확인 필요 — Codex 가 지목한 6건."""
    ws = wb.create_sheet("⑫ 확인 필요")
    for col, w in zip("ABCDE", [6, 34, 46, 20, 22]):
        ws.column_dimensions[col].width = w
    title(ws, "E", "Human 확인 필요 — 이 야장이 답하지 않는 것")
    note(ws, 2, "E",
         "Codex 독립 검토(2026-09-07)가 「기관 확인 없이는 확정할 수 없다」고 지목한 항목이다. "
         "야장 설계로는 해소되지 않으므로 착수 전에 결정해야 한다.")
    table(ws, 3, ["#", "항목", "왜 확인해야 하는가", "결정 주체", "상태"],
          [6, 34, 46, 20, 22])
    rows = [
        ("§19·§20 의 «세션» 및 20분 요건 공식 적용단위",
         "1일 6회의 «회»가 무엇인지, 20분이 어느 구간인지에 따라 야장 관측행 설계와 "
         "적합 판정이 달라진다. 법적·제도적 해석 사항.", "국토지리정보원"),
        ("기존 16점의 표석·표지 동일성과 공식 좌표",
         "과거 성과와 신규 관측을 견주려면 같은 물리점임이 확인돼야 한다. 미원은 좌표가 "
         "353 m 어긋나 HOLD 로 두었다.", "국토지리정보원"),
        ("현지 기준자력계 투입 여부",
         "이것 없이는 일변화와 공간변화를 분리할 수 없어 시험 결과를 선점 확정 근거로 "
         "쓸 수 없다. 추가 기기·인력·관측시간이 든다.", "발주기관 · 예산"),
        ("정량 임계값과 허용 오판위험 승인",
         "임계를 어디에 두느냐가 «부적합지를 통과시키는 위험»과 «양호지를 탈락시키는 위험» "
         "의 교환이다. 가치판단이 필요하다.", "발주기관"),
        ("§13② 시설 분류와 거리 산정 방법",
         "시설 유형 분류가 애매하면 법정 이격 판정이 갈린다. 대외 법정 적합성에 직결.",
         "국토지리정보원"),
        ("상시관측소 원자료의 제공조건·시간해상도",
         "자료가 «존재한다»와 «쓸 수 있다»는 다르다. §21 대조의 실현 가능성이 여기 달렸다.",
         "청양(INTERMAGNET) · KASA"),
    ]
    for i, (item, why, who) in enumerate(rows, 1):
        r = 3 + i
        for j, v in enumerate([i, item, why, who, ""], 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border = F_VAL, BORDER
            c.alignment = AL_L if j in (2, 3) else AL_C
            c.fill = FILL_FIELD if j == 5 else FILL_AUTO
        ws.row_dimensions[r].height = 44
    dv(ws, f"E4:E{3+len(rows)}", ["확인 완료", "질의 중", "PENDING"])
    return ws


# ══════════════════════════════════════════════════════════════
def main():
    sys.stdout.reconfigure(encoding="utf-8")
    pts = TP.load_points()
    print(f"대상 {len(pts)}점 로드")
    wb = Workbook()
    wb.remove(wb.active)
    sheet_intro(wb)
    sheet_points(wb, pts)
    sheet_site(wb, pts)
    sheet_existing(wb, pts)
    sheet_mark(wb, pts)
    sheet_base(wb)
    sheet_profile(wb)
    sheet_session(wb)
    sheet_dualF(wb)
    sheet_gate(wb, pts)
    sheet_result(wb, pts)
    sheet_human(wb)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / f"{dt.datetime.now():%Y%m%d_%H%M%S}_시험탐사_표준야장.xlsx"
    wb.save(path)
    print(f"시트 {len(wb.sheetnames)}장: {' · '.join(wb.sheetnames)}")
    print(f"[저장] {path}")
    return path


if __name__ == "__main__":
    main()
