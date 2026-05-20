#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
후보지 #16·#18 서브격자 Excel 지도링크 수정 재생성
 - Naver: map.naver.com/v5/search/{lat},{lon}  (핀 표시)
 - Kakao: map.kakao.com/link/map/{title},{lat},{lon}  (# 제거)
 - =HYPERLINK() 수식 방식으로 변경 (cell.hyperlink 대체)
"""
import sys, json, warnings
import numpy as np
import pandas as pd
import geopandas as gpd
import shapely
from shapely.geometry import Point, LineString, box
from shapely.prepared import prep
from pathlib import Path
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
warnings.filterwarnings("ignore")

MAIN_DIR  = Path("C:/LG_gram_backup_users/LX/2026_geomag")
DATA_DIR  = MAIN_DIR / "data"
CACHE_DIR = DATA_DIR / "zone_cache"
OUT_DIR   = MAIN_DIR / "docs" / "output"
UTM_CRS   = "EPSG:5179"
WGS84_CRS = "EPSG:4326"
TS        = datetime.now().strftime("%Y%m%d_%H%M%S")

RADIUS_M   = 5_000
GRID_STEP  = 1_000
ROAD_MAX_M = 2_000
TOP_N      = 10

SITES = [
    {"id": 16, "name": "후보지 #16 (전주 완주)",
     "lat": 35.59196, "lon": 127.15795, "score_national": 60.1, "priority": 1},
    {"id": 18, "name": "후보지 #18 (진안 무주)",
     "lat": 35.86272, "lon": 127.26755, "score_national": 58.8, "priority": 1},
]
ZONE_NAMES = ["power","railway","urban_dense","urban_resid",
              "pipeline","comm","wind","quarry","water","geology","fault","anomaly"]

# ── URL 생성 함수 ──────────────────────────────────────────────
def naver_url(lat, lon):
    """네이버 지도 — 좌표 검색 (핀 표시)"""
    return f"https://map.naver.com/v5/search/{lat},{lon}"

def kakao_url(lat, lon, rank, sid):
    """카카오 지도 — map link (# 없는 순수 레이블)"""
    label = f"P{sid}-{rank}"          # # 제거, 특수문자 없음
    return f"https://map.kakao.com/link/map/{label},{lat},{lon}"

# ── openpyxl 스타일 헬퍼 ──────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

C_HEADER = "1F4E79"
C_SITE16 = "2F75B6"
C_SITE18 = "375623"
C_TOP3   = "FFE699"
C_TOP5   = "EBF3FB"
C_WHITE  = "FFFFFF"
C_LABEL  = "D6DCE4"


# ════════════════════════════════════════════════════════════════
# 데이터 재구성 (캐시 활용)
# ════════════════════════════════════════════════════════════════
def load_zone(name, local_box):
    p = CACHE_DIR / f"{name}.wkb"
    if not p.exists():
        return None
    g = shapely.from_wkb(p.read_bytes())
    try:
        local = g.intersection(local_box)
        return None if local.is_empty else local
    except Exception:
        return g

def zone_dist(pts_list, zone_geom, simplify_m=300):
    if zone_geom is None or zone_geom.is_empty:
        return np.full(len(pts_list), 50_000.0)
    zs = zone_geom.simplify(simplify_m)
    return np.array([pt.distance(zs) for pt in pts_list])

def kigam_score(lat, lon, klat, klon, kanom, radius=0.05):
    mask = (np.abs(klat - lat) <= radius) & (np.abs(klon - lon) <= radius)
    pts_k = kanom[mask]
    if len(pts_k) < 3:
        return 6.0, np.nan
    v = float(np.percentile(pts_k, 90) - np.percentile(pts_k, 10))
    if   v <  30:  return 5.0, round(v,1)
    elif v <= 150: return 8.0, round(v,1)
    elif v <= 400: return 10.0, round(v,1)
    elif v <= 800: return 7.0, round(v,1)
    else:          return 0.0, round(v,1)

# KIGAM
kigam_path = DATA_DIR / "mag_1982-2018_1.5min_ed.dat"
klat_arr = klon_arr = kanom_arr = None
if kigam_path.exists():
    kdf = pd.read_csv(kigam_path, sep=r"\s+", skiprows=9, header=None,
                      names=["lon","lat","anomaly_nT"],
                      na_values=["99999","-99999"], on_bad_lines="skip")
    kdf = kdf.apply(pd.to_numeric, errors="coerce").dropna()
    klat_arr = kdf.lat.values; klon_arr = kdf.lon.values; kanom_arr = kdf.anomaly_nT.values

# 도로
road_json = DATA_DIR / "access_roads.json"
road_gdf_wgs = None
if road_json.exists():
    with open(road_json, encoding="utf-8") as f:
        rd = json.load(f)
    nodes_rd = {e["id"]: (e["lon"],e["lat"]) for e in rd.get("elements",[]) if e["type"]=="node"}
    road_lines = []
    for e in rd.get("elements",[]):
        if e["type"] == "way":
            coords = [nodes_rd[n] for n in e.get("nodes",[]) if n in nodes_rd]
            if len(coords) >= 2:
                road_lines.append(LineString(coords))
    if road_lines:
        road_gdf_wgs = gpd.GeoDataFrame(geometry=road_lines, crs=WGS84_CRS)

all_results = []
for site in SITES:
    sid = site["id"]; clat = site["lat"]; clon = site["lon"]
    print(f"\n재분석: 후보지 #{sid}...")

    center_gdf = gpd.GeoDataFrame(geometry=[Point(clon,clat)], crs=WGS84_CRS).to_crs(UTM_CRS)
    cx = center_gdf.geometry.iloc[0].x; cy = center_gdf.geometry.iloc[0].y
    xs = np.arange(cx-RADIUS_M, cx+RADIUS_M+1, GRID_STEP)
    ys = np.arange(cy-RADIUS_M, cy+RADIUS_M+1, GRID_STEP)
    center_pt = Point(cx, cy)
    pts_utm = [Point(x,y) for x in xs for y in ys if Point(x,y).distance(center_pt)<=RADIUS_M]
    n_all = len(pts_utm)
    LOCAL_BOX = box(cx-RADIUS_M*1.2, cy-RADIUS_M*1.2, cx+RADIUS_M*1.2, cy+RADIUS_M*1.2)

    gdf_all = gpd.GeoDataFrame({"gid":range(n_all)}, geometry=pts_utm, crs=UTM_CRS)
    gdf_wgs = gdf_all.to_crs(WGS84_CRS)
    gdf_all["lat"] = gdf_wgs.geometry.y.round(6)
    gdf_all["lon"] = gdf_wgs.geometry.x.round(6)

    # 경사도 캐시
    ELEV_CACHE = OUT_DIR / f"site{sid}_elev_cache.json"
    slopes = np.full(n_all, np.nan)
    if ELEV_CACHE.exists():
        with open(ELEV_CACHE) as f:
            elev_dict = {int(k):v for k,v in json.load(f).items()}
        ELEV_OFFSET = 500
        for gi in range(n_all):
            t = elev_dict.get(gi, {})
            ec = t.get("C")
            grads = []
            for dn in ["N","S","E","W"]:
                ev = t.get(dn)
                if ec is not None and ev is not None:
                    grads.append(np.degrees(np.arctan(abs(ev-ec)/ELEV_OFFSET)))
            if grads:
                slopes[gi] = np.mean(grads)
    gdf_all["slope_deg"] = np.round(slopes, 2)

    # 제외구역
    zones = {nm: load_zone(nm, LOCAL_BOX) for nm in ZONE_NAMES}
    excl_mask = np.zeros(n_all, dtype=bool)
    excl_reason = [""] * n_all
    for nm in ZONE_NAMES:
        z = zones[nm]
        if z is None or z.is_empty: continue
        pz = prep(z)
        for i, pt in enumerate(pts_utm):
            if not excl_mask[i] and pz.contains(pt):
                excl_mask[i]=True; excl_reason[i]=nm
    gdf_all["excl"] = excl_mask; gdf_all["excl_r"] = excl_reason
    candidates = gdf_all[~excl_mask].copy().reset_index(drop=True)

    # 도로 접근성
    road_dist = np.full(len(candidates), np.nan)
    if road_gdf_wgs is not None:
        road_utm = road_gdf_wgs.to_crs(UTM_CRS)
        local_roads = road_utm[road_utm.geometry.intersects(LOCAL_BOX)]
        if len(local_roads) > 0:
            road_union = local_roads.geometry.unary_union
            pts_c = candidates.geometry.tolist()
            road_dist = np.array([pt.distance(road_union) for pt in pts_c])
    candidates["road_dist_m"] = np.round(road_dist, 0)
    if not np.all(np.isnan(road_dist)):
        candidates = candidates[
            candidates["road_dist_m"].isna() | (candidates["road_dist_m"]<=ROAD_MAX_M)
        ].copy().reset_index(drop=True)

    # 점수
    nc = len(candidates); pts_c = candidates.geometry.tolist()
    sl = candidates["slope_deg"].values.astype(float)
    sl_fill = np.where(np.isnan(sl), np.nanmedian(sl) if not np.all(np.isnan(sl)) else 5.0, sl)
    s2 = np.clip((1.0-sl_fill/30.0)*15.0, 0.0, 15.0)
    d_power   = zone_dist(pts_c, zones["power"])
    d_railway = zone_dist(pts_c, zones["railway"])
    lp = np.log1p(d_power); lr3 = np.log1p(d_railway)
    s3 = (lp/(lp.max() or 1)*0.5 + lr3/(lr3.max() or 1)*0.5)*15
    d_urban = zone_dist(pts_c, zones["urban_resid"])
    lu = np.log1p(d_urban); s4 = (lu/(lu.max() or 1))*15
    s5_arr = np.full(nc, 6.0); p90_arr = np.full(nc, np.nan)
    cands_wgs = candidates.to_crs(WGS84_CRS)
    if klat_arr is not None:
        for i in range(nc):
            glat = cands_wgs.geometry.iloc[i].y; glon = cands_wgs.geometry.iloc[i].x
            sc5, pp = kigam_score(glat, glon, klat_arr, klon_arr, kanom_arr)
            s5_arr[i]=sc5; p90_arr[i]=pp
    gz = zones.get("geology"); fz = zones.get("fault")
    have_rock  = gz is not None and not gz.is_empty
    have_fault = fz is not None and not fz.is_empty
    lr_rock = np.ones(nc); lr_fault = np.ones(nc)
    d_rock_arr = np.full(nc, np.nan); d_fault_arr = np.full(nc, np.nan)
    if have_rock:
        d_rock = zone_dist(pts_c, gz.simplify(300))
        d_rock_arr = d_rock; lr_rock = np.log1p(d_rock)
        candidates["d_geo_rock_km"] = np.round(d_rock/1000, 1)
    else:
        candidates["d_geo_rock_km"] = np.nan
    if have_fault:
        d_fault = zone_dist(pts_c, fz.simplify(300))
        d_fault_arr = d_fault; lr_fault = np.log1p(d_fault)
        candidates["d_geo_fault_km"] = np.round(d_fault/1000, 1)
    else:
        candidates["d_geo_fault_km"] = np.nan
    if have_rock and have_fault:
        s6_arr = (lr_rock/(lr_rock.max() or 1)*0.5 + lr_fault/(lr_fault.max() or 1)*0.5)*5
    elif have_rock:
        s6_arr = (lr_rock/(lr_rock.max() or 1))*5
    elif have_fault:
        s6_arr = (lr_fault/(lr_fault.max() or 1))*5
    else:
        s6_arr = np.full(nc, 2.5)
    candidates["s2_지형"]=np.round(s2,1); candidates["s3_전력철도"]=np.round(s3,1)
    candidates["s4_인구이격"]=np.round(s4,1); candidates["s5_모델기여도"]=np.round(s5_arr,1)
    candidates["s6_암상"]=np.round(s6_arr,1); candidates["mag_p90p10_nT"]=np.round(p90_arr,1)
    candidates["d_power_km"]=np.round(d_power/1000,1); candidates["d_railway_km"]=np.round(d_railway/1000,1)
    candidates["d_urban_km"]=np.round(d_urban/1000,1)
    raw = s2+s3+s4+s5_arr+s6_arr
    candidates["score_raw"]=np.round(raw,1); candidates["score"]=np.round(raw/60*100,1)
    candidates=candidates.sort_values("score",ascending=False).reset_index(drop=True)
    candidates["rank"]=candidates.index+1
    top = candidates.head(TOP_N).copy()
    top_wgs = top.to_crs(WGS84_CRS)
    top["lat"] = top_wgs.geometry.y.round(6)
    top["lon"] = top_wgs.geometry.x.round(6)
    all_results.append({"site":site,"top":top,"clat":clat,"clon":clon})
    print(f"  → 상위 {min(TOP_N,len(top))}개 준비 완료")


# ════════════════════════════════════════════════════════════════
# Excel 생성 — =HYPERLINK() 수식 방식
# ════════════════════════════════════════════════════════════════
print("\nExcel 생성 중...")

def hl_cell(ws, row, col, url, label, bg_hex, bold=False):
    """=HYPERLINK() 수식으로 클릭 가능 링크 삽입"""
    c = ws.cell(row=row, column=col)
    # 수식: =HYPERLINK("url","label")
    c.value = f'=HYPERLINK("{url}","{label}")'
    c.font  = Font(name="맑은 고딕", size=10, color="0563C1",
                   underline="single", bold=bold)
    c.fill  = hdr_fill(bg_hex)
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

def make_sheet(wb, res, sheet_name, hdr_color):
    ws = wb.create_sheet(sheet_name)
    site = res["site"]; top = res["top"]
    sid = site["id"]; sname = site["name"]
    clat = res["clat"]; clon = res["clon"]

    # ── 타이틀 ──
    ws.merge_cells("A1:N1")
    ws["A1"] = f"측정 후보지 #{sid} — 1km 서브격자 현장 답사 세부 선정"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hdr_fill(hdr_color)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("사이트명", sname),
        ("10km격자 중심", f"{clat}°N, {clon}°E"),
        ("국가격자 점수", f"{site['score_national']}점 (우선순위 {site['priority']}등급)"),
        ("분석 조건", f"반경 {RADIUS_M//1000}km / 격자 {GRID_STEP//1000}km / 도로≤{ROAD_MAX_M//1000}km"),
        ("생성 일시", datetime.now().strftime("%Y년 %m월 %d일 %H:%M")),
    ]
    for r_off, (k, v) in enumerate(meta, start=2):
        ws[f"A{r_off}"] = k
        ws[f"A{r_off}"].font = Font(bold=True, name="맑은 고딕", size=10)
        ws[f"A{r_off}"].fill = hdr_fill(C_LABEL)
        ws[f"A{r_off}"].alignment = Alignment(horizontal="right")
        ws.merge_cells(f"B{r_off}:N{r_off}")
        ws[f"B{r_off}"] = v
        ws[f"B{r_off}"].font = Font(name="맑은 고딕", size=10)
    ROW_H = len(meta) + 3   # = 8

    # ── 컬럼 헤더 ──
    headers = [
        ("순위", 5), ("위도(°N)", 12), ("경도(°E)", 12),
        ("종합점수", 9), ("②지형\n/15", 8), ("③전력철도\n/15", 9),
        ("④인구이격\n/15", 9), ("⑤모델기여\n/10", 9), ("⑥암상\n/5", 7),
        ("경사도(°)", 9), ("전력(km)", 8), ("철도(km)", 8),
        ("네이버 지도 ▶", 20), ("카카오 지도 ▶", 20),
    ]
    for ci, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=ROW_H, column=ci, value=h)
        c.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        c.fill = hdr_fill(C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[ROW_H].height = 32

    # ── 데이터 행 ──
    top_wgs = top.to_crs(WGS84_CRS)
    for i, (_, row) in enumerate(top.iterrows(), start=1):
        r    = ROW_H + i
        rank = int(row["rank"])
        lat  = round(float(top_wgs.geometry.iloc[i-1].y), 6)
        lon  = round(float(top_wgs.geometry.iloc[i-1].x), 6)
        sc   = float(row["score"])
        bg   = C_TOP3 if rank <= 3 else (C_TOP5 if rank <= 5 else C_WHITE)

        def _v(key, fmt=".1f"):
            try:
                v = float(row.get(key, np.nan))
                return "-" if np.isnan(v) else f"{v:{fmt}}"
            except Exception:
                return "-"

        vals = [rank, lat, lon, sc,
                _v("s2_지형"), _v("s3_전력철도"), _v("s4_인구이격"),
                _v("s5_모델기여도"), _v("s6_암상"),
                _v("slope_deg"), _v("d_power_km"), _v("d_railway_km")]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=(ci in [1,4]))
            c.fill = hdr_fill(bg)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

        # 네이버 · 카카오 — =HYPERLINK() 수식
        hl_cell(ws, r, 13, naver_url(lat, lon),
                f"📍 네이버 #{rank}", bg)
        hl_cell(ws, r, 14, kakao_url(lat, lon, rank, sid),
                f"📍 카카오 #{rank}", bg)

    # ── 범례 ──
    leg = ROW_H + TOP_N + 2
    ws.merge_cells(f"A{leg}:N{leg}")
    ws[f"A{leg}"] = ("★1~3위(노랑) 최우선  ●4~5위(하늘) 대안  ○6~10위(흰) 참고  "
                     "│ 종합 = (②+③+④+⑤+⑥)/60×100점")
    ws[f"A{leg}"].font = Font(name="맑은 고딕", size=9, italic=True, color="595959")
    ws[f"A{leg}"].fill = hdr_fill(C_LABEL)
    ws.freeze_panes = f"A{ROW_H+1}"


def make_summary_sheet(wb, all_results):
    ws = wb.create_sheet("📋 전체 요약", 0)
    ws.merge_cells("A1:L1")
    ws["A1"] = "지구자기장 측정 후보지 서브격자 분석 — 전체 요약"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hdr_fill(C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:L2")
    ws["A2"] = (f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  /  "
                f"반경{RADIUS_M//1000}km / 격자{GRID_STEP//1000}km / 도로≤{ROAD_MAX_M//1000}km  │  "
                f"클릭 시 해당 좌표로 지도 이동")
    ws["A2"].font = Font(name="맑은 고딕", size=9, italic=True)

    hdrs = [("사이트",14),("순위",5),("위도(°N)",12),("경도(°E)",12),
            ("종합점수",9),("②지형",7),("③전력철도",8),("④인구",7),
            ("⑤모델기여",8),("⑥암상",6),("네이버 지도 ▶",20),("카카오 지도 ▶",20)]
    for ci, (h, w) in enumerate(hdrs, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        c.fill = hdr_fill(C_HEADER)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    SITE_HDR = {16: C_SITE16, 18: C_SITE18}
    r = 4
    for res in all_results:
        site = res["site"]; sid = site["id"]; top = res["top"]
        top_wgs = top.to_crs(WGS84_CRS)
        hc = SITE_HDR.get(sid, C_HEADER)

        ws.merge_cells(f"A{r}:L{r}")
        c = ws[f"A{r}"]
        c.value = f"▶ {site['name']}  (국가격자 {site['score_national']}점, {site['priority']}등급)"
        c.font  = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
        c.fill  = hdr_fill(hc)
        ws.row_dimensions[r].height = 20
        r += 1

        for i, (_, row) in enumerate(top.iterrows(), start=1):
            rank = int(row["rank"])
            lat  = round(float(top_wgs.geometry.iloc[i-1].y), 6)
            lon  = round(float(top_wgs.geometry.iloc[i-1].x), 6)
            sc   = float(row["score"])
            bg   = C_TOP3 if rank<=3 else (C_TOP5 if rank<=5 else C_WHITE)

            def _v(key, fmt=".1f"):
                try:
                    v = float(row.get(key, np.nan))
                    return "-" if np.isnan(v) else f"{v:{fmt}}"
                except Exception:
                    return "-"

            for ci, val in enumerate(
                [site["name"], rank, lat, lon, sc,
                 _v("s2_지형"), _v("s3_전력철도"), _v("s4_인구이격"),
                 _v("s5_모델기여도"), _v("s6_암상")], start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = Font(name="맑은 고딕", size=10, bold=(ci==5))
                c.fill = hdr_fill(bg)
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border()

            hl_cell(ws, r, 11, naver_url(lat, lon),
                    f"📍 네이버 #{rank}", bg)
            hl_cell(ws, r, 12, kakao_url(lat, lon, rank, sid),
                    f"📍 카카오 #{rank}", bg)
            r += 1
        r += 1   # 사이트 구분 빈 줄

    ws.freeze_panes = "A4"


# 워크북 생성
wb = openpyxl.Workbook()
wb.remove(wb.active)
make_summary_sheet(wb, all_results)
for res in all_results:
    sid = res["site"]["id"]
    color = C_SITE16 if sid == 16 else C_SITE18
    make_sheet(wb, res, f"후보지 #{sid}", color)

xl_path = OUT_DIR / f"{TS}_subgrid_site16_18_survey.xlsx"
wb.save(xl_path)
print(f"\n✅ Excel 저장: {xl_path}")
print(f"   저장 위치: {OUT_DIR}")
print("\n  링크 형식:")
print(f"  네이버: https://map.naver.com/v5/search/{{lat}},{{lon}}")
print(f"  카카오: https://map.kakao.com/link/map/P{{sid}}-{{rank}},{{lat}},{{lon}}")
