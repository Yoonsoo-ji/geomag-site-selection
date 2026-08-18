# -*- coding: utf-8 -*-
"""
도상선점 + 현장 자기교란 조사 통합 카드 v2
- 구성: ①기본(도상선점) 정보(소유권·관할본부 포함) ②현장 자기교란 조사(계량화·자동판정)
        ③방위표지 조사 ④종합 판정(자동) ⑤후보지 주변 사진 ⑥약도(추후)
- 사용: python make_survey_card_v2.py            → 거제 신규점1 1건 (프로토타입)
        python make_survey_card_v2.py --all      → 전체 103건 (추후)
"""
import sys, json, re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from make_survey_cards import (parse_kml, classify, dms_str, nearest_existing,
                               sheet_of, vehicle_access, cautions, reverse_geocode,
                               embed_map, EXISTING, TOPO)
from capture_map_images import IMG_DIR
from interference import Interference
from shapely.geometry import shape

ROOT    = Path(__file__).parent
OUT_DIR = ROOT / "docs/output"

# ── LX 본부 매핑 ─────────────────────────────────────────────────────────────
HQ_LIST = ["서울경기북부 본부", "부산울산 본부", "인천경기남부 본부", "강원 본부",
           "충북 본부", "대전세종충남 본부", "전북 본부", "광주전남 본부",
           "대구경북 본부", "경남 본부", "제주 본부"]
GG_NORTH = ["고양", "파주", "의정부", "양주", "동두천", "포천", "연천",
            "남양주", "구리", "가평"]

def assign_hq(addr):
    """지번 주소 → LX 관할 본부"""
    if not addr:
        return ""
    a = addr.strip()
    toks = a.split()
    sido = toks[0] if toks else ""
    if sido.startswith("서울"):
        return "서울경기북부 본부"
    if sido.startswith(("부산", "울산")):
        return "부산울산 본부"
    if sido.startswith("인천"):
        return "인천경기남부 본부"
    if sido.startswith("경기"):
        city = toks[1] if len(toks) > 1 else ""
        return ("서울경기북부 본부"
                if any(city.startswith(g) for g in GG_NORTH)
                else "인천경기남부 본부")
    if sido.startswith("강원"):
        return "강원 본부"
    if sido.startswith(("충북", "충청북도")):
        return "충북 본부"
    if sido.startswith(("대전", "세종", "충남", "충청남도")):
        return "대전세종충남 본부"
    if sido.startswith(("전북", "전라북도")):
        return "전북 본부"
    if sido.startswith(("광주", "전남", "전라남도")):
        return "광주전남 본부"
    if sido.startswith(("대구", "경북", "경상북도")):
        return "대구경북 본부"
    if sido.startswith(("경남", "경상남도")):
        return "경남 본부"
    if sido.startswith("제주"):
        return "제주 본부"
    return ""

# 역지오코딩 키 없이 실행할 때의 주소 보완 (프로토타입용)
ADDR_FALLBACK = {
    "거제 신규점1": ("경남 거제시 하청면 유계리 산 213", "-"),
}

# ── Nominatim 역지오코딩 (Kakao 키 없을 때 본부/주소 확보) ────────────────────
import requests
_NOM_CACHE = ROOT / "docs/output/_card_region_cache.json"

def _load_region_cache():
    if _NOM_CACHE.exists():
        try:
            return json.load(open(_NOM_CACHE, encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _save_region_cache(cache):
    json.dump(cache, open(_NOM_CACHE, "w", encoding="utf-8"), ensure_ascii=False, indent=0)

def region_via_nominatim(lat, lon, cache):
    """(시도, 시군구, 표시주소) 반환. 캐시 우선. 실패 시 ('','','')"""
    import time as _t
    key = f"{lat:.6f},{lon:.6f}"
    if key in cache:
        return tuple(cache[key])
    prov = city = disp = ""
    try:
        r = requests.get("https://nominatim.openstreetmap.org/reverse",
                         params={"format": "json", "lat": lat, "lon": lon,
                                 "addressdetails": 1, "accept-language": "ko"},
                         headers={"User-Agent": "LX-geomag-survey/1.0"}, timeout=15)
        d = r.json()
        a = d.get("address", {})
        prov = a.get("province") or a.get("state") or ""
        city = (a.get("city") or a.get("county") or a.get("town")
                or a.get("borough") or a.get("municipality") or "")
        disp = d.get("display_name", "") or ""
    except Exception as e:
        print("  [nominatim 실패]", lat, lon, e)
    cache[key] = [prov, city, disp]
    _t.sleep(1.05)   # Nominatim 사용정책: 최대 1req/sec
    return prov, city, disp


def hq_from_region(prov, city):
    """(시도, 시군구) → LX 관할 본부"""
    p = (prov or "").strip()
    c = (city or "").strip()
    if p.startswith("서울"):                          return "서울경기북부 본부"
    if p.startswith(("부산", "울산")):                return "부산울산 본부"
    if p.startswith("인천"):                          return "인천경기남부 본부"
    if p.startswith("경기"):
        return ("서울경기북부 본부"
                if any(c.startswith(g) for g in GG_NORTH) else "인천경기남부 본부")
    if p.startswith("강원"):                          return "강원 본부"
    if p.startswith(("충청북도", "충북")):            return "충북 본부"
    if p.startswith(("대전", "세종", "충청남도", "충남")):  return "대전세종충남 본부"
    if p.startswith(("전라북도", "전북")):            return "전북 본부"
    if p.startswith(("광주", "전라남도", "전남")):    return "광주전남 본부"
    if p.startswith(("대구", "경상북도", "경북")):    return "대구경북 본부"
    if p.startswith(("경상남도", "경남")):            return "경남 본부"
    if p.startswith("제주"):                          return "제주 본부"
    return ""


def nominatim_jibun(disp):
    """Nominatim display_name → '경남 거제시 하청면 유계리' 형태 축약 주소"""
    if not disp:
        return ""
    parts = [x.strip() for x in disp.split(",") if x.strip()]
    parts = [x for x in parts if x not in ("대한민국", "South Korea")]
    parts = [x for x in parts if not re.fullmatch(r"\d{3}-?\d{2,3}", x)]  # 우편번호 제거
    return " ".join(reversed(parts[-4:])) if parts else ""

# ── 자기교란 점검 항목 (항목, 확인 내용, 기준 이격 m; '-'=존재 시 부적합) ────
DISTURB_ITEMS = [
    ("차량 영향", "차량 통행 및 주차",          80),
    ("전력시설",  "송전선로·철탑",              60),
    ("통신시설",  "기지국·통신탑",              50),
    ("금속류",    "난간·가드레일 등",           30),
    ("매설물",    "매설 구조물",               "-"),
    ("건축물",    "주택·창고·콘크리트 구조물",  70),
    ("자연환경",  "수목 등 상공장애도 여부",     "-"),
]

# ── 스타일 ──────────────────────────────────────────────────────────────────
F_TITLE = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
F_SEC   = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
F_LBL   = Font(name='맑은 고딕', size=10, bold=True)
F_VAL   = Font(name='맑은 고딕', size=10)
F_CAP   = Font(name='맑은 고딕', size=9, italic=True, color='555555')
F_HINT  = Font(name='맑은 고딕', size=10, color='999999')
F_AUTO  = Font(name='맑은 고딕', size=10, bold=True, color='1F3864')
FILL_TITLE = PatternFill('solid', fgColor='1F3864')
FILL_SEC   = PatternFill('solid', fgColor='4472C4')
FILL_LBL   = PatternFill('solid', fgColor='F2F2F2')
FILL_FIELD = PatternFill('solid', fgColor='FFF8E1')   # 현장 기입란
FILL_AUTO  = PatternFill('solid', fgColor='E8F0E8')   # 자동 계산란
FILL_NA    = PatternFill('solid', fgColor='E0E0E0')   # 해당 없음
AL_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
_thin  = Side(style='thin', color='BBBBBB')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def lbl(ws, ref, text):
    c = ws[ref]; c.value = text; c.font = F_LBL; c.fill = FILL_LBL
    c.alignment = AL_C; c.border = BORDER

def val(ws, ref, text, field=False, auto=False, na=False, center=False):
    c = ws[ref]; c.value = text; c.font = F_AUTO if auto else F_VAL
    c.alignment = AL_C if center else AL_L; c.border = BORDER
    if field: c.fill = FILL_FIELD
    if auto:  c.fill = FILL_AUTO
    if na:    c.fill = FILL_NA; c.alignment = AL_C
    return c

def sec(ws, row, text):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]; c.value = text; c.font = F_SEC; c.fill = FILL_SEC
    c.alignment = AL_L

AL_TL = Alignment(horizontal='left', vertical='top', wrap_text=True)

def cap(ws, row, text, height=None):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]; c.value = text; c.font = F_CAP; c.alignment = AL_TL
    if height is None:
        # 병합 폭(A~G) 기준 줄 수 추정. 한글은 2배폭 → 한 줄 ≈ 44자(한글 기준)
        # 명시 줄바꿈(\n)도 반영
        segs = str(text).split('\n')
        lines = sum(max(1, -(-len(s) // 44)) for s in segs)
        height = 15 * lines + 6
    ws.row_dimensions[row].height = height

def border_range(ws, ref):
    for row in ws[ref]:
        for c in row:
            c.border = BORDER

def photo_box(ws, ref, hint):
    """사진 업로드 칸 — 조사자 입력 영역이므로 토글 셀과 동일한 배경색(FILL_FIELD)"""
    ws.merge_cells(ref)
    a = ref.split(':')[0]
    c = ws[a]; c.value = hint; c.font = F_HINT; c.alignment = AL_C
    for row in ws[ref]:
        for cc in row:
            cc.fill = FILL_FIELD
    border_range(ws, ref)


def build_card_v2(wb, idx, p, exi, dist, sh, inf, imgs):
    sid  = f"DS-{idx:03d}"
    safe = re.sub(r'[\\/*?:\[\]]', '', p['name'])[:22]
    ws = wb.create_sheet(f"{idx:02d}_{safe}")

    for col, w in zip("ABCDEFG", [13, 11, 11, 11, 11, 11, 11]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["H"].width = 2.5      # 카드 ↔ 지도 간격 열
    for col in "IJKL":
        ws.column_dimensions[col].width = 21

    cat = classify(p['name'], p['desc'])
    jibun = p.get('jibun') or "-"
    hq = p.get('hq') or assign_hq(jibun)

    dv_own, dv_hq, dv_exist, dv_ok, dv_target, dv_weather = make_dvs(ws)

    # ── 제목 (조사 카드 A~G) + 현장 위치 헤더 (지도 H~K, 분리 표시) ──
    ws.merge_cells("A1:G1")
    t = ws["A1"]; t.value = f"자기교란 현황 조사 카드   |   {sid}   {p['name']}"
    t.font = F_TITLE; t.fill = FILL_TITLE; t.alignment = AL_L
    ws.merge_cells("I1:L1")
    mt = ws["I1"]; mt.value = "현장 위치"
    mt.font = F_TITLE; mt.fill = FILL_TITLE; mt.alignment = AL_C
    ws.row_dimensions[1].height = 30

    # ── ① 기본(도상선점) 정보 ──
    sec(ws, 2, "① 기본(도상선점) 정보")
    lbl(ws, "A3", "관리번호"); ws.merge_cells("B3:C3"); val(ws, "B3", sid)
    lbl(ws, "D3", "후보지명"); ws.merge_cells("E3:G3"); val(ws, "E3", p['name'])
    lbl(ws, "A4", "도엽번호"); ws.merge_cells("B4:C4"); val(ws, "B4", sh.get('sheet_code', '-'))
    lbl(ws, "D4", "도엽명");   ws.merge_cells("E4:G4"); val(ws, "E4", sh.get('sheet_name', '-'))
    lbl(ws, "A5", "위도(십진)"); ws.merge_cells("B5:C5"); val(ws, "B5", f"{p['lat']:.6f}")
    lbl(ws, "D5", "경도(십진)"); ws.merge_cells("E5:G5"); val(ws, "E5", f"{p['lon']:.6f}")
    lbl(ws, "A6", "좌표(도분초)"); ws.merge_cells("B6:G6")
    val(ws, "B6", f"{dms_str(p['lat'],'N','S')}  /  {dms_str(p['lon'],'E','W')}")
    lbl(ws, "A7", "표고"); ws.merge_cells("B7:C7")
    val(ws, "B7", f"{p['elev']:.0f} m" if p['elev'] is not None else "-")
    lbl(ws, "D7", "후보지 유형"); ws.merge_cells("E7:G7"); val(ws, "E7", cat)
    lbl(ws, "A8", "지번 주소"); ws.merge_cells("B8:G8"); val(ws, "B8", jibun)
    lbl(ws, "A9", "도로명 주소"); ws.merge_cells("B9:G9"); val(ws, "B9", p.get('road') or "-")
    # 소유권 조사(드롭다운) + 관할 본부
    lbl(ws, "A10", "소유권 조사"); ws.merge_cells("B10:C10")
    c = val(ws, "B10", "", field=True, center=True); dv_own.add(ws["B10"])
    lbl(ws, "D10", "관할 본부"); ws.merge_cells("E10:G10")
    val(ws, "E10", hq, center=True); dv_hq.add(ws["E10"])
    # 조사대상(드롭다운) + 조사일
    default_target = "기존점" if "기존" in p['name'] else "신규 후보지"
    lbl(ws, "A11", "조사대상"); ws.merge_cells("B11:C11")
    val(ws, "B11", default_target, field=True, center=True); dv_target.add(ws["B11"])
    lbl(ws, "D11", "조사일\n(예: 20260707)"); ws.merge_cells("E11:G11")
    val(ws, "E11", "", field=True, center=True)
    ws["E11"].number_format = "0000-00-00"   # 20260707 입력 → 2026-07-07 표시
    dv_date = DataValidation(showInputMessage=True, allow_blank=True,
                             promptTitle="조사일 입력",
                             prompt="8자리 숫자로 입력하세요.\n예: 20260707 → 2026-07-07 자동 표시")
    ws.add_data_validation(dv_date); dv_date.add(ws["E11"])
    # 조사자 + 기상상태(드롭다운)
    lbl(ws, "A12", "조사자"); ws.merge_cells("B12:C12")
    val(ws, "B12", "", field=True, center=True)
    lbl(ws, "D12", "기상상태"); ws.merge_cells("E12:G12")
    val(ws, "E12", "", field=True, center=True); dv_weather.add(ws["E12"])
    lbl(ws, "A13", "연계 기존점"); ws.merge_cells("B13:G13")
    val(ws, "B13", f"{exi['name']} (직선거리 약 {dist:.1f} km)")
    lbl(ws, "A14", "접근 경로"); ws.merge_cells("B14:G14")
    val(ws, "B14", p['desc'] or "현장확인"); ws.row_dimensions[14].height = 58
    lbl(ws, "A15", "차량 진입"); ws.merge_cells("B15:C15"); val(ws, "B15", vehicle_access(p['desc']))
    lbl(ws, "D15", "주차→도보"); ws.merge_cells("E15:G15"); val(ws, "E15", "현장확인")
    lbl(ws, "A16", "유의사항"); ws.merge_cells("B16:G16"); val(ws, "B16", cautions(p['desc']))
    lbl(ws, "A17", "내비 바로가기")
    kakao  = f"https://map.kakao.com/link/map/{p['name']},{p['lat']:.6f},{p['lon']:.6f}"
    naver  = f"https://map.naver.com/v5/search/{p['lat']:.6f},{p['lon']:.6f}"
    google = f"https://www.google.com/maps/search/?api=1&query={p['lat']:.6f},{p['lon']:.6f}"
    for ref, txt, url in [("B", "카카오맵", kakao), ("D", "네이버지도", naver), ("F", "구글지도", google)]:
        ws.merge_cells(f"{ref}17:{chr(ord(ref)+1)}17")
        c = ws[f"{ref}17"]; c.value = txt; c.hyperlink = url
        c.font = Font(name='맑은 고딕', size=10, color='1A73E8', underline='single')
        c.alignment = AL_C; c.border = BORDER
    lbl(ws, "A18", "도상 간섭요인\n(자동 계산)")
    ws.merge_cells("B18:G18")
    val(ws, "B18",
        f"송전탑 {inf['tower']*1000:.0f} m ({'✓' if inf['tower']>=1 else '✗'} 기준 1km)  |  "
        f"철도 {inf['rail']:.1f} km ({'✓' if inf['rail']>=5 else '✗'} 기준 5km)  |  "
        f"통신탑 {inf['comm']:.1f} km  |  주거지 {inf['resid']:.2f} km", auto=True)
    ws.row_dimensions[18].height = 30

    # ── ② 현장 자기교란 조사 ──
    sec(ws, 19, "② 현장 자기교란 조사")
    cap(ws, 20, "기입 방법: 존재여부(있음/없음)를 드롭다운으로 선택 → 판정은 자동 (없음→적합, 있음→부적합). "
                "'있음'이면 실측 이격거리(m)를 참고용으로 기입. 방위표지는 가능→적합, 불가→부적합.")
    lbl(ws, "A21", "점검 항목")
    ws.merge_cells("B21:C21"); lbl(ws, "B21", "확인 내용")
    lbl(ws, "D21", "존재여부")
    lbl(ws, "E21", "실측 이격(m)")
    ws.merge_cells("F21:G21"); lbl(ws, "F21", "판정(자동)")
    r0 = 22
    for i, (name, detail, std) in enumerate(DISTURB_ITEMS):
        r = r0 + i
        lbl(ws, f"A{r}", name)
        ws.merge_cells(f"B{r}:C{r}"); val(ws, f"B{r}", detail)
        val(ws, f"D{r}", "", field=True, center=True); dv_exist.add(ws[f"D{r}"])
        val(ws, f"E{r}", "-", field=True, center=True)
        ws.merge_cells(f"F{r}:G{r}")
        val(ws, f"F{r}", f'=IF(D{r}="","",IF(D{r}="없음","적합","부적합"))', auto=True, center=True)
    r_az = r0 + len(DISTURB_ITEMS)          # 방위표지 행
    lbl(ws, f"A{r_az}", "방위표지")
    ws.merge_cells(f"B{r_az}:C{r_az}"); val(ws, f"B{r_az}", "방위표지 설치 가능 여부")
    val(ws, f"D{r_az}", "", field=True, center=True); dv_ok.add(ws[f"D{r_az}"])
    val(ws, f"E{r_az}", "-", na=True)
    ws.merge_cells(f"F{r_az}:G{r_az}")
    val(ws, f"F{r_az}", f'=IF(D{r_az}="","",IF(D{r_az}="가능","적합","부적합"))', auto=True, center=True)
    judge_rng = f"F{r0}:F{r_az}"
    n_items = len(DISTURB_ITEMS) + 1
    r_sum = r_az + 1                        # 계량화 결과
    lbl(ws, f"A{r_sum}", "판정 집계")
    ws.merge_cells(f"B{r_sum}:G{r_sum}")
    val(ws, f"B{r_sum}",
        f'="적합 "&COUNTIF({judge_rng},"적합")&"건  /  부적합 "&COUNTIF({judge_rng},"부적합")'
        f'&"건  /  미입력 "&({n_items}-COUNTIF({judge_rng},"적합")-COUNTIF({judge_rng},"부적합"))&"건"',
        auto=True)
    r_pic = r_sum + 1                       # 29
    lbl(ws, f"A{r_pic}", "교란원 사진")
    photo_box(ws, f"B{r_pic}:G{r_pic}", "📷 교란원 사진 삽입 (해당 시)")
    ws.row_dimensions[r_pic].height = 110

    # ── ③ 방위표지 조사 ──
    r = r_pic + 1                           # 30
    sec(ws, r, "③ 방위표지 조사 (경위도 입력 → 방위각 진북 기준 자동 계산)")
    r += 1                                  # 31 헤더
    lbl(ws, f"A{r}", "구분")
    ws.merge_cells(f"B{r}:C{r}"); lbl(ws, f"B{r}", "기준점 좌표")
    ws.merge_cells(f"D{r}:E{r}"); lbl(ws, f"D{r}", "방위표지 1")
    ws.merge_cells(f"F{r}:G{r}"); lbl(ws, f"F{r}", "방위표지 2")
    rows3 = ["경도", "위도", "표고(m)", "방위각(°)", "거리(m)"]
    r_lon, r_lat = r + 1, r + 2             # 경도·위도 행 (수식 참조용)

    def az_formula(lon_b, lat_b, lon_t, lat_t):
        """기준점→표지 진북방위각 (WGS84 국소평면 ENU 근사, 단거리 오차 <1″)"""
        ph = f"RADIANS(({lat_b}+{lat_t})/2)"
        w  = f"(1-0.00669437999014*SIN({ph})^2)"
        dn = f"RADIANS({lat_t}-{lat_b})*6378137*0.99330562000986/POWER({w},1.5)"
        de = f"RADIANS({lon_t}-{lon_b})*6378137/SQRT({w})*COS({ph})"
        return (f'=IF(OR({lon_b}="",{lat_b}="",{lon_t}="",{lat_t}=""),"",'
                f'MOD(DEGREES(ATAN2({dn},{de}))+360,360))')

    for i, name in enumerate(rows3):
        rr = r + 1 + i
        if name == "방위각(°)":
            lbl(ws, f"A{rr}", "방위각(°)\n(진북·자동)")
            ws.merge_cells(f"B{rr}:C{rr}"); val(ws, f"B{rr}", "-", na=True)
            for col in ("D", "F"):
                ws.merge_cells(f"{col}{rr}:{chr(ord(col)+1)}{rr}")
                val(ws, f"{col}{rr}",
                    az_formula(f"B{r_lon}", f"B{r_lat}", f"{col}{r_lon}", f"{col}{r_lat}"),
                    auto=True, center=True)
                ws[f"{col}{rr}"].number_format = "0.0000"
            continue
        lbl(ws, f"A{rr}", name)
        na_base = name == "거리(m)"           # 기준점에는 거리 없음
        ws.merge_cells(f"B{rr}:C{rr}")
        val(ws, f"B{rr}", "-" if na_base else "", field=not na_base, na=na_base, center=True)
        ws.merge_cells(f"D{rr}:E{rr}"); val(ws, f"D{rr}", "", field=True, center=True)
        ws.merge_cells(f"F{rr}:G{rr}"); val(ws, f"F{rr}", "", field=True, center=True)
    r_judge_sec = r + len(rows3) + 1        # 37

    # ── ④ 종합 판정 ──
    sec(ws, r_judge_sec, "④ 종합 판정")
    r1 = r_judge_sec + 1                    # 38 판정결과
    lbl(ws, f"A{r1}", "판정 결과\n(자동)")
    ws.merge_cells(f"B{r1}:G{r1}")
    val(ws, f"B{r1}",
        f'=IF(COUNTIF({judge_rng},"적합")+COUNTIF({judge_rng},"부적합")<{n_items},"조사 미완료",'
        f'IF(COUNTIF({judge_rng},"부적합")=0,"적합",'
        f'IF(COUNTIF({judge_rng},"부적합")<=2,"조건부 적합","부적합")))', auto=True, center=True)
    r2 = r1 + 1                             # 39 판정의견
    lbl(ws, f"A{r2}", "판정 의견\n(조사자)")
    ws.merge_cells(f"B{r2}:G{r2}"); val(ws, f"B{r2}", "", field=True)
    ws.row_dimensions[r2].height = 50
    r3 = r2 + 1                             # 40 향후조치
    lbl(ws, f"A{r3}", "향후 조치\n(자동)")
    ws.merge_cells(f"B{r3}:G{r3}")
    val(ws, f"B{r3}",
        f'=IF(B{r1}="적합","자기구배 조사 진행",IF(B{r1}="조건부 적합","추가 자기조사 후 재판정",'
        f'IF(B{r1}="부적합","대체 후보지 검토","② 조사 완료 후 자동 표시")))', auto=True, center=True)
    r4 = r3 + 1                             # 41 규칙 캡션
    cap(ws, r4, "판정 규칙: ② 각 항목 존재여부로 자동 판정(없음→적합, 있음→부적합). 부적합 0건=적합 / 1~2건=조건부 적합 / "
                "3건 이상=부적합 (미입력 존재 시 '조사 미완료')  |  조치: 적합→자기구배 조사, 조건부→추가 자기조사, 부적합→대체 후보지 검토")

    # ── ⑤ 후보지 주변 사진 ──
    r5 = r4 + 1                             # 42
    sec(ws, r5, "⑤ 후보지 주변 사진")
    rc = r5 + 1                             # 43 중심점
    ws.merge_cells(f"A{rc}:A{rc+1}"); lbl(ws, f"A{rc}", "지자기점\n(중심점)")
    photo_box(ws, f"B{rc}:G{rc+1}", "📷 지자기점(중심점) 사진 삽입")
    ws.row_dimensions[rc].height = 60; ws.row_dimensions[rc+1].height = 60
    rp = rc + 2                             # 45
    ws.merge_cells(f"A{rp}:A{rp+5}"); lbl(ws, f"A{rp}", "주변 사진\n(동서남북)")
    ws.merge_cells(f"B{rp}:D{rp}"); lbl(ws, f"B{rp}", "동측")
    ws.merge_cells(f"E{rp}:G{rp}"); lbl(ws, f"E{rp}", "서측")
    photo_box(ws, f"B{rp+1}:D{rp+2}", "📷 동측 사진")
    photo_box(ws, f"E{rp+1}:G{rp+2}", "📷 서측 사진")
    ws.merge_cells(f"B{rp+3}:D{rp+3}"); lbl(ws, f"B{rp+3}", "남측")
    ws.merge_cells(f"E{rp+3}:G{rp+3}"); lbl(ws, f"E{rp+3}", "북측")
    photo_box(ws, f"B{rp+4}:D{rp+5}", "📷 남측 사진")
    photo_box(ws, f"E{rp+4}:G{rp+5}", "📷 북측 사진")
    for rr in (rp+1, rp+2, rp+4, rp+5):
        ws.row_dimensions[rr].height = 55

    # ── ⑥ 약도 (추후 작업) ──
    r6 = rp + 6                             # 51
    sec(ws, r6, "⑥ 약도")
    ws.merge_cells(f"A{r6+1}:G{r6+4}")
    c = ws[f"A{r6+1}"]
    c.value = "※ 추후 작업 — 최종 기준점 좌표와 방위표지 1·2 좌표가 확정되면 지도 기반 약도를 자동 생성하여 삽입 예정."
    c.font = F_HINT; c.alignment = AL_C
    border_range(ws, f"A{r6+1}:G{r6+4}")
    for rr in range(r6+1, r6+5):
        ws.row_dimensions[rr].height = 35

    # ── 테두리 마무리: A~G 영역 모든 병합 범위 전 셀에 테두리 채움 ──
    # (openpyxl은 병합 시 좌상단 셀에만 스타일 적용 → 나머지 셀 테두리 보정)
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 2 and mr.max_col <= 7:
            for row in ws[mr.coord]:
                for cc in row:
                    cc.border = BORDER

    # ── 지도 (우측 2x2) ──
    embed_map(ws, imgs.get('sat'),  'I2',  "위성 근접 (현장 지형·임도)")
    embed_map(ws, imgs.get('sur'),  'K2',  "주변 (도로·마을)")
    embed_map(ws, imgs.get('int'),  'I19', "간섭요인 (송전탑·철도 버퍼)")
    embed_map(ws, imgs.get('wide'), 'K19', "광역 위치")

    return ws, sid, cat, hq


def make_dvs(ws_container):
    """드롭다운 정의 (워크북 공용으로 각 시트에 add)"""
    dv_own    = DataValidation(type="list", formula1='"국공유지,사유지"', allow_blank=True,
                               promptTitle="소유권", prompt="국공유지 / 사유지 선택")
    dv_hq     = DataValidation(type="list", formula1='"' + ",".join(HQ_LIST) + '"', allow_blank=True)
    dv_exist  = DataValidation(type="list", formula1='"있음,없음"', allow_blank=True)
    dv_ok     = DataValidation(type="list", formula1='"가능,불가"', allow_blank=True)
    dv_target = DataValidation(type="list", formula1='"기존점,신규 후보지"', allow_blank=True,
                               promptTitle="조사대상", prompt="기존점 / 신규 후보지 선택")
    dv_weather= DataValidation(type="list", formula1='"맑음,흐림,눈 또는 비"', allow_blank=True)
    for dv in (dv_own, dv_hq, dv_exist, dv_ok, dv_target, dv_weather):
        ws_container.add_data_validation(dv)
    return dv_own, dv_hq, dv_exist, dv_ok, dv_target, dv_weather


# ── 총괄(마스터) 목록 시트 ───────────────────────────────────────────────────
MASTER_COLS = ["관리번호", "후보지명", "관할 본부", "지번 주소", "도엽번호", "도엽명",
               "위도", "경도", "표고(m)", "유형", "연계 기존점", "기존점거리(km)",
               "최근접 송전탑(m)", "최근접 철도(km)", "최근접 주거지(km)", "카드"]

def build_master(wb, rows, title="도상선점 자기교란 조사 — 총괄 목록", pos=0):
    ws = wb.create_sheet("총괄 목록", pos)
    widths = [9, 15, 15, 26, 11, 9, 10, 10, 8, 10, 12, 11, 13, 12, 13, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells(f"A1:{get_column_letter(len(MASTER_COLS))}1")
    t = ws["A1"]; t.value = title; t.font = F_TITLE; t.fill = FILL_TITLE; t.alignment = AL_L
    ws.row_dimensions[1].height = 26
    for j, h in enumerate(MASTER_COLS, 1):
        c = ws.cell(2, j, h)
        c.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
        c.fill = FILL_SEC; c.alignment = AL_C; c.border = BORDER
    for r, (row, sheet_title) in enumerate(rows, 3):
        for j, v in enumerate(row, 1):
            c = ws.cell(r, j, v); c.font = F_VAL; c.alignment = AL_C; c.border = BORDER
        lc = ws.cell(r, len(MASTER_COLS), "이동")
        lc.hyperlink = f"#'{sheet_title}'!A1"
        lc.font = Font(name='맑은 고딕', size=10, color='1A73E8', underline='single')
        lc.alignment = AL_C; lc.border = BORDER
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(MASTER_COLS))}{len(rows)+2}"
    return ws


def build_workbook(recs, title):
    """recs: [(idx, p, exi, dist, sh, inf), ...] → master + 카드 워크북"""
    wb = Workbook(); wb.remove(wb.active)
    master_rows = []
    for idx, p, exi, dist, sh, inf in recs:
        imgs = {k: IMG_DIR / f"{idx:02d}_{k}.png" for k in ("sat", "sur", "int", "wide")}
        ws, sid, cat, hq = build_card_v2(wb, idx, p, exi, dist, sh, inf, imgs)
        master_rows.append(([
            sid, p['name'], hq or "미지정", p.get('jibun') or "-",
            sh.get('sheet_code', '-'), sh.get('sheet_name', '-'),
            f"{p['lat']:.5f}", f"{p['lon']:.5f}",
            f"{p['elev']:.0f}" if p['elev'] is not None else "-",
            cat, exi['name'], f"{dist:.1f}",
            f"{inf['tower']*1000:.0f}", f"{inf['rail']:.1f}", f"{inf['resid']:.2f}",
        ], ws.title))
    build_master(wb, master_rows, title=title, pos=0)
    return wb


def main():
    build_all = "--all" in sys.argv
    pts = parse_kml()
    targets = list(range(1, len(pts) + 1)) if build_all else \
              [i for i, p in enumerate(pts, 1) if p['name'] == "거제 신규점1"]
    if not targets:
        print("대상 지점을 찾지 못했습니다."); return

    ex = [f['properties'] for f in json.load(open(EXISTING, encoding='utf-8'))['features']]
    polys = [(f['properties'], shape(f['geometry']))
             for f in json.load(open(TOPO, encoding='utf-8'))['features']]
    print("간섭요인 데이터 로딩...")
    inf_calc = Interference()
    region_cache = _load_region_cache()

    # ── 지점별 데이터 사전 계산 (역지오코딩·본부 포함) ──
    recs = []
    hq_count = {}
    print(f"지점 {len(targets)}건 처리 (역지오코딩 포함)...")
    for n, idx in enumerate(targets, 1):
        p = pts[idx - 1]
        exi, dist = nearest_existing(p['lat'], p['lon'], ex)
        # 도엽은 **후보지 자신의 좌표**로 판정한다. 기존점(exi) 좌표를 쓰면
        # 연계 기존점의 도엽이 찍혀 실제 도엽과 어긋난다(2026-08 정정).
        sh  = sheet_of(p['lon'], p['lat'], polys)
        inf = inf_calc.all(p['lat'], p['lon'])
        # 주소·본부: Kakao → 없으면 Nominatim (시도/시군 + display_name 축약주소)
        jibun, road = reverse_geocode(p['lat'], p['lon'])
        hq = assign_hq(jibun)
        if not hq:
            prov, city, disp = region_via_nominatim(p['lat'], p['lon'], region_cache)
            hq = hq_from_region(prov, city)
            if not jibun:
                jibun = nominatim_jibun(disp)
            if not hq:                      # 시도 필드 누락 시 축약주소로 재판정
                hq = assign_hq(jibun)
        if not jibun and p['name'] in ADDR_FALLBACK:
            jibun, road = ADDR_FALLBACK[p['name']]
            hq = hq or assign_hq(jibun)
        p['jibun'], p['road'], p['hq'] = jibun or "-", road or "-", hq
        recs.append((idx, p, exi, dist, sh, inf))
        hq_count[hq or "미지정"] = hq_count.get(hq or "미지정", 0) + 1
        if n % 10 == 0 or n == len(targets):
            print(f"  {n}/{len(targets)} 처리")
    _save_region_cache(region_cache)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not build_all:
        wb = build_workbook(recs, "도상선점 자기교란 조사 — 총괄 목록")
        out = OUT_DIR / f"{ts}_조사카드_v2_DS-001_거제신규점1.xlsx"
        wb.save(out); print(f"저장 완료: {out}")
        return

    # ── ① 총괄 파일 (전체 103건 + 총괄 목록) ──
    wb_all = build_workbook(recs, "도상선점 자기교란 조사 — 총괄 목록 (전체 103건)")
    out_all = OUT_DIR / f"{ts}_조사카드_v2_총괄_전체103건.xlsx"
    wb_all.save(out_all)
    print(f"\n[총괄] 저장: {out_all}")

    # ── ② 본부별 분리 파일 ──
    hq_dir = OUT_DIR / f"{ts}_조사카드_v2_본부별"
    hq_dir.mkdir(exist_ok=True)
    order = HQ_LIST + ["미지정"]
    for hq in order:
        sub = [r for r in recs if (r[1].get('hq') or "미지정") == hq]
        if not sub:
            continue
        wb_hq = build_workbook(sub, f"[{hq}] 도상선점 자기교란 조사 — {len(sub)}건")
        safe_hq = hq.replace(" ", "")
        out_hq = hq_dir / f"{ts}_조사카드_v2_{safe_hq}_{len(sub)}건.xlsx"
        wb_hq.save(out_hq)
        print(f"  [{hq}] {len(sub)}건 → {out_hq.name}")

    print("\n== 본부별 지점 수 ==")
    for hq in order:
        if hq_count.get(hq):
            print(f"  {hq}: {hq_count[hq]}건")


if __name__ == "__main__":
    main()
