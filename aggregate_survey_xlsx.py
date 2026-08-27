# -*- coding: utf-8 -*-
"""
현장 조사서(엑셀 카드) → 취합 + 선점 검토 + 웹 표출용 정리
==========================================================

각 본부가 회신한 조사서 워크북(총괄 목록 + 카드 시트들)을 모두 읽어

  1) [현장조사 취합]  한 카드 = 한 행. 도상선점 정보 + 현장 조사 결과 전부.
  2) [선점 검토]      취합 결과를 근거로 선점 가능성 등급·결론·핵심 교란요인 정리
                      (웹 페이지 표출을 염두에 둔 열 구성).
  3) [검토 요약]      본부별·판정별 분포, 주요 교란요인 통계.

카드 시트는 make_survey_card_v2.py 로 생성된 고정 레이아웃이므로 셀 위치가 일정하다.
자동판정(F열·집계·종합판정)은 캐시된 수식값 대신 존재여부에서 규칙으로 재계산해
파일이 Excel 에서 평가되지 않았어도 안전하게 판정한다.

사용:
    python aggregate_survey_xlsx.py                        # 기본 폴더 전체
    python aggregate_survey_xlsx.py --dir <폴더> --out <파일.xlsx>
"""
import argparse
import math
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
# 회신 조사서가 놓인 베이스 폴더(하위 폴더명은 자주 바뀌므로 재귀 탐색한다)
DEF_BASE = Path(r"D:\LX_yoons\2026_research\2026_지자기 연구\20260811_사전 현장 조사서 취합")
DEF_DIR = DEF_BASE   # 하위호환 별칭
OUT_DIR = ROOT / "docs" / "output"


def survey_files(base=None):
    """베이스 폴더 아래에서 'N.조사서_지역.xlsx' 회신본을 재귀 탐색(폴더명 변경에 견고)."""
    base = Path(base) if base else DEF_BASE
    fs = [p for p in base.rglob("*.조사서_*.xlsx") if not p.name.startswith("~$")]
    return sorted(fs, key=lambda p: int(re.match(r"(\d+)", p.name).group(1))
                  if re.match(r"(\d+)", p.name) else 999)

# (카드 라벨, 취합 열 이름)
DIST_ITEMS = [("차량 영향", "차량영향"), ("전력시설", "전력시설"), ("통신시설", "통신시설"),
              ("금속류", "금속류"), ("매설물", "매설물"), ("건축물", "건축물"),
              ("자연환경", "자연환경")]

# ── 검토 단계 수동 재판정 (도상선점↔현장 기준점 좌표 1km 이상 불일치 5건) ──
# {관리번호: (종합판정, 재판정 사유)}. 원본 카드는 불변, 취합·검토·총괄·웹에만 반영.
_SHADOW = "도상 후보지 통신 음영지역으로 확인 어려운 지역 → 부적합"
JUDGMENT_OVERRIDE = {
    "DS-002": ("부적합", _SHADOW),
    "DS-041": ("부적합", _SHADOW),
    "DS-061": ("부적합", _SHADOW),
    # DS-011 은 후보지↔기준점 2,049m 이격만이 사유였는데, 이격은 도상 후보지로
    # **진입이 어려워 옮긴 것**일 뿐 부적합 사유가 아니다(성과 기준은 현장 기준점).
    # 카드 원판정 조건부 적합 → 등급 C 로 되돌린다.
    "DS-087": ("부적합", "상공장애로 취득 불가 → 부적합"),
}


# ── 카드 시트 파싱 ────────────────────────────────────────────────────────────
def _norm(v):
    return re.split(r"[\n(]", str(v))[0].strip() if v is not None else ""


def label_rows(ws, col):
    m = {}
    for r in range(1, ws.max_row + 1):
        k = _norm(ws.cell(r, col).value)
        if k:
            m.setdefault(k, r)
    return m


def _s(v):
    return "" if v is None else str(v).strip()


def dms2dd(v):
    """도분초 문자열 → 십진도. 형식 무관(대시·°'\"·공백). 앞 숫자 3개를 도·분·초로 본다."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "-":
        return None
    nums = re.findall(r"\d*\.?\d+", s)   # '.6614'(앞자리 0 생략)도 초로 인식
    if len(nums) < 3:
        return None
    d, mi, se = float(nums[0]), float(nums[1]), float(nums[2])
    return d + mi / 60 + se / 3600


def _fmt_date(v):
    if v is None or v == "":
        return ""
    s = str(v).strip()
    if re.fullmatch(r"20\d{6}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return s


def parse_card(ws):
    A = label_rows(ws, 1)
    D = label_rows(ws, 4)
    d = {}

    def bA(name):  # A라벨 → B값
        r = A.get(name)
        return _s(ws.cell(r, 2).value) if r else ""

    def eD(name):  # D라벨 → E값
        r = D.get(name)
        return _s(ws.cell(r, 5).value) if r else ""

    d["관리번호"] = bA("관리번호")
    d["후보지명"] = eD("후보지명")
    d["도엽번호"] = bA("도엽번호")
    d["도엽명"] = eD("도엽명")
    d["위도"] = bA("위도")          # '위도(십진)' → norm '위도'
    d["경도"] = eD("경도")          # '경도(십진)' → norm '경도'
    d["표고"] = bA("표고").replace(" m", "").replace("m", "").strip()
    d["유형"] = eD("후보지 유형")
    d["지번주소"] = bA("지번 주소")
    d["도로명주소"] = bA("도로명 주소")
    d["소유권"] = bA("소유권 조사")
    d["관할본부"] = eD("관할 본부")
    d["조사대상"] = bA("조사대상")
    d["조사일"] = _fmt_date(ws.cell(D["조사일"], 5).value) if "조사일" in D else ""
    d["조사자"] = bA("조사자")
    d["기상"] = eD("기상상태")
    d["연계기존점"] = bA("연계 기존점")
    d["접근경로"] = bA("접근 경로").replace("\n", " ")
    d["차량진입"] = bA("차량 진입")
    d["유의사항"] = bA("유의사항")
    d["도상간섭"] = bA("도상 간섭요인").replace("\n", " ")

    # ── 자기교란 8항목: D=존재여부, E=실측이격 ──
    dist = {}
    for lab, key in DIST_ITEMS:
        r = A.get(lab)
        if r:
            dist[key] = (_s(ws.cell(r, 4).value), _s(ws.cell(r, 5).value))
    d["자기교란"] = dist
    rb = A.get("방위표지")
    d["방위표지"] = _s(ws.cell(rb, 4).value) if rb else ""   # 가능/불가

    # ── 종합판정 규칙 재계산 (캐시 수식값에 의존하지 않음) ──
    hits, missing = [], 0
    for lab, key in DIST_ITEMS:
        ex, gap = dist.get(key, ("", ""))
        if ex == "":
            missing += 1
        elif ex == "있음":
            hits.append((key, gap))
    if d["방위표지"] == "":
        missing += 1
    bang_bad = d["방위표지"] == "불가"
    nbad = len(hits) + (1 if bang_bad else 0)
    if missing > 0:
        verdict = "조사 미완료"
    elif nbad == 0:
        verdict = "적합"
    elif nbad <= 2:
        verdict = "조건부 적합"
    else:
        verdict = "부적합"
    d["종합판정"] = verdict
    d["부적합수"] = nbad
    d["교란hits"] = hits
    d["방위불가"] = bang_bad
    d["판정의견"] = bA("판정 의견")

    # ── 검토 단계 수동 재판정(원본 카드는 불변, 취합·검토·총괄·웹에만 반영) ──
    if d["관리번호"] in JUDGMENT_OVERRIDE:
        verdict_ov, reason = JUDGMENT_OVERRIDE[d["관리번호"]]
        d["종합판정"] = verdict_ov
        d["재판정"] = reason
        note = d["판정의견"].strip()
        d["판정의견"] = (note + "  ※재판정: " + reason) if note else reason
    else:
        d["재판정"] = ""

    # ── 방위표지 좌표 상세 (③) ──
    # A맵의 '위도'는 A5(위도(십진))와 충돌하므로 '경도'(A34) 앵커에서 offset 으로 읽는다.
    r_lon = A.get("경도")   # 34
    detail = {}
    d["기준점ll"] = d["표지1ll"] = d["표지2ll"] = None
    if r_lon:
        def coord(col):   # col 2=기준점(B), 4=표지1(D), 6=표지2(F)
            return {"경도": _s(ws.cell(r_lon, col).value),
                    "위도": _s(ws.cell(r_lon + 1, col).value),
                    "방위각": _s(ws.cell(r_lon + 3, col).value),
                    "거리": _s(ws.cell(r_lon + 4, col).value)}
        detail = {"기준점": coord(2), "표지1": coord(4), "표지2": coord(6)}

        def ll(col):   # (경도, 위도) 십진 or None. 경도/위도 칸 스왑 입력 자동 교정.
            a = dms2dd(ws.cell(r_lon, col).value)      # '경도' 라벨 칸
            b = dms2dd(ws.cell(r_lon + 1, col).value)  # '위도' 라벨 칸
            if a is None or b is None:
                return None
            # 한반도 범위(경도 124~132 · 위도 33~40)로 판별 — 본부별 칸 뒤바뀜 대응
            if 124 <= a <= 132 and 33 <= b <= 40:
                return (a, b)
            if 124 <= b <= 132 and 33 <= a <= 40:
                return (b, a)
            return None   # 범위 밖이면 무효
        d["기준점ll"] = ll(2)
        d["표지1ll"] = ll(4)
        d["표지2ll"] = ll(6)
    d["방위표지상세"] = detail
    az1 = detail.get("표지1", {}).get("방위각", "")
    d["방위표지좌표"] = "입력" if az1 and az1 != "-" else "미입력"
    fix_sheet(d)      # 도엽을 후보지 좌표로 재판정(카드 기재값은 별도 보존)
    fix_azimuth(d)    # 방위각 역방위·오기입 정정 + 거리 좌표 계산
    return d


# ── 카드 임베드 사진 추출 (중심·동·서·남·북·교란원) ──────────────────────────
def _photo_slot(row, col):
    """anchor (row,col) 0-index → 사진 슬롯. 지도(col>=7)·해당없음은 None."""
    if col >= 7:
        return None
    if col in (0, 1):   # A·B 열
        if 28 <= row <= 31:
            return "교란원"
        if 43 <= row <= 45:
            return "중심"
        if 46 <= row <= 48:
            return "동"
        if 49 <= row <= 52:
            return "남"
    if col in (3, 4):   # D·E 열
        if 46 <= row <= 48:
            return "서"
        if 49 <= row <= 52:
            return "북"
    return None


def extract_photos(ws, code, outdir, maxpx=300, quality=75):
    """카드 시트의 방위/중심 사진을 다운스케일 JPEG 로 저장. {슬롯: 파일명} 반환."""
    import io as _io
    from PIL import Image
    outdir.mkdir(parents=True, exist_ok=True)
    got = {}
    for im in getattr(ws, "_images", []):
        try:
            a = im.anchor._from
        except AttributeError:
            continue
        slot = _photo_slot(a.row, a.col)
        if not slot or slot in got:
            continue
        try:
            img = Image.open(_io.BytesIO(im._data())).convert("RGB")
        except Exception:   # noqa: BLE001
            continue
        img.thumbnail((maxpx, maxpx))
        fname = f"{code}_{slot}.jpg"
        img.save(outdir / fname, "JPEG", quality=quality, optimize=True)
        got[slot] = fname
    return got


# ── 도엽(1:50,000) 좌표 판정 ──────────────────────────────────────────────────
# ⚠ 배포된 카드의 도엽번호·도엽명은 **연계 기존점의 도엽**이 찍혀 있다
#   (make_survey_card_v2.py 가 sheet_of() 에 기존점 좌표를 넘긴 버그. 2026-08 정정).
#   회신본은 이미 그 값으로 굳었으므로, 취합 단계에서 **후보지 자신의 좌표**로
#   다시 판정해 덮어쓴다. 원본 카드는 불변이고 기재값은 '도엽번호기재'로 보존한다.
#   도엽 폴리곤은 전부 경위선 직교 직사각형(15'×15')이라 bbox 판정으로 충분하다.
TOPO_GEOJSON = ROOT / "docs" / "data" / "topo_sheets.geojson"
_TOPO_CELLS = None


def _topo_cells():
    """[(w, s, e, n, sheet_code, sheet_name, mapidcd)] — 최초 1회 로드."""
    global _TOPO_CELLS
    if _TOPO_CELLS is None:
        cells = []
        if TOPO_GEOJSON.exists():
            import json
            fc = json.load(open(TOPO_GEOJSON, encoding="utf-8"))
            for ft in fc["features"]:
                ring = ft["geometry"]["coordinates"][0]
                xs = [c[0] for c in ring]
                ys = [c[1] for c in ring]
                p = ft["properties"]
                cells.append((min(xs), min(ys), max(xs), max(ys),
                              str(p.get("sheet_code") or ""),
                              str(p.get("sheet_name") or ""),
                              str(p.get("sheet_mapidcd") or "")))
        _TOPO_CELLS = cells
    return _TOPO_CELLS


def sheet_at(lat, lon):
    """좌표가 들어있는 1:50,000 도엽 → (도엽번호, 도엽명, NGII번호). 없으면 None."""
    try:
        lat, lon = float(lat), float(lon)
    except (TypeError, ValueError):
        return None
    for w, s_, e, n, code, name, mid in _topo_cells():
        if w <= lon <= e and s_ <= lat <= n:
            return code, name, mid
    return None


def fix_sheet(d):
    """카드 1건의 도엽 기재값을 좌표 판정 결과로 정정(제자리 수정)."""
    d["도엽번호기재"] = d.get("도엽번호", "")
    d["도엽명기재"] = d.get("도엽명", "")
    d["도엽정정"] = ""
    hit = sheet_at(d.get("위도"), d.get("경도"))
    if not hit:
        return d
    code, name, _mid = hit
    if (d["도엽번호기재"] or "").strip() != code or (d["도엽명기재"] or "").strip() != name:
        d["도엽정정"] = f"{d['도엽명기재'] or '-'}({d['도엽번호기재'] or '-'}) → {name}({code})"
    d["도엽번호"], d["도엽명"] = code, name
    return d


# ── 방위각·거리 정정 ──────────────────────────────────────────────────────────
# 카드 기재 방위각 196쌍 중 29쌍이 좌표와 5° 넘게 어긋나고 그중 16쌍이 정확히
# ±180°(역방위 — 표지→기준점 방향을 적음)다. 좌표 맞바꿈이 아님은 확인됐다
# (맞바꿈이면 표지 좌표가 후보지 자리에 와야 하나 최소 22m, DS-004 는 기준점이
# 후보지와 0m 로 일치). 즉 **좌표는 옳고 카드 숫자만 뒤집혔다**.
#
# 거리도 카드값이 100~117m 에 몰려 있어(좌표상 실제 18~75m) 측정값이 아니라
# 이격 기준 100m 충족을 뜻해 적은 것으로 보인다. 현장 성과의 거리는 **기준점과
# 방위표지 사이**를 재는 것이므로 좌표 계산 측지거리가 정답이다.
#
# 원본 카드는 불변. 취합·검토·총괄·웹에만 반영하고 기재값은 '*기재' 로 보존한다.
AZ_TOL_DEG = 5.0        # 이 이내면 카드 방위각을 그대로 둔다(측정값이 GPS 좌표보다 정밀)
DIST_TOL_M = 5.0        # 거리 정정 문턱 — 절대 5m 이고
DIST_TOL_RATIO = 0.10   # 동시에 상대 10% 를 넘을 때만 정정(반올림 차이 무시)


def az_from_ll(a, b):
    """a→b 진북기준 방위각(도). **북=0°, 시계방향** (§19 진북 기준).

    a, b = (경도, 위도). 좌표 없으면 None.
    """
    if not a or not b:
        return None
    lon1, lat1 = a
    lon2, lat2 = b
    la = math.radians((lat1 + lat2) / 2)
    m_lat = 111132.92 - 559.82 * math.cos(2 * la) + 1.175 * math.cos(4 * la)
    m_lon = 111412.84 * math.cos(la) - 93.5 * math.cos(3 * la)
    return math.degrees(math.atan2((lon2 - lon1) * m_lon,
                                   (lat2 - lat1) * m_lat)) % 360


def az_deg(v):
    """카드 방위각 문자열('ddd°mm′ss″' 등) → 도(십진). 실패 시 None."""
    nums = re.findall(r"\d+\.?\d*", str(v or ""))
    if not nums:
        return None
    out = float(nums[0])
    if len(nums) >= 2:
        out += float(nums[1]) / 60
    if len(nums) >= 3:
        out += float(nums[2]) / 3600
    return out % 360


def fmt_dms(deg):
    """도(십진) → 'ddd°mm′ss″' (카드 표기 형식)."""
    if deg is None:
        return ""
    deg %= 360
    d = int(deg)
    m = int((deg - d) * 60)
    sec = (deg - d - m / 60) * 3600
    if round(sec) >= 60:      # 초 반올림이 60 이 되면 자리올림
        sec, m = 0.0, m + 1
    if m >= 60:
        m, d = 0, (d + 1) % 360
    return f"{d}°{m}′{sec:.0f}″"


def fix_azimuth(d):
    """카드 1건의 방위표지 방위각·거리를 좌표 기준으로 정정(제자리 수정).

    방위각 — 차이 5° 이내면 카드값 유지(측정 방위각이 휴대GPS 좌표보다 정밀),
    ±180°(±5°) 면 **카드값 +180°** 로 뒤집어 측정 정밀도를 살리고, 그 밖의
    불일치는 복원할 원값이 없으므로 **좌표 방위각**으로 대체한다.
    거리 — 항상 기준점↔표지 좌표 측지거리를 쓴다(문턱 초과 시 정정 표시).
    """
    det = d.get("방위표지상세") or {}
    base = d.get("기준점ll")
    notes = []
    for llk, tag in (("표지1ll", "표지1"), ("표지2ll", "표지2")):
        c = det.get(tag)
        if not c:
            continue
        c.setdefault("방위각기재", c.get("방위각", ""))
        c.setdefault("거리기재", c.get("거리", ""))
        c["방위각정정"] = c["거리정정"] = ""
        mll = d.get(llk)
        geo = az_from_ll(base, mll)
        if geo is None:
            continue
        card = az_deg(c["방위각기재"])
        if card is None:
            c["방위각"], c["방위각정정"] = fmt_dms(geo), "좌표 산출(카드 미기재)"
        else:
            diff = abs((geo - card + 180) % 360 - 180)
            if abs(diff - 180) <= AZ_TOL_DEG:
                c["방위각"] = fmt_dms(card + 180)
                c["방위각정정"] = "역방위 +180°"
            elif diff > AZ_TOL_DEG:
                c["방위각"] = fmt_dms(geo)
                c["방위각정정"] = f"좌표 재계산(카드와 {diff:.0f}° 차)"
        if c["방위각정정"]:
            notes.append(f"{tag} 방위각 {c['방위각기재'] or '-'} → {c['방위각']}"
                         f" ({c['방위각정정']})")
        # 거리 — 기준점↔표지 좌표 측지거리
        geo_m = _geo_dist(base, mll)
        if geo_m is not None:
            card_m = _card_dist(det, tag)
            c["거리"] = f"{geo_m:.0f}"
            if (card_m is not None and abs(card_m - geo_m) > DIST_TOL_M
                    and abs(card_m - geo_m) > geo_m * DIST_TOL_RATIO):
                c["거리정정"] = f"카드 {card_m:.0f}m → 좌표 {geo_m:.0f}m"
                notes.append(f"{tag} 거리 {c['거리정정']}")
    d["방위정정"] = " · ".join(notes)

    # 후보지(도상) ↔ 현장 기준점 이격 — 오류가 아니라 **접근성 지표**다.
    # 후보지는 현장조사 전 도상에서 고른 점이고, 성과의 기준은 현장 기준점이다.
    # 이격이 크다는 것은 도상 후보지로 진입이 어려워 옮겼다는 뜻이다.
    try:
        cand = (float(d["경도"]), float(d["위도"]))
    except (TypeError, ValueError):
        cand = None
    d["후보지이격"] = _geo_dist(base, cand) if (base and cand) else None
    return d


def parse_workbook(path, photo_dir=None):
    wb = load_workbook(path, data_only=True)
    out = []
    for name in wb.sheetnames:
        if name == "총괄 목록":
            continue
        ws = wb[name]
        d = parse_card(ws)
        if d.get("관리번호"):
            d["_src"] = path.name
            if photo_dir is not None:
                d["사진"] = extract_photos(ws, d["관리번호"], Path(photo_dir))
            out.append(d)
    wb.close()
    return out


# ── 검토(선점 가능성) 파생 ────────────────────────────────────────────────────
# ======================================================================
# 평탄화 취합본에서 되읽기 — 카드 회신본이 없을 때의 재생성 경로
# ======================================================================
#
# `parse_workbook()` 은 본부별 **카드형 회신본**을 읽는다. 그 원본이 없을 때
# (배포 저장소에는 없다) `*_현장조사_일괄취합_103건.xlsx` 로 같은 레코드를
# 되만들어 `make_survey_map.py` 를 재생성할 수 있다.
#
# ⚠️ 취합본에 **없는 것 둘**을 복원해야 한다:
#   · 사진      — `docs/survey_photos/{관리번호}_{슬롯}.jpg` 파일 존재로 재구성
#   · 기준점 좌표 — 방위표지 좌표 + 방위각 + 거리로 **역산**(표지에서 방위각+180°
#                  방향으로 거리만큼). 표지1·표지2 두 해가 서로 맞는지 대조한다.
#
# ⚠️ 이것은 **카드 원본의 대체가 아니다.** 카드에만 있는 항목(원 기재 방위각의
#    일부 맥락 등)은 취합 단계에서 이미 정규화된 값으로만 남는다. 회신본이
#    확보되면 `parse_workbook()` 경로로 돌아가는 것이 옳다.

AGG_SHEET = "일괄취합"
AGG_HEADER_ROW = 1          # 0-based — 0행은 제목 띠

# 취합본 열 → 레코드 키
AGG_COLMAP = {
    "관할 본부": "관할본부", "관리번호": "관리번호", "후보지명": "후보지명",
    "도엽번호": "도엽번호기재", "도엽명": "도엽명기재",
    "표고(m)": "표고", "유형": "유형",
    "지번 주소": "지번주소", "도로명 주소": "도로명주소",
    "연계 기존점": "연계기존점", "소유권": "소유권",
    "조사대상": "조사대상", "조사일": "조사일", "조사자": "조사자",
    "기상": "기상", "판정 의견": "판정의견",
    "접근 경로": "접근경로", "차량 진입": "차량진입",
    "유의사항": "유의사항", "도상 간섭요인": "도상간섭",
}


def _agg_backsolve_base(ll, az_deg, dist_m):
    """표지 좌표 + (기준점→표지) 방위각·거리 → 기준점 (경도, 위도). 실패 시 None."""
    import math

    if ll is None:
        return None
    # ⚠ 취합본의 방위각은 DMS 문자열(`162° 55′ 36.41″`)이다 — float() 로는 못 읽는다
    az = dms2dd(az_deg)
    if az is None:
        try:
            az = float(str(az_deg).strip())
        except (TypeError, ValueError):
            return None
    try:
        d = float(str(dist_m).strip())
    except (TypeError, ValueError):
        return None
    if d <= 0:
        return None
    lon, lat = ll
    back = math.radians((az + 180.0) % 360.0)
    dn = d * math.cos(back) / 110574.0
    de = d * math.sin(back) / (111320.0 * math.cos(math.radians(lat)))
    return (lon + de, lat + dn)


def load_aggregate(path, photo_dir=None):
    """
    평탄화 취합본 → `parse_workbook()` 과 같은 레코드 리스트.

    Parameters
    ----------
    path      : `*_현장조사_일괄취합_103건.xlsx`
    photo_dir : `docs/survey_photos` — 있으면 파일명 규약으로 사진을 되붙인다.
    """
    import pandas as pd

    df = pd.read_excel(path, sheet_name=AGG_SHEET,
                       header=AGG_HEADER_ROW, engine="openpyxl")
    photo_dir = Path(photo_dir) if photo_dir else None
    have_ph = set(f.name for f in photo_dir.glob("*.jpg")) if photo_dir else set()

    def S(v):
        if v is None or (isinstance(v, float) and v != v):
            return ""
        return str(v).strip()

    recs, base_gap = [], []
    for _, r in df.iterrows():
        if not S(r.get("관리번호")):
            continue
        d = {k2: S(r.get(k1)) for k1, k2 in AGG_COLMAP.items()}
        d["위도"] = S(r.get("위도"))
        d["경도"] = S(r.get("경도"))

        # ── 자기교란 8항목 (존재 / 이격) ──
        dist = {}
        for lab, key in DIST_ITEMS:
            dist[key] = (S(r.get(f"{key} 존재")), S(r.get(f"{key} 이격")))
        d["자기교란"] = dist
        d["방위표지"] = S(r.get("방위표지 가능"))

        # ── 종합판정은 **규칙으로 재계산** (저장된 열을 믿지 않는다) ──
        hits, missing = [], 0
        for lab, key in DIST_ITEMS:
            ex, gap = dist.get(key, ("", ""))
            if ex == "":
                missing += 1
            elif ex == "있음":
                hits.append((key, gap))
        if d["방위표지"] == "":
            missing += 1
        bang_bad = d["방위표지"] == "불가"
        nbad = len(hits) + (1 if bang_bad else 0)
        d["종합판정"] = ("조사 미완료" if missing else
                        "적합" if nbad == 0 else
                        "조건부 적합" if nbad <= 2 else "부적합")
        d["부적합수"] = nbad
        d["교란hits"] = hits
        d["방위불가"] = bang_bad

        if d["관리번호"] in JUDGMENT_OVERRIDE:
            v_ov, reason = JUDGMENT_OVERRIDE[d["관리번호"]]
            d["종합판정"] = v_ov
            d["재판정"] = reason
            note = d["판정의견"].strip()
            d["판정의견"] = (note + "  ※재판정: " + reason) if note else reason
        else:
            d["재판정"] = ""

        # ── 방위표지 좌표·상세 ──
        def ll_of(tag):
            a, b = S(r.get(f"{tag} 경도")), S(r.get(f"{tag} 위도"))
            a, b = dms2dd(a), dms2dd(b)
            if a is None or b is None:
                return None
            if 124 <= a <= 132 and 33 <= b <= 40:
                return (a, b)
            if 124 <= b <= 132 and 33 <= a <= 40:
                return (b, a)
            return None

        d["표지1ll"], d["표지2ll"] = ll_of("방위1"), ll_of("방위2")
        detail = {}
        for tag, key in (("방위1", "표지1"), ("방위2", "표지2")):
            detail[key] = {
                "경도": S(r.get(f"{tag} 경도")), "위도": S(r.get(f"{tag} 위도")),
                "방위각": S(r.get(f"{tag} 방위각")), "거리": S(r.get(f"{tag} 거리(m)")),
                "방위각기재": S(r.get(f"{tag} 방위각(카드)")),
                "거리기재": S(r.get(f"{tag} 거리(카드)")),
            }
        detail["기준점"] = {"경도": "", "위도": "", "방위각": "", "거리": ""}
        d["방위표지상세"] = detail
        d["방위정정"] = S(r.get("방위 정정"))
        d["방위표지좌표"] = ("입력" if detail["표지1"]["방위각"] not in ("", "-")
                            else "미입력")

        # ── 기준점 좌표는 취합본에 없다 → 방위표지에서 역산 ──
        b1 = _agg_backsolve_base(d["표지1ll"], detail["표지1"]["방위각"],
                                 detail["표지1"]["거리"])
        b2 = _agg_backsolve_base(d["표지2ll"], detail["표지2"]["방위각"],
                                 detail["표지2"]["거리"])
        if b1 and b2:
            base_gap.append((d["관리번호"], _geo_dist(b1, b2)))
        d["기준점ll"] = b1 or b2
        if d["기준점ll"]:
            detail["기준점"] = {"경도": f"{d['기준점ll'][0]:.7f}",
                                "위도": f"{d['기준점ll'][1]:.7f}",
                                "방위각": "", "거리": ""}

        # 수치로 둔다 — 팝업이 `gap >= 50` 로 비교한다
        try:
            d["후보지이격"] = float(r.get("후보지↔기준점(m)"))
        except (TypeError, ValueError):
            d["후보지이격"] = None

        # ── 사진은 파일명 규약으로 ──
        ph = {}
        for slot in ("중심", "동", "서", "남", "북"):
            fn = f"{d['관리번호']}_{slot}.jpg"
            if fn in have_ph:
                ph[slot] = fn
        d["사진"] = ph

        d["_src"] = S(r.get("회신 파일")) or Path(path).name
        fix_sheet(d)        # 도엽을 후보지 좌표로 재판정 (카드값은 기재 열에 보존)
        recs.append(d)

    if base_gap:
        import statistics as _st
        g = [v for _, v in base_gap if v == v]
        if g:
            print(f"    기준점 역산 교차검증 {len(g)}건 — "
                  f"표지1·2 해 차이 중앙 {_st.median(g):.1f} m · "
                  f"최대 {max(g):.1f} m")
    return recs


def key_disturb(d):
    parts = []
    for key, gap in d["교란hits"]:
        parts.append(f"{key}({gap}m)" if gap and gap != "-" else key)
    if d["방위불가"]:
        parts.append("방위표지 불가")
    return ", ".join(parts)


MARK_MIN_DIST = 100.0   # 방위표지 최소 이격(m). 둘 중 하나라도 이 이상이면 선점 가능.


def _geo_dist(a, b):
    """두 (경도, 위도) 점 사이 측지 거리(m). WGS84 국소 평면 근사(한반도 위도)."""
    if not a or not b:
        return None
    lon1, lat1 = a
    lon2, lat2 = b
    la = math.radians((lat1 + lat2) / 2)
    m_lat = 111132.92 - 559.82 * math.cos(2 * la) + 1.175 * math.cos(4 * la)
    m_lon = 111412.84 * math.cos(la) - 93.5 * math.cos(3 * la)
    return math.hypot((lon2 - lon1) * m_lon, (lat2 - lat1) * m_lat)


def _card_dist(det, tag):
    """카드 기재 거리(m) 파싱. 없으면 None."""
    s = det.get(tag, {}).get("거리")
    if not s or s == "-":
        return None
    nums = re.findall(r"\d+\.?\d*", str(s))
    return float(nums[0]) if nums else None


def mark_max_dist(d):
    """방위표지 1·2 중 최장 이격(m).

    ⚠ 카드 기재 '거리' 는 오기입 사례가 있어(경남 DS-004·045·084 등 좌표상 실제
    거리의 4~5배로 적힘) 신뢰도가 낮다. 따라서 **기준점·표지 좌표로 계산한 거리를
    우선**하고, 좌표가 없는 표지만 카드 기재값으로 폴백한다. 방위각은 좌표·카드가
    일치하므로 방향은 신뢰 가능, 거리 크기만 좌표를 신뢰한다.
    """
    base = d.get("기준점ll")
    det = d.get("방위표지상세", {})
    best = None
    for ll_key, tag in (("표지1ll", "표지1"), ("표지2ll", "표지2")):
        v = _geo_dist(base, d.get(ll_key))     # 좌표 우선
        if v is None:                           # 좌표 없으면 카드값 폴백
            v = _card_dist(det, tag)
        if v is not None:
            best = v if best is None else max(best, v)
    return best


def _sheet_key(d):
    """도엽 식별자 — 도엽번호 우선, 없으면 도엽명."""
    return (d.get("도엽번호") or "").strip() or (d.get("도엽명") or "").strip()


def sheet_priority(recs):
    """도엽(1:50,000) 단위로 **B 후보를 대표/예비로 구분**한다.

    지형도 자침편각 표기는 **도엽당 1점**이면 충분하므로, 같은 도엽에 후보가 겹치면
    전부를 유효 후보로 세지 않는다.

      대표 : 그 도엽에 A 가 없고, 도엽 내 B 중 **방위표지 최장거리 1순위** → A 승격 1순위
      예비 : ① 도엽에 이미 A 확보(그 도엽은 해결됨)  ② 도엽 중복 후순위

    반환 {관리번호: (구분, 순위, 도엽내B수, 사유)} — 구분은 "대표"/"예비".
    """
    from collections import defaultdict
    by = defaultdict(list)
    for d in recs:
        k = _sheet_key(d)
        if k:
            by[k].append(d)
    a_sheets = {k for k, l in by.items() if any(review(d)[0] == "A" for d in l)}
    out = {}
    for k, lst in by.items():
        bs = sorted([d for d in lst if review(d)[0] == "B"],
                    key=lambda d: -(mark_max_dist(d) or 0))
        n = len(bs)
        for i, d in enumerate(bs, 1):
            if k in a_sheets:
                out[d["관리번호"]] = ("예비", i, n, "도엽에 A 확보됨")
            elif i == 1:
                out[d["관리번호"]] = ("대표", i, n, "도엽 대표" if n > 1 else "도엽 단독")
            else:
                out[d["관리번호"]] = ("예비", i, n, "도엽 중복 후순위")
    return out


def review(d):
    """(등급, 선점 검토 결론, 검토 의견) — 웹 표출용. **4단계 분류**.

    A 선점 가능        : 자기 청정(적합) + 방위표지 좌표거리 하나라도 ≥100m
    B 조건부 선점 가능  : 자기 청정(적합)이나 방위표지 최장 <100m
                       → 자기적으로 양호. 정밀 방위각(천문·자이로·RTK) 또는 ≥100m 표지
                         확보 시 선점 가능. 새 측점을 찾는 것보다 방위 기준 정밀화가 저비용.
    C 현장 확인 필요    : 조건부 적합(자기교란 재확인 필요)
    D 부적합           : 부적합 또는 방위표지 설치 불가(대체 후보지 검토)
    미완료             : 조사 항목 누락

    거리는 카드 기재값이 아니라 좌표 계산거리(`mark_max_dist`)를 쓴다(카드값 오기입 사례).
    """
    v = d["종합판정"]
    kd = key_disturb(d)
    op = d["판정의견"].replace("\n", " ").strip()
    if v == "조사 미완료":
        return "미완료", "재조사 필요 — 조사 항목 누락", op or "조사 항목 미입력"
    if d["방위불가"]:
        return "D", "부적합 — 방위표지 설치 불가로 대체 후보지 검토", op or "방위표지 설치 불가"
    if v == "적합":
        md = mark_max_dist(d)
        if md is not None and md < MARK_MIN_DIST:
            concl = (f"조건부 선점 가능 — 방위표지 거리 확보 필요"
                     f" (현 최장 {md:.0f}m<{MARK_MIN_DIST:.0f}m)")
            note = op or (f"자기교란 없음 · 방위표지 최장 이격 {md:.0f}m — 정밀 방위각"
                          f"(천문·자이로·RTK) 또는 ≥{MARK_MIN_DIST:.0f}m 표지 확보 시 선점 가능")
            return "B", concl, note
        return "A", "선점 가능 — 자기구배 조사 진행", op or "주변 자기교란 요소 없음"
    if v == "조건부 적합":
        return "C", f"현장 확인 필요 — 자기교란 재확인 후 결정 ({kd})", op or kd
    return "D", f"부적합 — 대체 후보지 검토 ({kd})", op or kd


# ── 스타일 ────────────────────────────────────────────────────────────────────
F_TITLE = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")
F_HDR = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
F_VAL = Font(name="맑은 고딕", size=9)
F_BOLD = Font(name="맑은 고딕", size=9, bold=True)
FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_HDR = PatternFill("solid", fgColor="4472C4")
FILL_BAD = PatternFill("solid", fgColor="FDE0DC")
FILL_COND = PatternFill("solid", fgColor="FFF2CC")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FILL_BLUE = PatternFill("solid", fgColor="DDEBF7")   # B 조건부 선점 가능
FILL_GRAY = PatternFill("solid", fgColor="EDEDED")
# 4단계: A 선점가능(녹) · B 조건부 선점가능(청) · C 현장확인(황) · D 부적합(적)
GRADE_FILL = {"A": FILL_OK, "B": FILL_BLUE, "C": FILL_COND, "D": FILL_BAD, "미완료": FILL_GRAY}
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _title(ws, ncol, text):
    ws.merge_cells(f"A1:{get_column_letter(ncol)}1")
    c = ws["A1"]
    c.value = text
    c.font = F_TITLE
    c.fill = FILL_TITLE
    c.alignment = AL_L
    ws.row_dimensions[1].height = 26


def _header(ws, cols, row=2):
    for j, h in enumerate(cols, 1):
        c = ws.cell(row, j, h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDER


def dcell(item):
    if not item:
        return ""
    ex, gap = item
    if ex == "있음":
        return f"있음 ({gap}m)" if gap and gap != "-" else "있음"
    return ex


# ── 시트 1: 현장조사 취합 ────────────────────────────────────────────────────
FULL_COLS = ["관할 본부", "관리번호", "후보지명", "도엽번호", "도엽명", "도엽 정정",
             "방위 정정", "후보지↔기준점(m)",
             "지번 주소",
             "위도", "경도", "표고(m)", "유형", "연계 기존점", "소유권",
             "조사일", "조사자", "기상",
             "차량영향", "전력시설", "통신시설", "금속류", "매설물", "건축물", "자연환경",
             "방위표지", "방위표지좌표", "부적합 건수", "종합 판정", "판정 의견",
             "접근 경로", "유의사항", "도상 간섭요인", "회신 파일"]


def sheet_full(wb, recs):
    ws = wb.create_sheet("현장조사 취합")
    n = len(FULL_COLS)
    _title(ws, n, f"지자기 도상선점 — 현장 조사서 취합 ({len(recs)}건)   ·   {datetime.now():%Y-%m-%d}")
    _header(ws, FULL_COLS)
    ds = FULL_COLS.index("차량영향") + 1
    de = FULL_COLS.index("자연환경") + 1
    bcol = FULL_COLS.index("방위표지") + 1
    jcol = FULL_COLS.index("종합 판정") + 1
    for ri, d in enumerate(recs, 3):
        dist = d["자기교란"]
        row = [d["관할본부"], d["관리번호"], d["후보지명"], d["도엽번호"], d["도엽명"],
               d["도엽정정"],
               d["방위정정"],
               (round(d["후보지이격"]) if d["후보지이격"] is not None else ""),
               d["지번주소"], d["위도"], d["경도"], d["표고"], d["유형"], d["연계기존점"],
               d["소유권"], d["조사일"], d["조사자"], d["기상"],
               dcell(dist.get("차량영향")), dcell(dist.get("전력시설")), dcell(dist.get("통신시설")),
               dcell(dist.get("금속류")), dcell(dist.get("매설물")), dcell(dist.get("건축물")),
               dcell(dist.get("자연환경")), d["방위표지"], d["방위표지좌표"], d["부적합수"],
               d["종합판정"], d["판정의견"].replace("\n", " "),
               d["접근경로"], d["유의사항"], d["도상간섭"], d["_src"]]
        for j, v in enumerate(row, 1):
            c = ws.cell(ri, j, v)
            c.font = F_VAL
            c.border = BORDER
            c.alignment = AL_L if FULL_COLS[j - 1] in (
                "지번 주소", "도엽 정정", "방위 정정", "판정 의견", "접근 경로",
                "유의사항", "도상 간섭요인") else AL_C
            if ds <= j <= de and isinstance(v, str) and v.startswith("있음"):
                c.fill = FILL_BAD
            if j == bcol and v == "불가":
                c.fill = FILL_BAD
            if j == jcol:
                c.fill = {"적합": FILL_OK, "조건부 적합": FILL_COND,
                          "부적합": FILL_BAD}.get(v, FILL_GRAY)
    widths = [14, 8, 13, 11, 8, 22, 34, 12, 24, 9, 9, 7, 8, 9, 8, 11, 12, 6,
              9, 8, 8, 8, 8, 8, 8, 8, 9, 8, 10, 30, 24, 18, 30, 16]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}{len(recs) + 2}"


# ── 시트 2: 선점 검토 (웹 표출용) ────────────────────────────────────────────
WEB_COLS = ["등급", "도엽 구분", "관할 본부", "관리번호", "후보지명", "위도", "경도", "표고(m)",
            "종합 판정", "핵심 교란요인", "방위표지", "선점 검토 결론", "검토 의견", "조사일"]


def _rank_str(d, prio):
    """B 의 도엽 대표/예비 구분 문자열. B 가 아니면 '-'."""
    p = prio.get(d["관리번호"])
    if not p:
        return "-"
    kind, i, n, why = p
    sheet = d.get("도엽명") or d.get("도엽번호") or ""
    if kind == "대표":
        return f"★ 대표 · {sheet}" + (f" 1/{n}" if n > 1 else " 단독")
    if why == "도엽에 A 확보됨":
        return f"예비 · {sheet} (A 확보)"
    return f"예비 · {sheet} {i}/{n}"


def sheet_web(wb, recs):
    ws = wb.create_sheet("선점 검토")
    n = len(WEB_COLS)
    _title(ws, n, "지자기 도상선점 — 선점 위치 검토 (웹 표출용)")
    # 등급 범례
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    lg = ws["A2"]
    lg.value = ("등급  A = 선점 가능(자기구배 조사)   ·   B = 조건부 선점 가능(방위표지 거리 확보 필요)"
                "   ·   C = 현장 확인 필요(자기교란 재확인)   ·   D = 부적합(대체 후보지 검토)"
                "   ·   방위표지 '불가'는 D"
                "   ·   도엽 구분: 자침편각 표기는 도엽당 1점이면 충분 → B 는 도엽 "
                "대표(★, 방위표지 최장 1순위)와 예비(중복 후순위·A 확보 도엽)로 나눔")
    lg.font = Font(name="맑은 고딕", size=8.5, color="555555")
    lg.alignment = AL_L
    _header(ws, WEB_COLS, row=3)
    order = {"A": 0, "B": 1, "C": 2, "D": 3, "미완료": 4}
    prio = sheet_priority(recs)   # B 도엽 대표/예비 구분

    def skey(d):
        go = order.get(review(d)[0], 9)
        p = prio.get(d["관리번호"])
        if p:   # B 는 대표 먼저 → 도엽 → 순위
            kind, i, n, _ = p
            sheet = (d.get("도엽번호") or d.get("도엽명") or "")
            return (go, 0 if kind == "대표" else 1, sheet, i, d["관리번호"])
        return (go, 0, d["관할본부"] or "", 99, d["관리번호"])
    rr = sorted(recs, key=skey)
    for ri, d in enumerate(rr, 4):
        grade, concl, note = review(d)
        row = [grade, _rank_str(d, prio), d["관할본부"], d["관리번호"], d["후보지명"],
               d["위도"], d["경도"], d["표고"], d["종합판정"], key_disturb(d) or "-",
               d["방위표지"], concl, note, d["조사일"]]
        for j, v in enumerate(row, 1):
            c = ws.cell(ri, j, v)
            c.font = F_BOLD if j in (1, 2) else F_VAL
            c.border = BORDER
            c.alignment = AL_L if j in (10, 12, 13) else AL_C
            if j == 1:
                c.fill = GRADE_FILL.get(grade, FILL_GRAY)
            if j == 2 and isinstance(v, str) and v.startswith("★"):
                c.fill = FILL_OK
            elif j == 2 and isinstance(v, str) and v.startswith("예비"):
                c.fill = FILL_GRAY
            if j == 9:
                c.fill = {"적합": FILL_OK, "조건부 적합": FILL_COND,
                          "부적합": FILL_BAD}.get(v, FILL_GRAY)
            if j == 11 and v == "불가":
                c.fill = FILL_BAD
    widths = [6, 13, 14, 8, 14, 9, 9, 7, 10, 26, 8, 34, 40, 11]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(rr) + 3}"


# ── 시트 3: 검토 요약 ─────────────────────────────────────────────────────────
def sheet_summary(wb, recs):
    ws = wb.create_sheet("검토 요약", 0)
    _title(ws, 6, "지자기 도상선점 현장조사 — 검토 요약")
    grades = Counter(review(d)[0] for d in recs)
    verd = Counter(d["종합판정"] for d in recs)
    bang = sum(1 for d in recs if d["방위불가"])
    dist_cnt = Counter()
    for d in recs:
        for key, _ in d["교란hits"]:
            dist_cnt[key] += 1

    r = 3
    def line(label, val, fill=None, bold=False):
        nonlocal r
        a = ws.cell(r, 1, label)
        a.font = F_BOLD
        a.alignment = AL_L
        ws.merge_cells(f"A{r}:C{r}")
        b = ws.cell(r, 4, val)
        b.font = F_BOLD if bold else F_VAL
        b.alignment = AL_L
        ws.merge_cells(f"D{r}:F{r}")
        if fill:
            for cc in (a, b):
                cc.fill = fill
        r += 1

    ws.cell(r, 1, "■ 전체 현황").font = F_BOLD
    r += 1
    line("총 조사 후보지", f"{len(recs)} 건", bold=True)
    _prio = sheet_priority(recs)
    _rep = sum(1 for v in _prio.values() if v[0] == "대표")
    _sp_a = sum(1 for v in _prio.values() if v[0] == "예비" and v[3] == "도엽에 A 확보됨")
    _sp_d = sum(1 for v in _prio.values() if v[0] == "예비" and v[3] == "도엽 중복 후순위")
    line("선점 가능 (등급 A)", f"{grades['A']} 건", FILL_OK)
    line("조건부 선점 가능 (등급 B — 도엽 대표)", f"{_rep} 건", FILL_BLUE)
    line("    └ 예비: 도엽 중복 후순위", f"{_sp_d} 건", FILL_GRAY)
    line("    └ 예비: 도엽에 A 이미 확보", f"{_sp_a} 건", FILL_GRAY)
    line("현장 확인 필요 (등급 C)", f"{grades['C']} 건", FILL_COND)
    line("부적합 (등급 D)", f"{grades['D']} 건", FILL_BAD)
    if grades.get("미완료"):
        line("조사 미완료", f"{grades['미완료']} 건", FILL_GRAY)
    r += 1
    ws.cell(r, 1, "■ 종합 판정 분포 (조사자 카드 기준)").font = F_BOLD
    r += 1
    for k in ("적합", "조건부 적합", "부적합", "조사 미완료"):
        if verd.get(k):
            line(k, f"{verd[k]} 건")
    r += 1
    ws.cell(r, 1, "■ 특이사항").font = F_BOLD
    r += 1
    line("방위표지 설치 불가", f"{bang} 건", FILL_BAD if bang else None)
    r += 1
    ws.cell(r, 1, "■ 주요 교란요인 (있음 건수)").font = F_BOLD
    r += 1
    for key, c in dist_cnt.most_common():
        line(key, f"{c} 건")
    r += 1
    ws.cell(r, 1, "■ 본부별 등급 분포").font = F_BOLD
    r += 1
    hqs = sorted({d["관할본부"] for d in recs})
    hdr = ["관할 본부", "A", "B", "C", "D", "미완료", "계"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r, j, h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDER
    r += 1
    for hq in hqs:
        sub = [d for d in recs if d["관할본부"] == hq]
        g = Counter(review(d)[0] for d in sub)
        vals = [hq, g["A"], g["B"], g["C"], g["D"], g.get("미완료", 0), len(sub)]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.font = F_VAL
            c.alignment = AL_L if j == 1 else AL_C
            c.border = BORDER
        r += 1
    for j, w in enumerate([16, 6, 6, 6, 6, 8, 6], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── 도엽별 B 대표/예비 (자침편각 표기는 도엽당 1점) ──
    from collections import defaultdict
    sheets = defaultdict(list)
    for d in recs:
        if d["관리번호"] in _prio:
            sheets[f"{d.get('도엽명') or ''}({d.get('도엽번호') or ''})"].append(d)
    r += 1
    ws.cell(r, 1, "■ 도엽별 B 대표/예비 (★=도엽 대표 · 방위표지 최장거리 1순위)").font = F_BOLD
    r += 1
    hdr2 = ["도엽 (1:50,000)", "구분", "관리번호", "후보지명", "방위표지 최장(m)", "사유"]
    for j, h in enumerate(hdr2, 1):
        c = ws.cell(r, j, h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDER
    r += 1
    for key in sorted(sheets):
        lst = sorted(sheets[key], key=lambda d: -(mark_max_dist(d) or 0))
        for i, d in enumerate(lst, 1):
            kind, rk, n, why = _prio[d["관리번호"]]
            top = (kind == "대표")
            vals = [key if i == 1 else "", ("★ 대표" if top else f"예비 {rk}/{n}"),
                    d["관리번호"], d["후보지명"], f"{mark_max_dist(d) or 0:.0f}", why]
            for j, v in enumerate(vals, 1):
                c = ws.cell(r, j, v)
                c.font = F_BOLD if top and j in (2, 3) else F_VAL
                c.alignment = AL_L if j in (1, 4, 6) else AL_C
                c.border = BORDER
                c.fill = FILL_OK if top else FILL_GRAY
            r += 1
    for j, w in enumerate([18, 10, 10, 16, 14, 18], 1):
        ws.column_dimensions[get_column_letter(j)].width = max(
            ws.column_dimensions[get_column_letter(j)].width or 0, w)


# ── 일괄취합(플랫) 워크북 — 분석·등급 없이 모든 원자료를 한 시트에 ──────────
def _action(v):
    return {"적합": "자기구배 조사 진행", "조건부 적합": "추가 자기조사 후 재판정",
            "부적합": "대체 후보지 검토"}.get(v, "② 조사 완료 후")


FLAT_COLS = ["관할 본부", "관리번호", "후보지명", "도엽번호", "도엽명", "도엽 정정",
             "위도", "경도",
             "표고(m)", "유형", "지번 주소", "도로명 주소", "연계 기존점",
             "소유권", "조사대상", "조사일", "조사자", "기상",
             "차량영향 존재", "차량영향 이격", "전력시설 존재", "전력시설 이격",
             "통신시설 존재", "통신시설 이격", "금속류 존재", "금속류 이격",
             "매설물 존재", "매설물 이격", "건축물 존재", "건축물 이격",
             "자연환경 존재", "자연환경 이격", "방위표지 가능",
             "부적합 건수", "종합 판정", "판정 의견", "향후 조치",
             "방위1 경도", "방위1 위도", "방위1 방위각", "방위1 거리(m)",
             "방위1 방위각(카드)", "방위1 거리(카드)",
             "방위2 경도", "방위2 위도", "방위2 방위각", "방위2 거리(m)",
             "방위2 방위각(카드)", "방위2 거리(카드)",
             "방위 정정", "후보지↔기준점(m)",
             "접근 경로", "차량 진입", "유의사항", "도상 간섭요인", "회신 파일"]


def build_flat_workbook(recs, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "일괄취합"
    n = len(FLAT_COLS)
    _title(ws, n, f"지자기 도상선점 — 현장 조사서 일괄취합 ({len(recs)}건)   ·   {datetime.now():%Y-%m-%d}")
    _header(ws, FLAT_COLS)
    for ri, d in enumerate(recs, 3):
        dist = d["자기교란"]

        def g(key):
            ex, gap = dist.get(key, ("", ""))
            return ex, gap
        b1 = d.get("방위표지상세", {}).get("표지1", {})
        b2 = d.get("방위표지상세", {}).get("표지2", {})
        row = [d["관할본부"], d["관리번호"], d["후보지명"], d["도엽번호"], d["도엽명"],
               d["도엽정정"],
               d["위도"], d["경도"], d["표고"], d["유형"], d["지번주소"], d["도로명주소"],
               d["연계기존점"], d["소유권"], d["조사대상"], d["조사일"], d["조사자"], d["기상"],
               *g("차량영향"), *g("전력시설"), *g("통신시설"), *g("금속류"),
               *g("매설물"), *g("건축물"), *g("자연환경"), d["방위표지"],
               d["부적합수"], d["종합판정"], d["판정의견"].replace("\n", " "), _action(d["종합판정"]),
               b1.get("경도", ""), b1.get("위도", ""), b1.get("방위각", ""), b1.get("거리", ""),
               b1.get("방위각기재", ""), b1.get("거리기재", ""),
               b2.get("경도", ""), b2.get("위도", ""), b2.get("방위각", ""), b2.get("거리", ""),
               b2.get("방위각기재", ""), b2.get("거리기재", ""),
               d["방위정정"], (round(d["후보지이격"]) if d["후보지이격"] is not None else ""),
               d["접근경로"], d["차량진입"], d["유의사항"], d["도상간섭"], d["_src"]]
        for j, v in enumerate(row, 1):
            c = ws.cell(ri, j, v)
            c.font = F_VAL
            c.border = BORDER
            c.alignment = AL_L if FLAT_COLS[j - 1] in (
                "지번 주소", "도로명 주소", "도엽 정정", "방위 정정", "판정 의견",
                "접근 경로", "유의사항", "도상 간섭요인") else AL_C
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}{len(recs) + 2}"
    for j, col in enumerate(FLAT_COLS, 1):
        w = 30 if col in ("판정 의견", "도상 간섭요인") else \
            34 if col == "방위 정정" else \
            24 if col in ("지번 주소", "접근 경로") else \
            22 if col == "도엽 정정" else \
            18 if col == "유의사항" else \
            13 if "방위" in col or col in ("조사일", "조사자", "관할 본부", "후보지명") else 9
        ws.column_dimensions[get_column_letter(j)].width = w
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=str(DEF_BASE))
    ap.add_argument("--out")
    a = ap.parse_args()

    files = survey_files(a.dir)
    if not files:
        print(f"조사서 xlsx 를 찾지 못했습니다: {a.dir}")
        return

    recs = []
    for f in files:
        cards = parse_workbook(f)
        print(f"  {f.name:28} 카드 {len(cards):3}건")
        recs += cards
    recs.sort(key=lambda d: (d["관할본부"] or "", d["관리번호"] or ""))
    print(f"총 {len(recs)}건 취합")

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, recs)
    sheet_web(wb, recs)
    sheet_full(wb, recs)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(a.out) if a.out else OUT_DIR / f"{ts}_현장조사_취합검토_{len(recs)}건.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\n저장(취합·검토): {out}")

    flat = OUT_DIR / f"{ts}_현장조사_일괄취합_{len(recs)}건.xlsx"
    build_flat_workbook(recs, flat)
    print(f"저장(일괄취합): {flat}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
