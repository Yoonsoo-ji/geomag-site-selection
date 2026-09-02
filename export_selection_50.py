# -*- coding: utf-8 -*-
"""
선점 검토 50지점 엑셀 산출 — `docs/output/*_선점검토_50지점.xlsx`
================================================================

survey_review.html 이 「선점 검토」로 세는 세 갈래를 한 파일로 낸다:

  · 등급 A          5점 — 자기 청정 + 방위표지 좌표거리 ≥100 m
  · 등급 B 도엽대표 29점 — 자기 청정이나 표지 <100 m, 도엽 대표(A 승격 1순위)
  · 기존점 선점대상 16점 — 1등 지자기점 관측망 30점 중 표석 지속성 확보분
                           (2026-09-02 함양 추가)
                                                        합계 **50점**

⚠️ **기존점에서 실시한 답사는 기존점 행으로 합친다**(`AG.MERGED_INTO_EXISTING`).
사전 답사는 신규 후보지뿐 아니라 기존 지자기점에서도 했으므로, 그 카드를 등급
후보로 따로 세우면 같은 표석이 두 번 계상된다 — DS-047 「화천 기존점」은 관측망
화천과 **12 m** 떨어진 같은 자리였다. 답사 내용은 기존점 행에 열로 남긴다.

⚠️ **B 예비 7점은 넣지 않는다.** 자침편각 표기는 도엽당 1점이면 충분하므로
같은 도엽의 후순위 B 를 유효 후보로 세면 과대계상이 된다(14절 참조).

⚠️ 등급·도엽 대표 판정은 `aggregate_survey_xlsx.review()` ·
`sheet_priority()` **단일 출처**를 그대로 호출한다 — 여기서 다시 매기지 않는다.

    python export_selection_50.py
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import aggregate_survey_xlsx as AG
import existing_network as EN

ROOT = Path(__file__).parent
OUT_DIR = ROOT / "docs" / "output"
AGGREGATE = OUT_DIR / "20260818_211750_현장조사_일괄취합_103건.xlsx"

FONT = "맑은 고딕"
NAVY = "16324F"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
BAND = {"A": "E2F0E4", "B": "E4EDF7", "기존점": "FDF2E0"}
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _f(v):
    try:
        x = float(v)
        return None if pd.isna(x) else x
    except (TypeError, ValueError):
        return None


def _s(v):
    if v is None:
        return ""
    t = str(v).strip()
    return "" if t.lower() in ("nan", "none") else t


# ──────────────────────────────────────────────────────────────
def gather_survey():
    """현장조사 103건 → 등급 A · B 도엽대표만 추린 레코드."""
    if not AGGREGATE.exists():
        sys.exit(f"[중단] 취합본 없음: {AGGREGATE}")
    recs = AG.load_aggregate(AGGREGATE)
    prio = AG.sheet_priority(recs)
    # 기존점에서 한 답사는 여기서 빼고 `gather_existing()` 이 기존점 행에 붙인다.
    merged = {AG.MERGED_INTO_EXISTING[d["관리번호"]]: d
              for d in recs if d["관리번호"] in AG.MERGED_INTO_EXISTING}
    recs = [d for d in recs if d["관리번호"] not in AG.MERGED_INTO_EXISTING]
    out = []
    for d in recs:
        grade, concl, note = AG.review(d)
        rk = prio.get(d["관리번호"])
        if grade == "A":
            kind, why = "A", "방위표지 ≥100 m 확보"
        elif grade == "B" and rk and rk[0] == "대표":
            kind, why = "B", ("도엽 단독" if rk[2] == 1 else f"도엽 대표(B {rk[2]}건 중 1순위)")
        else:
            continue
        d = dict(d)
        d["_구분"], d["_사유"], d["_결론"], d["_의견"] = kind, why, concl, note
        out.append(d)
    return sorted(out, key=lambda d: (d["_구분"], d["관리번호"])), merged


def gather_existing(merged=None):
    """기존 관측망 중 선점 대상 — 좌표·성과는 **지도와 같은 파일**에서 읽는다.

    ⚠️ 대조표(점조서) 좌표를 그대로 쓰면 안 된다. 프로젝트 규칙은
    **「22~25 성과표에 있으면 성과표, 없으면 대조표」**(여주·언양만 예외)이고
    그 병합 결과가 `docs/data/existing_sites.geojson` 이다. 대조표만 읽으면
    상주·함양이 지도와 약 0.39 km 어긋난다.

    지점코드·도엽번호·소재지·관측장비는 geojson 에 없으므로 대조표에서 붙인다.
    """
    import json

    merged = merged or {}
    gj = ROOT / "docs" / "data" / "existing_sites.geojson"
    if not gj.exists():
        sys.exit(f"[중단] 측점 geojson 없음: {gj} — geomag_site_selection.py 를 먼저 실행")
    feats = {}
    for f in json.loads(gj.read_text(encoding="utf-8"))["features"]:
        pr = f["properties"]
        lon, lat = f["geometry"]["coordinates"]
        feats[_s(pr.get("name"))] = (pr, lat, lon)

    reg = EN.load_register().set_index("지점명")
    rows = []
    for nm in EN.EXISTING_TARGET:
        if nm not in feats:
            print(f"  ! 지도 자료 미매칭: {nm}", file=sys.stderr)
            continue
        pr, lat, lon = feats[nm]
        r = reg.loc[nm] if nm in reg.index else {}
        row = {
            "지점명": nm,
            "지점코드": _s(r.get("지점코드")),
            "도엽번호": _s(r.get("도엽번호")),
            "위도": _f(lat),
            "경도": _f(lon),
            "표고": _f(pr.get("elev")),
            "소재지": _s(pr.get("address")) or _s(r.get("소재지")),
            "최종관측": _s(r.get("최종관측")),
            "관측연도": _f(pr.get("obs_year")),
            "관측횟수": _f(r.get("관측횟수")),
            "편각": _f(pr.get("decl")),
            "복각": _f(pr.get("incl")),
            "총자력": _f(pr.get("total")),
            "관측장비": _s(r.get("관측장비")),
        }
        sv = merged.get(nm)
        if sv is not None:
            g, concl, note = AG.review(sv)
            md = AG.mark_max_dist(sv)
            row.update({
                "답사 관리번호": _s(sv["관리번호"]),
                "답사 후보지명": _s(sv["후보지명"]),
                "답사 종합판정": _s(sv["종합판정"]),
                "답사 등급": g,
                "답사 방위표지": _s(sv["방위표지"]),
                "답사 방위표지 최장(m)": md,
                "답사일": _s(sv["조사일"]),
                "답사자": _s(sv["조사자"]),
                "답사 검토": concl,
            })
        rows.append(row)
    return rows


# ──────────────────────────────────────────────────────────────
def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_table(ws, hdr, rows, widths, start=1, fills=None, numfmt=None):
    for j, h in enumerate(hdr, 1):
        ws.cell(row=start, column=j, value=h)
    style_header(ws, start, len(hdr))
    for i, rec in enumerate(rows, start + 1):
        fill = PatternFill("solid", fgColor=fills[i - start - 1]) if fills else None
        for j, h in enumerate(hdr, 1):
            cell = ws.cell(row=i, column=j, value=rec.get(h))
            cell.font = Font(name=FONT, size=9.5)
            cell.border = BOX
            cell.alignment = Alignment(
                vertical="center", wrap_text=h in ("소재지", "지번주소", "판정의견",
                                                   "자기교란", "방위표지상세", "선정 사유"),
                horizontal="center" if h not in (
                    "소재지", "지번주소", "판정의견", "자기교란", "후보지명",
                    "방위표지상세", "선정 사유", "비고") else "left")
            if fill:
                cell.fill = fill
            if numfmt and h in numfmt:
                cell.number_format = numfmt[h]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    sv, merged = gather_survey()
    ex = gather_existing(merged)
    n_a = sum(1 for d in sv if d["_구분"] == "A")
    n_b = sum(1 for d in sv if d["_구분"] == "B")
    total = len(sv) + len(ex)
    print(f"등급 A {n_a} · B 도엽대표 {n_b} · 기존점 선점대상 {len(ex)} = 합계 {total}점")
    if total != 50:
        print(f"  ⚠ 합계가 50 이 아니다 ({total}) — 명단·판정 변경 여부 확인할 것")

    from openpyxl import Workbook
    wb = Workbook()

    # ═══ 1) 통합 목록 50점 ══════════════════════════════════════
    ws = wb.active
    ws.title = "선점검토 50점"
    ws["A1"] = f"지자기 선점 검토 대상 {total}점 — 등급 A {n_a} · B 도엽대표 {n_b} · 기존점 {len(ex)}"
    ws["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws["A2"] = ("현장조사 103건 중 A·B 도엽대표 + 1등 지자기점 관측망 30점 중 선점 대상. "
                "B 예비 7점은 도엽 중복이라 제외. 좌표는 현장 기준점이며, 기존점은 지도와 "
                "같은 병합 규칙(22~25 성과표 우선, 없으면 점조서)을 따른다.")
    ws["A2"].font = Font(name=FONT, size=9, color="666666")
    ws["A3"] = f"작성 {dt.date.today():%Y-%m-%d}"
    ws["A3"].font = Font(name=FONT, size=9, color="666666")

    hdr = ["연번", "구분", "관리번호/코드", "지점명", "관할본부", "도엽번호", "도엽명",
           "위도", "경도", "표고(m)", "소재지", "선정 사유", "비고"]
    rows, fills = [], []
    n = 0
    for d in sv:
        n += 1
        rows.append({
            "연번": n, "구분": d["_구분"], "관리번호/코드": _s(d["관리번호"]),
            "지점명": _s(d["후보지명"]), "관할본부": _s(d["관할본부"]),
            "도엽번호": _s(d["도엽번호"]), "도엽명": _s(d["도엽명"]),
            "위도": _f(d["위도"]), "경도": _f(d["경도"]), "표고(m)": _f(d["표고"]),
            "소재지": _s(d["지번주소"]) or _s(d["도로명주소"]),
            "선정 사유": d["_사유"], "비고": d["_결론"],
        })
        fills.append(BAND[d["_구분"]])
    for r in ex:
        n += 1
        rows.append({
            "연번": n, "구분": "기존점", "관리번호/코드": r["지점코드"],
            "지점명": r["지점명"], "관할본부": "",
            "도엽번호": r["도엽번호"], "도엽명": r["지점명"],
            "위도": r["위도"], "경도": r["경도"], "표고(m)": r["표고"],
            "소재지": r["소재지"],
            "선정 사유": ("기존 관측망 · 표석 지속성 확보"
                        + (" · 사전답사 실시" if r.get("답사 관리번호") else "")),
            "비고": (f"성과 {int(r['관측연도'])}년" if r["관측연도"] else "")
                    + (f" · 점조서 {r['최종관측']}" if r["최종관측"] else ""),
        })
        fills.append(BAND["기존점"])
    write_table(ws, hdr, rows,
                [5, 7, 13, 18, 12, 13, 10, 11, 11, 9, 40, 26, 18],
                start=5, fills=fills,
                numfmt={"위도": "0.000000", "경도": "0.000000", "표고(m)": "0.0"})

    # ═══ 2) 현장조사 상세 ══════════════════════════════════════
    ws2 = wb.create_sheet("현장조사 상세 (A·B대표)")
    ws2["A1"] = f"현장조사 카드 상세 — 등급 A {n_a} · B 도엽대표 {n_b}"
    ws2["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws2["A2"] = ("종합판정·거리·방위각은 카드 기재값이 아니라 좌표 계산값 기준으로 "
                 "정정된 값이다(도엽·방위 정정 열 참조).")
    ws2["A2"].font = Font(name=FONT, size=9, color="666666")
    h2 = ["구분", "관리번호", "지점명", "관할본부", "도엽번호", "도엽명",
          "위도", "경도", "표고(m)", "지번주소", "종합판정", "자기교란",
          "부적합 항목수", "방위표지", "방위표지 최장거리(m)", "방위표지상세",
          "후보지↔기준점(m)", "차량진입", "소유권", "조사일", "조사자", "기상",
          "판정의견", "도엽 정정", "방위 정정"]
    r2, f2 = [], []
    for d in sv:
        r2.append({
            "구분": d["_구분"], "관리번호": _s(d["관리번호"]),
            "지점명": _s(d["후보지명"]), "관할본부": _s(d["관할본부"]),
            "도엽번호": _s(d["도엽번호"]), "도엽명": _s(d["도엽명"]),
            "위도": _f(d["위도"]), "경도": _f(d["경도"]), "표고(m)": _f(d["표고"]),
            "지번주소": _s(d["지번주소"]), "종합판정": _s(d["종합판정"]),
            "자기교란": _s(d["자기교란"]), "부적합 항목수": _f(d["부적합수"]),
            "방위표지": _s(d["방위표지"]),
            "방위표지 최장거리(m)": AG.mark_max_dist(d),
            "방위표지상세": _s(d["방위표지상세"]),
            "후보지↔기준점(m)": _f(d["후보지이격"]),
            "차량진입": _s(d["차량진입"]), "소유권": _s(d["소유권"]),
            "조사일": _s(d["조사일"]), "조사자": _s(d["조사자"]),
            "기상": _s(d["기상"]), "판정의견": _s(d["판정의견"]),
            "도엽 정정": _s(d["도엽정정"]), "방위 정정": _s(d["방위정정"]),
        })
        f2.append(BAND[d["_구분"]])
    write_table(ws2, h2, r2,
                [6, 11, 18, 12, 13, 10, 11, 11, 9, 34, 12, 26, 8, 10, 15, 30,
                 14, 10, 12, 12, 10, 10, 40, 22, 22],
                start=4, fills=f2,
                numfmt={"위도": "0.000000", "경도": "0.000000", "표고(m)": "0.0",
                        "방위표지 최장거리(m)": "0.0", "후보지↔기준점(m)": "0.0"})

    # ═══ 3) 기존 관측점 ════════════════════════════════════════
    ws3 = wb.create_sheet("기존 관측점 선점대상")
    ws3["A1"] = f"1등 지자기점 관측망 30점 중 선점 대상 {len(ex)}점"
    ws3["A1"].font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws3["A2"] = ("좌표·성과는 지도(existing_sites.geojson)와 같은 값 — 22~25 성과표가 "
                 "있으면 그 값, 없으면 점조서(2025.12.15) 최종관측 행. "
                 "나머지 14점은 저수지(제방) 설치점으로 표석 지속성 미확보 판정. "
                 "「답사」 열은 그 기존점에서 실시한 사전 현장답사 기록이다.")
    ws3["A2"].font = Font(name=FONT, size=9, color="666666")
    # ⚠️ 성과값은 geojson(성과표 우선 병합), 「점조서 최종관측」은 대조표에서
    #    온다 — 두 출처가 다를 수 있으므로 **열을 나눠** 적는다. 하나로 합치면
    #    조도처럼 「2010연구성과」 옆에 성과표 값이 서는 짝이 생긴다.
    h3 = ["지점명", "지점코드", "도엽번호", "위도", "경도", "표고(m)", "소재지",
          "성과 연도", "편각(°)", "복각(°)", "총자력(nT)",
          "점조서 최종관측", "점조서 관측횟수", "점조서 관측장비",
          "답사 관리번호", "답사 후보지명", "답사 종합판정", "답사 등급",
          "답사 방위표지", "답사 방위표지 최장(m)", "답사일", "답사자", "답사 검토"]
    alias = {"편각(°)": "편각", "복각(°)": "복각", "총자력(nT)": "총자력",
             "표고(m)": "표고", "성과 연도": "관측연도",
             "점조서 최종관측": "최종관측", "점조서 관측횟수": "관측횟수",
             "점조서 관측장비": "관측장비"}
    r3 = [{k: r.get(alias.get(k, k)) for k in h3} for r in ex]
    write_table(ws3, h3, r3,
                [10, 10, 13, 11, 11, 9, 44, 10, 11, 11, 12, 15, 12, 13,
                 12, 14, 12, 8, 11, 15, 12, 14, 34],
                start=4, fills=[BAND["기존점"]] * len(r3),
                numfmt={"위도": "0.000000", "경도": "0.000000", "표고(m)": "0.0",
                        "편각(°)": "0.0000", "복각(°)": "0.0000",
                        "총자력(nT)": "#,##0.0", "성과 연도": "0",
                        "점조서 관측횟수": "0", "답사 방위표지 최장(m)": "0.0"})

    # ── 중복 위치 점검 — 세 갈래를 합치면 같은 표석이 두 번 셀 수 있다 ──
    import numpy as np
    pts = [(r["지점명"], r["구분"], r["위도"], r["경도"]) for r in rows]
    dup = []
    for i in range(len(pts)):
        for j in range(i + 1, len(pts)):
            (n1, k1, a1, o1), (n2, k2, a2, o2) = pts[i], pts[j]
            if None in (a1, o1, a2, o2):
                continue
            mid = np.radians((a1 + a2) / 2)
            km = float(np.hypot((o2 - o1) * 111.320 * np.cos(mid),
                                (a2 - a1) * 110.574))
            if km < 0.5:
                dup.append(f"{n1}({k1}) ↔ {n2}({k2}) {km*1000:.0f} m")
    if dup:
        print("  ⚠ 같은 자리로 보이는 쌍 — 실제 지점 수는 그만큼 적다:")
        for t in dup:
            print(f"      {t}")
        ws["A4"] = "⚠ 중복 위치: " + " · ".join(dup)
        ws["A4"].font = Font(name=FONT, size=9, bold=True, color="B2182B")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / f"{dt.datetime.now():%Y%m%d_%H%M%S}_선점검토_{total}지점.xlsx"
    wb.save(path)
    print(f"[저장] {path}")
    return path


if __name__ == "__main__":
    main()
