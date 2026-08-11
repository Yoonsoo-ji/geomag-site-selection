# -*- coding: utf-8 -*-
"""
회신 조사서 워크북 병합 → 배포 양식(총괄 목록 + 카드 103장) 재구성
====================================================================

11개 본부가 회신한 조사서(각 = 총괄 목록 + 채운 카드 시트들)를 원본 배포 양식
`20260707_조사카드_v2_총괄_전체103건.xlsx` 과 **동일한 구조**로 한 파일에 합친다.

  · 총괄 목록: 원본 템플릿의 총괄 목록(103행)을 그대로 복사
  · 카드 시트: 각 회신본의 카드를 DS 번호순으로 복사 —
    값·서식(글꼴·채움·테두리·정렬·표시형식)·병합셀·행높이·열너비·
    임베드 사진·드롭다운(데이터 검증)·하이퍼링크까지 보존

openpyxl 은 워크북 간 시트 복사를 지원하지 않으므로 셀 단위로 충실 복사한다.

사용:
    python merge_survey_workbooks.py            # 전체 11파일
    python merge_survey_workbooks.py --test     # 소형 파일만(1·2·11) 미리보기
"""
import argparse
import copy
import io
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from aggregate_survey_xlsx import (parse_workbook, parse_card, key_disturb,
                                   survey_files, DEF_BASE)

ROOT = Path(__file__).parent
SKETCH_DIR = ROOT / "docs" / "output" / "_sketches"
SRC_DIR = Path(r"D:\LX_yoons\2026_research\2026_지자기 연구\20260811_사전 현장 조사서 취합"
               r"\2.조사서_부산울산")
TEMPLATE = Path(r"D:\LX_yoons\2026_research\2026_지자기 연구\20260707_지자기 사전 현장조사 엑셀"
                r"\20260707_조사카드_v2_총괄_전체103건.xlsx")
OUT_DIR = ROOT / "docs" / "output"


def copy_sheet(src_ws, dst_wb, title):
    """워크북 간 시트 충실 복사."""
    dst = dst_wb.create_sheet(title)

    # 열 너비 / 숨김
    for k, dim in src_ws.column_dimensions.items():
        d = dst.column_dimensions[k]
        if dim.width is not None:
            d.width = dim.width
        d.hidden = dim.hidden
    # 행 높이
    for k, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst.row_dimensions[k].height = dim.height

    # 셀 값 + 스타일 + 하이퍼링크
    for row in src_ws.iter_rows():
        for c in row:
            if c.value is None and not c.has_style and not c.hyperlink:
                continue
            nc = dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                nc.font = copy.copy(c.font)
                nc.fill = copy.copy(c.fill)
                nc.border = copy.copy(c.border)
                nc.alignment = copy.copy(c.alignment)
                nc.protection = copy.copy(c.protection)
                nc.number_format = c.number_format
            if c.hyperlink:
                nc.hyperlink = copy.copy(c.hyperlink)

    # 병합 셀
    for mc in list(src_ws.merged_cells.ranges):
        dst.merge_cells(str(mc))

    # 임베드 이미지 (지도·사진)
    for img in getattr(src_ws, "_images", []):
        try:
            ni = XLImage(io.BytesIO(img._data()))
            ni.anchor = img.anchor
            dst.add_image(ni)
        except Exception as e:   # noqa: BLE001
            print(f"    ! 이미지 복사 실패({title}): {e}", file=sys.stderr)

    # 데이터 검증(드롭다운)
    try:
        for dv in src_ws.data_validations.dataValidation:
            ndv = copy.copy(dv)
            dst.add_data_validation(ndv)
            ndv.sqref = dv.sqref
    except Exception as e:   # noqa: BLE001
        print(f"    ! 데이터검증 복사 실패({title}): {e}", file=sys.stderr)

    # 시트 뷰 속성
    if src_ws.freeze_panes:
        dst.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst.auto_filter.ref = src_ws.auto_filter.ref
    dst.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    if src_ws.sheet_properties.tabColor:
        dst.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
    return dst


def card_index(name):
    m = re.match(r"(\d+)_", name)
    return int(m.group(1)) if m else 9999


# ── ① 총괄 목록에 현장 결과 열 추가 ──────────────────────────────────────────
_HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
_VAL_FONT = Font(name="맑은 고딕", size=9)
_FILL_HDR = PatternFill("solid", fgColor="2E5A3E")
_FILL_OK = PatternFill("solid", fgColor="E2EFDA")
_FILL_COND = PatternFill("solid", fgColor="FFF2CC")
_FILL_BAD = PatternFill("solid", fgColor="FDE0DC")
_AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

ADD_COLS = ["소유권", "조사일", "조사자", "기상", "종합 판정", "방위표지",
            "핵심 교란요인", "판정 의견"]
ADD_W = [8, 11, 13, 6, 10, 8, 24, 34]


def augment_master(ws, records):
    """총괄 목록(A~P 기본) 오른쪽에 현장 결과 열 추가. 우측 helper(Q~T)는 정리."""
    by_id = {d["관리번호"]: d for d in records}
    start = 17   # Q 열부터
    # 기존 우측 helper 표(관할본부 건수) 제거
    for r in range(2, ws.max_row + 1):
        for c in range(start, start + 6):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()
    # 헤더
    for j, h in enumerate(ADD_COLS, start):
        c = ws.cell(2, j, h)
        c.font = _HDR_FONT
        c.fill = _FILL_HDR
        c.alignment = _AL_C
        c.border = _BORDER
    # 데이터 (A열 관리번호로 매칭)
    jcol = start + ADD_COLS.index("종합 판정")
    bcol = start + ADD_COLS.index("방위표지")
    last = 2
    for r in range(3, ws.max_row + 1):
        mid = ws.cell(r, 1).value
        if not mid:
            continue
        d = by_id.get(str(mid).strip())
        if not d:
            continue
        last = r
        vals = [d["소유권"], d["조사일"], d["조사자"], d["기상"], d["종합판정"],
                d["방위표지"], key_disturb(d) or "-", d["판정의견"].replace("\n", " ")]
        for j, v in enumerate(vals, start):
            c = ws.cell(r, j, v)
            c.font = _VAL_FONT
            c.border = _BORDER
            c.alignment = _AL_L if ADD_COLS[j - start] in ("핵심 교란요인", "판정 의견") else _AL_C
            if j == jcol:
                c.fill = {"적합": _FILL_OK, "조건부 적합": _FILL_COND,
                          "부적합": _FILL_BAD}.get(v, PatternFill())
            if j == bcol and v == "불가":
                c.fill = _FILL_BAD
    for j, w in enumerate(ADD_W, start):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.auto_filter.ref = f"A2:{get_column_letter(start + len(ADD_COLS) - 1)}{last}"


# ── ② 카드 ⑥ 약도 칸에 약도 이미지 삽입 ─────────────────────────────────────
def embed_sketch(ws, code):
    p = SKETCH_DIR / f"{code}.png"
    if not p.exists():
        return False
    r6 = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith("⑥"):
            r6 = r
            break
    if r6 is None:
        return False
    note_r = r6 + 1
    ws.cell(note_r, 1).value = None   # "추후 작업" 안내문 제거
    for rr in range(note_r, note_r + 4):
        ws.row_dimensions[rr].height = 95
    with PILImage.open(p) as im:
        w, h = im.size
    tw = 470
    th = int(tw * h / w)
    xim = XLImage(str(p))
    xim.width = tw
    xim.height = th
    ws.add_image(xim, f"A{note_r}")
    return True


# ── 이미 만든 취합 엑셀에 결과 열·약도만 추가(원본 회신본 미접근) ──────────────
def augment_existing(path):
    path = Path(path)
    print(f"취합 엑셀 여는 중: {path.name} ({path.stat().st_size/1e6:.0f} MB)...")
    wb = load_workbook(path)   # 기본 로드(이미지 보존)
    # 카드 시트에서 직접 현장 결과 파싱 → 관리번호 매칭용
    records, cards = [], []
    for name in wb.sheetnames:
        if name == "총괄 목록":
            continue
        d = parse_card(wb[name])
        if d.get("관리번호"):
            records.append(d)
            cards.append(name)
    print(f"  카드 {len(cards)}장에서 결과 파싱")

    augment_master(wb["총괄 목록"], records)
    print("  총괄 목록 결과 열 추가 완료")

    nsk = 0
    for name in cards:
        if embed_sketch(wb[name], f"DS-{card_index(name):03d}"):
            nsk += 1
    print(f"  약도 {nsk}장 삽입")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUT_DIR / f"{ts}_현장조사_취합_총괄_약도포함_전체{len(cards)}건.xlsx"
    print("저장 중(대용량, 수 분 소요 가능)...")
    wb.save(out)
    print(f"\n저장: {out}  ({out.stat().st_size/1e6:.1f} MB, 시트 {len(wb.sheetnames)}개)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--test", action="store_true")
    ap.add_argument("--dir", default=None)
    ap.add_argument("--augment", nargs="?", const="auto",
                    help="기존 취합 엑셀에 결과 열·약도만 추가(원본 미접근). "
                         "경로 생략 시 최신 취합_총괄 파일 자동 선택")
    a = ap.parse_args()

    if a.augment:
        target = a.augment
        if target == "auto":
            cands = sorted(OUT_DIR.glob("*_현장조사_취합_총괄_전체*.xlsx"))
            if not cands:
                print("취합_총괄 파일을 찾지 못했습니다."); return
            target = cands[-1]
        augment_existing(target)
        return

    files = survey_files(a.dir)
    if a.test:
        files = [f for f in files if re.match(r"(\d+)", f.name)
                 and int(re.match(r"(\d+)", f.name).group(1)) in (1, 2, 11)]

    out = Workbook()
    out.remove(out.active)

    # ── 현장 결과 레코드(관리번호 매칭용) ──
    print("현장 결과 파싱(총괄 목록 열 추가용)...")
    records = []
    for f in files:
        records += parse_workbook(f)

    # ── 총괄 목록: 템플릿에서 복사 후 현장 결과 열 추가 ──
    print("총괄 목록 복사(템플릿) + 결과 열 추가...")
    tpl = load_workbook(TEMPLATE)
    copy_sheet(tpl["총괄 목록"], out, "총괄 목록")
    tpl.close()
    augment_master(out["총괄 목록"], records)

    # ── 카드 시트: 파일별로 열고 닫으며 복사(메모리 보호), 이후 DS 순 재정렬 ──
    total = 0
    for fi, f in enumerate(files, 1):
        print(f"[{fi}/{len(files)}] {f.name} 여는 중...")
        wb = load_workbook(f)
        names = [s for s in wb.sheetnames if s != "총괄 목록"]
        nsk = 0
        for name in names:
            dst = copy_sheet(wb[name], out, name)
            if embed_sketch(dst, f"DS-{card_index(name):03d}"):
                nsk += 1
            total += 1
        wb.close()
        del wb
        print(f"    카드 {len(names)}장 복사, 약도 {nsk}장 삽입 (누적 {total})")

    # DS 번호순 정렬 (총괄 목록은 맨 앞 고정)
    head = out["총괄 목록"]
    rest = sorted((ws for ws in out.worksheets if ws is not head),
                  key=lambda ws: card_index(ws.title))
    out._sheets = [head] + rest
    print(f"카드 {total}장 병합 완료")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = "미리보기" if a.test else f"약도포함_전체{total}건"
    dst = OUT_DIR / f"{ts}_현장조사_취합_총괄_{tag}.xlsx"
    dst.parent.mkdir(parents=True, exist_ok=True)
    print("저장 중(대용량, 수 분 소요 가능)...")
    out.save(dst)
    mb = dst.stat().st_size / 1e6
    print(f"\n저장: {dst}  ({mb:.1f} MB, 시트 {len(out.sheetnames)}개)")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
