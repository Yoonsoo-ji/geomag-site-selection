#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1등급 P1 후보지 63개 전체 서브격자 분석
  1. 각 P1 셀 중심 ±5km / 1km 격자 → zone_cache WKB 필터 → 도로접근성
  2. s2~s6 점수 산정 → 상위 5개 선정
  3. 기준점.kml 에서 각 사이트 5km 반경 내 기준점 추출
  4. 합본 KML 생성 (사이트별 폴더)
  5. Excel 생성 (전체 요약 + 기준점 매칭)
"""
import sys, json, warnings, math
sys.stdout.reconfigure(encoding="utf-8")
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import geopandas as gpd
import shapely
from shapely.geometry import Point, LineString, box
from shapely.prepared import prep
from pathlib import Path
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── 경로 ──────────────────────────────────────────────────────
MAIN_DIR  = Path("C:/LG_gram_backup_users/LX/2026_geomag")
DATA_DIR  = MAIN_DIR / "data"
CACHE_DIR = DATA_DIR / "zone_cache"
OUT_DIR   = MAIN_DIR / "docs" / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)

UTM_CRS   = "EPSG:5179"
WGS84_CRS = "EPSG:4326"
TS        = datetime.now().strftime("%Y%m%d_%H%M%S")

RADIUS_M   = 5_000
GRID_STEP  = 1_000
ROAD_MAX_M = 2_000
TOP_N      = 5          # 사이트당 상위 후보 수
REF_RADIUS = 5_000      # 기준점 검색 반경 (m)

ZONE_NAMES = ["power","railway","urban_dense","urban_resid",
              "pipeline","comm","wind","quarry","water",
              "geology","fault","anomaly"]

ELEV_CACHE_PATH = OUT_DIR / "all_p1_elev_cache.json"
ELEV_OFFSET     = 500   # 경사도 오프셋 (m)
DIRS = {"C":(0,0),"N":(0,ELEV_OFFSET),"S":(0,-ELEV_OFFSET),
        "E":(ELEV_OFFSET,0),"W":(-ELEV_OFFSET,0)}

print("=" * 66)
print("  1등급 P1 후보지 63개 — 전체 서브격자 분석 + 기준점 매칭")
print("=" * 66)
t0 = datetime.now()


# ════════════════════════════════════════════════════════════════
# 공통 데이터 로드
# ════════════════════════════════════════════════════════════════
# P1 사이트
print("\n[준비] 데이터 로드...")
df_cands = pd.read_csv(MAIN_DIR / "output" / "candidate_sites.csv")
p1_sites  = df_cands[df_cands.priority == 1].reset_index(drop=True)
print(f"  P1 사이트: {len(p1_sites)}개")

# KIGAM
kigam_path = DATA_DIR / "mag_1982-2018_1.5min_ed.dat"
klat_arr = klon_arr = kanom_arr = None
if kigam_path.exists():
    kdf = pd.read_csv(kigam_path, sep=r"\s+", skiprows=9, header=None,
                      names=["lon","lat","anomaly_nT"],
                      na_values=["99999","-99999"], on_bad_lines="skip")
    kdf = kdf.apply(pd.to_numeric, errors="coerce").dropna()
    klat_arr = kdf.lat.values; klon_arr = kdf.lon.values; kanom_arr = kdf.anomaly_nT.values
    print(f"  KIGAM: {len(kdf):,}개")

# 도로
road_json = DATA_DIR / "access_roads.json"
road_gdf_wgs = None
if road_json.exists():
    with open(road_json, encoding="utf-8") as f: rd = json.load(f)
    nodes_rd = {e["id"]:(e["lon"],e["lat"]) for e in rd.get("elements",[]) if e["type"]=="node"}
    road_lines = []
    for e in rd.get("elements",[]):
        if e["type"]=="way":
            coords=[nodes_rd[n] for n in e.get("nodes",[]) if n in nodes_rd]
            if len(coords)>=2: road_lines.append(LineString(coords))
    if road_lines:
        road_gdf_wgs = gpd.GeoDataFrame(geometry=road_lines, crs=WGS84_CRS)
        print(f"  도로: {len(road_gdf_wgs):,}개")

# 기준점 파싱
print("\n  기준점.kml 파싱...")
KML_REF = OUT_DIR / "기준점.kml"
NS_KML  = "http://www.opengis.net/kml/2.2"
def ktag(n): return f"{{{NS_KML}}}{n}"

ref_points = []  # [{"category","name","lon","lat","point_utm"}]
if KML_REF.exists():
    tree_ref = ET.parse(KML_REF)
    for folder_top in tree_ref.getroot().iter(ktag("Folder")):
        for subfolder in folder_top.findall(ktag("Folder")):
            sf_nm_el = subfolder.find(ktag("name"))
            cat = sf_nm_el.text.strip() if sf_nm_el is not None else ""
            for pm in subfolder.findall(ktag("Placemark")):
                pt_el = pm.find(ktag("Point"))
                if pt_el is None: continue
                ce = pt_el.find(ktag("coordinates"))
                if ce is None or not ce.text: continue
                parts = ce.text.strip().split(",")
                if len(parts) < 2: continue
                try: lon, lat = float(parts[0]), float(parts[1])
                except: continue
                nm = ""
                ed = pm.find(ktag("ExtendedData"))
                if ed is not None:
                    for sd in ed.iter(ktag("SimpleData")):
                        if sd.get("name")=="CTRLPNT_NM":
                            nm = (sd.text or "").strip(); break
                nm_el = pm.find(ktag("name"))
                if nm_el is not None and nm_el.text: nm = nm_el.text.strip() or nm
                ref_points.append({"category":cat,"name":nm,"lon":lon,"lat":lat})

    # UTM 변환 (거리 계산용)
    if ref_points:
        ref_df = pd.DataFrame(ref_points)
        ref_gdf = gpd.GeoDataFrame(ref_df,
                    geometry=gpd.points_from_xy(ref_df.lon, ref_df.lat),
                    crs=WGS84_CRS).to_crs(UTM_CRS)
        ref_utm_xy = np.column_stack([ref_gdf.geometry.x, ref_gdf.geometry.y])
        print(f"  기준점: {len(ref_points):,}개 (삼각점+통합기준점)")
else:
    print("  ⚠ 기준점.kml 없음")
    ref_utm_xy = np.empty((0,2))

# 고도 캐시 로드
elev_cache = {}
if ELEV_CACHE_PATH.exists():
    with open(ELEV_CACHE_PATH, encoding="utf-8") as f:
        elev_cache = json.load(f)
    print(f"  고도 캐시: {len(elev_cache):,}개 포인트 로드됨")


# ════════════════════════════════════════════════════════════════
# 고도 API 배치 요청 (전 사이트 통합)
# ════════════════════════════════════════════════════════════════
def elev_key(lat, lon): return f"{lat:.5f},{lon:.5f}"

def fetch_elev_batch_list(latlons):
    """최대 200개씩 Open-Elevation 배치 요청"""
    url = "https://api.open-elevation.com/api/v1/lookup"
    results = [None]*len(latlons)
    BATCH = 200
    for i in range(0, len(latlons), BATCH):
        batch = latlons[i:i+BATCH]
        try:
            r = requests.post(url, json={"locations":[{"latitude":la,"longitude":lo}
                              for la,lo in batch]}, timeout=120,
                              headers={"User-Agent":"geomag-research/1.0"})
            if r.status_code==200:
                for j,item in enumerate(r.json()["results"]):
                    results[i+j] = item.get("elevation")
        except Exception as e:
            pass  # 캐시 미스 → 기본값 사용
    return results

# 전 사이트 격자점 생성 → 미캐시 포인트 수집
print("\n[1] 전 사이트 서브격자 생성 및 고도 미캐시 수집...")
all_site_grids = []   # [{site_row, cx, cy, pts_utm, gdf_all}]

for _, row in p1_sites.iterrows():
    cg = gpd.GeoDataFrame(geometry=[Point(row.lon, row.lat)],
                           crs=WGS84_CRS).to_crs(UTM_CRS)
    cx, cy = cg.geometry.iloc[0].x, cg.geometry.iloc[0].y
    xs = np.arange(cx-RADIUS_M, cx+RADIUS_M+1, GRID_STEP)
    ys = np.arange(cy-RADIUS_M, cy+RADIUS_M+1, GRID_STEP)
    cp = Point(cx, cy)
    pts_utm = [Point(x,y) for x in xs for y in ys
               if Point(x,y).distance(cp) <= RADIUS_M]
    n = len(pts_utm)
    gdf = gpd.GeoDataFrame({"gid":range(n)}, geometry=pts_utm, crs=UTM_CRS)
    wgs = gdf.to_crs(WGS84_CRS)
    gdf["lat"] = wgs.geometry.y.round(6)
    gdf["lon"] = wgs.geometry.x.round(6)
    all_site_grids.append({"row":row,"cx":cx,"cy":cy,
                            "pts_utm":pts_utm,"gdf":gdf,"n":n})

# 미캐시 포인트 수집 (전 사이트 × 5방향)
missing_keys = []   # (key, lat, lon)
for sg in all_site_grids:
    gdf = sg["gdf"]
    for di, (dname,(dx,dy)) in enumerate(DIRS.items()):
        pts5 = gpd.GeoDataFrame(
            geometry=[Point(p.x+dx, p.y+dy) for p in sg["pts_utm"]],
            crs=UTM_CRS).to_crs(WGS84_CRS)
        for lat5, lon5 in zip(pts5.geometry.y, pts5.geometry.x):
            k = elev_key(lat5, lon5)
            if k not in elev_cache:
                missing_keys.append((k, lat5, lon5))

# 중복 제거
unique_missing = list({k:(la,lo) for k,la,lo in missing_keys}.items())
print(f"  미캐시 고도 요청: {len(unique_missing):,}개 포인트")

if unique_missing:
    print(f"  Open-Elevation API 요청 중 (배치 200개씩)...")
    latlons_req = [v for k,v in unique_missing]
    elevs = fetch_elev_batch_list(latlons_req)
    new_cached = 0
    for (k, (la,lo)), el in zip(unique_missing, elevs):
        if el is not None:
            elev_cache[k] = el
            new_cached += 1
    print(f"  → {new_cached:,}개 취득 완료 (실패 {len(unique_missing)-new_cached}개 → 기본값 사용)")
    with open(ELEV_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(elev_cache, f)
    print(f"  고도 캐시 저장: {ELEV_CACHE_PATH.name} ({len(elev_cache):,}개)")


# ════════════════════════════════════════════════════════════════
# 사이트별 분석
# ════════════════════════════════════════════════════════════════
def zone_dist(pts_list, zone_geom, simplify_m=300):
    if zone_geom is None or zone_geom.is_empty:
        return np.full(len(pts_list), 50_000.0)
    return np.array([pt.distance(zone_geom.simplify(simplify_m)) for pt in pts_list])

def kigam_s5(lat, lon):
    if klat_arr is None: return 6.0, np.nan
    mask = (np.abs(klat_arr-lat)<=0.05)&(np.abs(klon_arr-lon)<=0.05)
    pk = kanom_arr[mask]
    if len(pk)<3: return 6.0, np.nan
    v = float(np.percentile(pk,90)-np.percentile(pk,10))
    sc = 5.0 if v<30 else (8.0 if v<=150 else (10.0 if v<=400 else (7.0 if v<=800 else 0.0)))
    return sc, round(v,1)

print(f"\n[2] 63개 사이트 분석 시작...")
results_all = []   # 최종 결과 리스트

for si, sg in enumerate(all_site_grids):
    row    = sg["row"]
    sid    = int(row.site_id)
    clat   = row.lat; clon = row.lon
    cx,cy  = sg["cx"], sg["cy"]
    pts_utm= sg["pts_utm"]
    gdf    = sg["gdf"].copy()
    n      = sg["n"]
    LOCAL_BOX = box(cx-RADIUS_M*1.2, cy-RADIUS_M*1.2,
                    cx+RADIUS_M*1.2, cy+RADIUS_M*1.2)

    print(f"  [{si+1:02d}/63] #{sid}  ({clat:.3f}°N, {clon:.3f}°E)", end="")

    # ── 경사도 계산 ──────────────────────────────────────────
    slopes = np.full(n, np.nan)
    for gi in range(n):
        pt = pts_utm[gi]
        grads = []
        pts5g = gpd.GeoDataFrame(
            geometry=[Point(pt.x+dx, pt.y+dy) for dx,dy in DIRS.values()],
            crs=UTM_CRS).to_crs(WGS84_CRS)
        elevs5 = [elev_cache.get(elev_key(r.geometry.y, r.geometry.x))
                  for _,r in pts5g.iterrows()]
        ec = elevs5[0]  # "C" 방향
        if ec is not None:
            for ev in elevs5[1:]:
                if ev is not None:
                    grads.append(math.degrees(math.atan(abs(ev-ec)/ELEV_OFFSET)))
        if grads: slopes[gi] = np.mean(grads)
    # 미취득 → 사이트 전체 중앙값으로 대체
    med_sl = np.nanmedian(slopes) if not np.all(np.isnan(slopes)) else 5.0
    slopes = np.where(np.isnan(slopes), med_sl, slopes)
    gdf["slope_deg"] = np.round(slopes, 2)

    # ── 제외구역 필터 ────────────────────────────────────────
    zones = {}
    for nm in ZONE_NAMES:
        p = CACHE_DIR / f"{nm}.wkb"
        if not p.exists(): zones[nm]=None; continue
        g = shapely.from_wkb(p.read_bytes())
        try:
            lc = g.intersection(LOCAL_BOX)
            zones[nm] = None if lc.is_empty else lc
        except: zones[nm] = g

    excl = np.zeros(n, dtype=bool)
    excl_r = [""]*n
    for nm in ZONE_NAMES:
        z = zones[nm]
        if z is None or z.is_empty: continue
        pz = prep(z)
        for i,pt in enumerate(pts_utm):
            if not excl[i] and pz.contains(pt):
                excl[i]=True; excl_r[i]=nm
    gdf["excl"]  = excl
    gdf["excl_r"]= excl_r
    cands = gdf[~excl].copy().reset_index(drop=True)
    n_excl = n - len(cands)

    # ── 도로 접근성 ──────────────────────────────────────────
    road_dist_arr = np.full(len(cands), np.nan)
    if road_gdf_wgs is not None:
        road_utm = road_gdf_wgs.to_crs(UTM_CRS)
        local_roads = road_utm[road_utm.geometry.intersects(LOCAL_BOX)]
        if len(local_roads)>0:
            ru = local_roads.geometry.unary_union
            pts_c = cands.geometry.tolist()
            road_dist_arr = np.array([pt.distance(ru) for pt in pts_c])
    cands["road_dist_m"] = np.round(road_dist_arr, 0)
    if not np.all(np.isnan(road_dist_arr)):
        before = len(cands)
        cands = cands[cands["road_dist_m"].isna()|(cands["road_dist_m"]<=ROAD_MAX_M)
                      ].copy().reset_index(drop=True)

    nc = len(cands)
    if nc == 0:
        print(f"  → 후보 0개 (건너뜀)")
        results_all.append({"site_id":sid,"clat":clat,"clon":clon,
                            "site_name":f"#{sid}","score_national":row.score,
                            "top":pd.DataFrame(),"ref_nearby":[]})
        continue

    pts_c = cands.geometry.tolist()

    # ── 점수 산정 ─────────────────────────────────────────────
    sl = cands["slope_deg"].values.astype(float)
    s2 = np.clip((1-sl/30)*15, 0, 15)

    d_pow = zone_dist(pts_c, zones["power"])
    d_rai = zone_dist(pts_c, zones["railway"])
    lp=np.log1p(d_pow); lr=np.log1p(d_rai)
    s3 = (lp/(lp.max() or 1)*0.5 + lr/(lr.max() or 1)*0.5)*15

    d_urb = zone_dist(pts_c, zones["urban_resid"])
    lu=np.log1p(d_urb)
    s4 = (lu/(lu.max() or 1))*15

    cands_wgs = cands.to_crs(WGS84_CRS)
    s5_arr=np.full(nc,6.0); p90_arr=np.full(nc,np.nan)
    for i in range(nc):
        glat=cands_wgs.geometry.iloc[i].y; glon=cands_wgs.geometry.iloc[i].x
        s5_arr[i],p90_arr[i] = kigam_s5(glat,glon)

    gz=zones.get("geology"); fz=zones.get("fault")
    hr=gz is not None and not gz.is_empty
    hf=fz is not None and not fz.is_empty
    lr_r=np.ones(nc); lr_f=np.ones(nc)
    d_rock_arr=np.full(nc,np.nan); d_flt_arr=np.full(nc,np.nan)
    if hr:
        dr=zone_dist(pts_c, gz.simplify(300))
        d_rock_arr=dr; lr_r=np.log1p(dr)
        cands["d_geo_rock_km"]=np.round(dr/1000,1)
    else: cands["d_geo_rock_km"]=np.nan
    if hf:
        df2=zone_dist(pts_c, fz.simplify(300))
        d_flt_arr=df2; lr_f=np.log1p(df2)
        cands["d_geo_fault_km"]=np.round(df2/1000,1)
    else: cands["d_geo_fault_km"]=np.nan
    if hr and hf: s6=(lr_r/(lr_r.max() or 1)*0.5+lr_f/(lr_f.max() or 1)*0.5)*5
    elif hr:      s6=(lr_r/(lr_r.max() or 1))*5
    elif hf:      s6=(lr_f/(lr_f.max() or 1))*5
    else:         s6=np.full(nc,2.5)

    cands["s2_지형"]=np.round(s2,1); cands["s3_전력철도"]=np.round(s3,1)
    cands["s4_인구이격"]=np.round(s4,1); cands["s5_모델기여도"]=np.round(s5_arr,1)
    cands["s6_암상"]=np.round(s6,1); cands["mag_p90p10_nT"]=np.round(p90_arr,1)
    cands["d_power_km"]=np.round(d_pow/1000,1); cands["d_railway_km"]=np.round(d_rai/1000,1)
    cands["d_urban_km"]=np.round(d_urb/1000,1)

    raw = s2+s3+s4+s5_arr+s6
    cands["score"]=np.round(raw/60*100,1)
    cands = cands.sort_values("score",ascending=False).reset_index(drop=True)
    cands["rank"]=cands.index+1

    top = cands.head(TOP_N).copy()
    tw = top.to_crs(WGS84_CRS)
    top["lat"]=tw.geometry.y.round(6); top["lon"]=tw.geometry.x.round(6)

    # ── 근접 기준점 검색 (5km 반경) ──────────────────────────
    ref_nearby = []
    if len(ref_utm_xy)>0:
        dists = np.sqrt((ref_utm_xy[:,0]-cx)**2 + (ref_utm_xy[:,1]-cy)**2)
        within = np.where(dists <= REF_RADIUS)[0]
        for idx in within:
            rp = ref_points[idx]
            ref_nearby.append({**rp, "dist_m":round(dists[idx],0)})
        ref_nearby.sort(key=lambda x: x["dist_m"])

    n_ref = len(ref_nearby)
    print(f"  → 후보{nc}→Top{len(top)}개 / 기준점{n_ref}개")

    results_all.append({
        "site_id": sid, "clat": clat, "clon": clon,
        "site_name": f"#{sid}", "score_national": row.score,
        "top": top, "ref_nearby": ref_nearby,
        "n_grid": n, "n_excl": n_excl, "n_cands": nc,
    })

elapsed1 = (datetime.now()-t0).seconds
print(f"\n  분석 완료 ({elapsed1//60}분 {elapsed1%60}초)")


# ════════════════════════════════════════════════════════════════
# KML 생성
# ════════════════════════════════════════════════════════════════
print("\n[3] KML 생성...")

def _v(val, fmt=".1f", unit=""):
    try:
        f = float(val)
        return "-" if math.isnan(f) else f"{f:{fmt}}{unit}"
    except: return str(val) if val else "-"

KML_STYLES = """
  <Style id="top1"><IconStyle><color>ff0000ff</color><scale>1.4</scale>
    <Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-stars.png</href></Icon>
  </IconStyle><LabelStyle><scale>1.1</scale></LabelStyle></Style>
  <Style id="top5"><IconStyle><color>ff00aaff</color><scale>1.0</scale>
    <Icon><href>http://maps.google.com/mapfiles/kml/paddle/orange-circle.png</href></Icon>
  </IconStyle><LabelStyle><scale>0.85</scale></LabelStyle></Style>
  <Style id="center"><IconStyle><color>ffff8800</color><scale>1.2</scale>
    <Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-circle.png</href></Icon>
  </IconStyle><LabelStyle><scale>0.9</scale></LabelStyle></Style>
  <Style id="tri"><IconStyle><color>ff00cc00</color><scale>0.8</scale>
    <Icon><href>http://maps.google.com/mapfiles/kml/shapes/triangle.png</href></Icon>
  </IconStyle><LabelStyle><scale>0.65</scale><color>ff007700</color></LabelStyle></Style>
  <Style id="uni"><IconStyle><color>ff0088ff</color><scale>0.75</scale>
    <Icon><href>http://maps.google.com/mapfiles/kml/shapes/donut.png</href></Icon>
  </IconStyle><LabelStyle><scale>0.65</scale><color>ff0066cc</color></LabelStyle></Style>
"""

lines = ['<?xml version="1.0" encoding="UTF-8"?>',
         '<kml xmlns="http://www.opengis.net/kml/2.2">',
         '<Document>',
         '  <name>1등급 후보지 63개 서브격자 + 기준점</name>',
         f'  <description>P1 63개 / 사이트당 상위{TOP_N}개 / 기준점 5km 반경 / {TS}</description>',
         KML_STYLES]

for res in results_all:
    sid   = res["site_id"]; clat=res["clat"]; clon=res["clon"]
    top   = res["top"]; refs=res["ref_nearby"]
    n_ref = len(refs); n_top=len(top)

    lines.append(f'  <Folder>')
    lines.append(f'    <name>#{sid} ({clat:.2f}°N {clon:.2f}°E) 국가{res["score_national"]}점</name>')
    lines.append(f'    <description>서브격자 상위{n_top}개 / 근접 기준점{n_ref}개</description>')

    # 원점 마커
    nu = f"https://map.naver.com/v5/search/{clat},{clon}"
    lines.append(f'    <Placemark><name>▶ P1 #{sid} 중심</name>')
    lines.append(f'      <description><![CDATA[P1 국가격자 후보지<br>점수: {res["score_national"]}점<br>'
                 f'좌표: {clat}°N, {clon}°E<br>'
                 f'서브격자 후보: {n_top}개 / 근접 기준점: {n_ref}개<br><br>'
                 f'<a href="{nu}">📍 네이버 지도</a>]]></description>')
    lines.append(f'      <styleUrl>#center</styleUrl>')
    lines.append(f'      <Point><coordinates>{clon},{clat},0</coordinates></Point>')
    lines.append(f'    </Placemark>')

    # 서브격자 후보
    if n_top > 0:
        lines.append(f'    <Folder><name>📍 서브격자 후보 ({n_top}개)</name>')
        for _, row in top.iterrows():
            rnk=int(row["rank"]); sc=float(row["score"])
            lat=float(row["lat"]); lon=float(row["lon"])
            style = "top1" if rnk==1 else "top5"
            nu2 = f"https://map.naver.com/v5/search/{lat},{lon}"
            ku2 = f"https://map.kakao.com/link/map/P{sid}-{rnk},{lat},{lon}"
            desc = (f"사이트 #{sid} / 서브격자 {rnk}위\n"
                    f"종합점수: {sc:.1f}점\n"
                    f"좌표: {lat:.6f}°N, {lon:.6f}°E\n\n"
                    f"② 지형: {_v(row.get('s2_지형'))} / 15점 (경사 {_v(row.get('slope_deg'),'','')}°)\n"
                    f"③ 전력·철도: {_v(row.get('s3_전력철도'))} / 15점\n"
                    f"④ 인구이격: {_v(row.get('s4_인구이격'))} / 15점\n"
                    f"⑤ 모델기여: {_v(row.get('s5_모델기여도'))} / 10점\n"
                    f"⑥ 암상: {_v(row.get('s6_암상'))} / 5점\n"
                    f"도로: {_v(row.get('road_dist_m'),'.0f','m')}\n\n"
                    f"네이버: {nu2}\n카카오: {ku2}")
            lines.append(f'      <Placemark><name>#{rnk} {sc:.1f}점</name>')
            lines.append(f'        <description><![CDATA[{desc}]]></description>')
            lines.append(f'        <styleUrl>#{style}</styleUrl>')
            lines.append(f'        <Point><coordinates>{lon},{lat},0</coordinates></Point>')
            lines.append(f'      </Placemark>')
        lines.append(f'    </Folder>')

    # 근접 기준점
    if n_ref > 0:
        lines.append(f'    <Folder><name>📐 근접 기준점 ({n_ref}개)</name>')
        for rp in refs:
            cat=rp["category"]; nm=rp["name"] or cat
            rlat=rp["lat"]; rlon=rp["lon"]; dm=rp["dist_m"]
            style = "tri" if cat=="삼각점" else "uni"
            icon  = "△" if cat=="삼각점" else "◎"
            nu3 = f"https://map.naver.com/v5/search/{rlat},{rlon}"
            lines.append(f'      <Placemark><name>{icon} {nm}</name>')
            lines.append(f'        <description><![CDATA['
                         f'<b>{cat}</b><br>명칭: {nm}<br>'
                         f'P1 #{sid} 원점에서 {dm:.0f}m<br>'
                         f'좌표: {rlat:.6f}°N, {rlon:.6f}°E<br><br>'
                         f'<a href="{nu3}">📍 네이버 지도</a>]]></description>')
            lines.append(f'        <styleUrl>#{style}</styleUrl>')
            lines.append(f'        <Point><coordinates>{rlon},{rlat},0</coordinates></Point>')
            lines.append(f'      </Placemark>')
        lines.append(f'    </Folder>')

    lines.append(f'  </Folder>')

lines += ['</Document>', '</kml>']

KML_PATH = OUT_DIR / f"{TS}_P1_63sites_subgrid_기준점.kml"
KML_PATH.write_text("\n".join(lines), encoding="utf-8")
print(f"  ✅ KML: {KML_PATH.name}  ({KML_PATH.stat().st_size/1024:.0f} KB)")


# ════════════════════════════════════════════════════════════════
# Excel 생성
# ════════════════════════════════════════════════════════════════
print("\n[4] Excel 생성...")

def hf(hex): return PatternFill("solid", fgColor=hex)
def tb():
    s=Side(style="thin",color="BFBFBF")
    return Border(left=s,right=s,top=s,bottom=s)

C_H="#1F4E79"; C_L="#D6DCE4"; C_G="#E2EFDA"; C_Y="#FFF2CC"; C_B="#DEEAF1"

def hl(ws, r, c, url, label, bg):
    cell = ws.cell(row=r, column=c)
    cell.value = f'=HYPERLINK("{url}","{label}")'
    cell.font = Font(name="맑은 고딕",size=9,color="0563C1",underline="single")
    cell.fill = hf(bg); cell.alignment = Alignment(horizontal="center")
    cell.border = tb()

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── 시트1: 전체 요약 ─────────────────────────────────────────
ws1 = wb.create_sheet("📋 전체 요약")
ws1.merge_cells("A1:N1")
ws1["A1"] = f"1등급 P1 후보지 63개 — 서브격자 최우선 후보 + 기준점 매칭 요약"
ws1["A1"].font = Font(name="맑은 고딕",bold=True,size=13,color="FFFFFF")
ws1["A1"].fill = hf(C_H[1:])
ws1["A1"].alignment = Alignment(horizontal="center",vertical="center")
ws1.row_dimensions[1].height = 26
ws1.merge_cells("A2:N2")
ws1["A2"] = (f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  "
             f"/ 반경{RADIUS_M//1000}km / 격자{GRID_STEP//1000}km / 도로≤{ROAD_MAX_M//1000}km "
             f"/ 기준점반경{REF_RADIUS//1000}km / 사이트당 상위{TOP_N}개")
ws1["A2"].font = Font(name="맑은 고딕",size=9,italic=True)

hdrs1 = [
    ("사이트",8),("국가점수",8),("서브#1위도",11),("서브#1경도",11),
    ("서브점수",8),("②지형",7),("③전력철도",8),("④인구",7),
    ("⑤모델기여",8),("⑥암상",6),("도로(m)",7),
    ("기준점수",7),("네이버(#1)",18),("카카오(#1)",18),
]
for ci,(h,w) in enumerate(hdrs1,start=1):
    c=ws1.cell(row=3,column=ci,value=h)
    c.font=Font(name="맑은 고딕",bold=True,size=9,color="FFFFFF")
    c.fill=hf(C_H[1:]); c.alignment=Alignment(horizontal="center",wrap_text=True)
    c.border=tb(); ws1.column_dimensions[get_column_letter(ci)].width=w
ws1.row_dimensions[3].height=28

r1=4
for res in results_all:
    sid=res["site_id"]; sc_nat=res["score_national"]
    top=res["top"]; n_ref=len(res["ref_nearby"])
    bg = C_G if n_ref>=3 else (C_Y if n_ref>=1 else "FFFFFF")

    if len(top)==0:
        vals=[f"#{sid}",sc_nat,"-","-","-","-","-","-","-","-","-",n_ref]
        for ci,v in enumerate(vals,start=1):
            c=ws1.cell(row=r1,column=ci,value=v)
            c.font=Font(name="맑은 고딕",size=9); c.fill=hf("F2F2F2")
            c.alignment=Alignment(horizontal="center"); c.border=tb()
        r1+=1; continue

    tr=top.iloc[0]
    lat1=float(tr["lat"]); lon1=float(tr["lon"]); sc1=float(tr["score"])
    def _vv(k,fmt=".1f"):
        try:
            v=float(tr.get(k,float("nan")))
            return "-" if math.isnan(v) else f"{v:{fmt}}"
        except: return "-"
    vals=[f"#{sid}",sc_nat,lat1,lon1,sc1,
          _vv("s2_지형"),_vv("s3_전력철도"),_vv("s4_인구이격"),
          _vv("s5_모델기여도"),_vv("s6_암상"),_vv("road_dist_m",".0f"),n_ref]
    for ci,v in enumerate(vals,start=1):
        c=ws1.cell(row=r1,column=ci,value=v)
        c.font=Font(name="맑은 고딕",size=9,bold=(ci in [1,5]))
        c.fill=hf(bg[1:] if bg.startswith("#") else bg)
        c.alignment=Alignment(horizontal="center"); c.border=tb()

    nu=f"https://map.naver.com/v5/search/{lat1},{lon1}"
    ku=f"https://map.kakao.com/link/map/P{sid}-1,{lat1},{lon1}"
    hl(ws1,r1,13,nu,f"📍 네이버 #{sid}",bg[1:] if bg.startswith("#") else bg)
    hl(ws1,r1,14,ku,f"📍 카카오 #{sid}",bg[1:] if bg.startswith("#") else bg)
    r1+=1

# 범례
leg=r1+1; ws1.merge_cells(f"A{leg}:N{leg}")
ws1[f"A{leg}"]="🟢 기준점 3개↑  🟡 기준점 1~2개  ⬜ 기준점 없음  │ 서브격자 #1위 기준 표시"
ws1[f"A{leg}"].font=Font(name="맑은 고딕",size=8,italic=True,color="595959")
ws1[f"A{leg}"].fill=hf(C_L[1:])
ws1.freeze_panes="A4"

# ── 시트2: 전체 서브격자 후보 ────────────────────────────────
ws2 = wb.create_sheet("📍 서브격자 후보")
ws2.merge_cells("A1:M1")
ws2["A1"]="전체 사이트 서브격자 상위 후보 (사이트당 상위 5개)"
ws2["A1"].font=Font(name="맑은 고딕",bold=True,size=12,color="FFFFFF")
ws2["A1"].fill=hf(C_H[1:]); ws2["A1"].alignment=Alignment(horizontal="center")
ws2.row_dimensions[1].height=24

hdrs2=[("사이트",7),("순위",5),("위도",11),("경도",11),("점수",7),
       ("②지형",7),("③전력철도",8),("④인구",7),("⑤모델기여",8),
       ("⑥암상",6),("도로(m)",7),("네이버",18),("카카오",18)]
for ci,(h,w) in enumerate(hdrs2,start=1):
    c=ws2.cell(row=2,column=ci,value=h)
    c.font=Font(name="맑은 고딕",bold=True,size=9,color="FFFFFF")
    c.fill=hf(C_H[1:]); c.alignment=Alignment(horizontal="center")
    c.border=tb(); ws2.column_dimensions[get_column_letter(ci)].width=w
ws2.row_dimensions[2].height=22

r2=3
for res in results_all:
    sid=res["site_id"]; top=res["top"]
    if len(top)==0: continue
    for _,row in top.iterrows():
        rnk=int(row["rank"]); sc=float(row["score"])
        lat=float(row["lat"]); lon=float(row["lon"])
        bg = C_Y[1:] if rnk==1 else ("FFFFFF")
        def _vv2(k,fmt=".1f"):
            try:
                v=float(row.get(k,float("nan")))
                return "-" if math.isnan(v) else f"{v:{fmt}}"
            except: return "-"
        vals=[f"#{sid}",rnk,lat,lon,sc,
              _vv2("s2_지형"),_vv2("s3_전력철도"),_vv2("s4_인구이격"),
              _vv2("s5_모델기여도"),_vv2("s6_암상"),_vv2("road_dist_m",".0f")]
        for ci,v in enumerate(vals,start=1):
            c=ws2.cell(row=r2,column=ci,value=v)
            c.font=Font(name="맑은 고딕",size=9,bold=(ci in [1,5]))
            c.fill=hf(bg); c.alignment=Alignment(horizontal="center"); c.border=tb()
        nu=f"https://map.naver.com/v5/search/{lat},{lon}"
        ku=f"https://map.kakao.com/link/map/P{sid}-{rnk},{lat},{lon}"
        hl(ws2,r2,12,nu,f"📍 네이버",bg)
        hl(ws2,r2,13,ku,f"📍 카카오",bg)
        r2+=1
ws2.freeze_panes="A3"

# ── 시트3: 기준점 매칭 ───────────────────────────────────────
ws3 = wb.create_sheet("📐 기준점 매칭")
ws3.merge_cells("A1:H1")
ws3["A1"]="사이트별 근접 기준점 목록 (5km 반경)"
ws3["A1"].font=Font(name="맑은 고딕",bold=True,size=12,color="FFFFFF")
ws3["A1"].fill=hf(C_H[1:]); ws3["A1"].alignment=Alignment(horizontal="center")
ws3.row_dimensions[1].height=24

hdrs3=[("사이트",7),("종류",8),("기준점명",14),("거리(m)",8),
       ("위도",11),("경도",11),("네이버",18),("카카오",18)]
for ci,(h,w) in enumerate(hdrs3,start=1):
    c=ws3.cell(row=2,column=ci,value=h)
    c.font=Font(name="맑은 고딕",bold=True,size=9,color="FFFFFF")
    c.fill=hf(C_H[1:]); c.alignment=Alignment(horizontal="center")
    c.border=tb(); ws3.column_dimensions[get_column_letter(ci)].width=w
ws3.row_dimensions[2].height=22

r3=3
for res in results_all:
    sid=res["site_id"]; refs=res["ref_nearby"]
    if not refs: continue
    # 사이트 소제목
    ws3.merge_cells(f"A{r3}:H{r3}")
    c=ws3[f"A{r3}"]
    c.value=f"▶ P1 #{sid}  ({res['clat']:.3f}°N, {res['clon']:.3f}°E) / 국가점수 {res['score_national']}점 / 기준점 {len(refs)}개"
    c.font=Font(name="맑은 고딕",bold=True,size=10,color="FFFFFF")
    c.fill=hf("2F75B6"); ws3.row_dimensions[r3].height=18; r3+=1

    for rp in refs:
        cat=rp["category"]; nm=rp["name"] or "-"
        rlat=rp["lat"]; rlon=rp["lon"]; dm=rp["dist_m"]
        bg = C_G[1:] if cat=="삼각점" else C_B[1:]
        vals=[f"#{sid}",cat,nm,dm,rlat,rlon]
        for ci,v in enumerate(vals,start=1):
            c=ws3.cell(row=r3,column=ci,value=v)
            c.font=Font(name="맑은 고딕",size=9); c.fill=hf(bg)
            c.alignment=Alignment(horizontal="center"); c.border=tb()
        nu=f"https://map.naver.com/v5/search/{rlat},{rlon}"
        ku=f"https://map.kakao.com/link/map/{cat}{nm},{rlat},{rlon}"
        hl(ws3,r3,7,nu,f"📍 네이버",bg)
        hl(ws3,r3,8,ku,f"📍 카카오",bg)
        r3+=1
    r3+=1
ws3.freeze_panes="A3"

XL_PATH = OUT_DIR / f"{TS}_P1_63sites_subgrid_기준점.xlsx"
wb.save(XL_PATH)
print(f"  ✅ Excel: {XL_PATH.name}  ({XL_PATH.stat().st_size/1024:.0f} KB)")

# ── 최종 요약 ──────────────────────────────────────────────────
elapsed = (datetime.now()-t0).seconds
total_top = sum(len(r["top"]) for r in results_all)
total_ref = sum(len(r["ref_nearby"]) for r in results_all)
sites_with_ref = sum(1 for r in results_all if r["ref_nearby"])

print("\n" + "="*66)
print(f"  ✅ 완료  ({elapsed//60}분 {elapsed%60}초)")
print(f"  KML  : {KML_PATH.name}")
print(f"  Excel: {XL_PATH.name}")
print(f"  저장 : {OUT_DIR}")
print()
print(f"  ── 결과 요약 ────────────────────────────────────")
print(f"  P1 사이트 : 63개")
print(f"  서브격자 후보 : {total_top}개 (사이트당 최대 {TOP_N}개)")
print(f"  근접 기준점 총합 : {total_ref}개")
print(f"  기준점 있는 사이트 : {sites_with_ref}/63개")
print("="*66)
