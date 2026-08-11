# -*- coding: utf-8 -*-
"""
현장 조사서(PDF) → 취합 엑셀
=============================

각 본부가 회신한 「자기교란 현황 조사 카드」 PDF 를 읽어 한 카드=한 행으로 취합한다.

  · 도상선점 기본정보(관리번호·좌표·도엽·주소·간섭요인 등)는 총괄 양식 xlsx 에서
    관리번호로 조인 → 원본 수치를 그대로 사용(파싱 오류 배제).
  · 현장 조사 결과(소유권·조사일·조사자·기상·자기교란 8항목·판정·의견)는
    카드 PDF 에서 좌표 기반으로 파싱.

카드는 고정 레이아웃이므로 라벨 토큰을 앵커로, 값은 정해진 x-구간에서 읽는다.

사용:
    python aggregate_survey_reports.py                # 기본 폴더 전체
    python aggregate_survey_reports.py --pdf "1. 서울 경기북.pdf"
    python aggregate_survey_reports.py --pdf-dir <폴더> --master <총괄.xlsx> --out <파일.xlsx>
"""
import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
DEF_PDF_DIR = Path(r"D:\LX_yoons\2026_research\2026_지자기 연구\20260811_사전 현장 조사서 취합")
DEF_MASTER = Path(r"D:\LX_yoons\2026_research\2026_지자기 연구\20260707_지자기 사전 현장조사 엑셀"
                  r"\20260707_조사카드_v2_총괄_전체103건.xlsx")
OUT_DIR = ROOT / "docs" / "output"

DIST_ITEMS = ["차량", "전력시설", "통신시설", "금속류", "매설물", "건축물", "자연환경"]
DIST_LABEL = {"차량": "차량영향", "전력시설": "전력시설", "통신시설": "통신시설",
              "금속류": "금속류", "매설물": "매설물", "건축물": "건축물", "자연환경": "자연환경"}


# ── PDF 파싱 ──────────────────────────────────────────────────────────────────
def rows_of(page):
    """단어들을 시각적 행으로 묶는다(y 근접 4.5pt)."""
    ws = sorted(page.get_text("words"), key=lambda w: (w[1], w[0]))
    rows, cur, cy = [], [], None
    for w in ws:
        if cy is None or abs(w[1] - cy) <= 4.5:
            cur.append(w)
            cy = w[1] if cy is None else cy
        else:
            rows.append(cur)
            cur, cy = [w], w[1]
    if cur:
        rows.append(cur)
    return rows


def is_card(page):
    return "자기교란 현황 조사 카드" in page.get_text()


def txt(row, x0, x1):
    return " ".join(w[4] for w in sorted(row, key=lambda w: w[0]) if x0 <= w[0] < x1).strip()


def rows_with_left(rows, needle, xmax=85):
    """왼쪽 라벨칸(x<xmax)에 needle 을 포함하는 행."""
    return [r for r in rows if needle in "".join(w[4] for w in r if w[0] < xmax)]


def parse_card(page):
    rows = rows_of(page)
    d = {}

    # 관리번호 · 후보지명 (제목행)
    for r in rows:
        j = " ".join(w[4] for w in sorted(r, key=lambda w: w[0]))
        m = re.search(r"DS-\d{3}", j)
        if m and "조사 카드" in j:
            d["관리번호"] = m.group(0)
            name = j.split(m.group(0), 1)[1]
            name = re.sub(r"\s*현장\s*위치.*$", "", name).strip()  # 우측 지도 헤더 제거
            d["후보지명"] = name
            break

    # 소유권
    for r in rows_with_left(rows, "소유권"):
        for v in ("국공유지", "사유지"):
            if v in " ".join(w[4] for w in r):
                d["소유권"] = v

    # 조사일 (YYYY-MM-DD; '(예: 20260707)' 은 하이픈 없어 제외됨)
    for r in rows:
        m = re.search(r"20\d{2}-\d{2}-\d{2}", " ".join(w[4] for w in r))
        if m:
            d["조사일"] = m.group(0)
            break

    # 조사자 — 라벨 토큰이 정확히 '조사자'(괄호 없는) 인 행. '의견(조사자)' 행 제외.
    for r in rows:
        if "의견" in " ".join(w[4] for w in r if w[0] < 85):
            continue
        if any(w[4] == "조사자" and w[0] < 85 for w in r):
            v = txt(r, 95, 150)
            if v:
                d["조사자"] = v
                break

    # 기상
    for r in rows:
        line = " ".join(w[4] for w in r)
        if "기상" in "".join(w[4] for w in r if w[0] < 200):
            for v in ("눈 또는 비", "맑음", "흐림", "눈", "비"):
                if v in line:
                    d["기상"] = v
                    break

    # 자기교란 8항목: 라벨 앵커 → 존재여부 x[150,172], 실측이격 x[180,205]
    dist = {}
    for name in DIST_ITEMS:
        for r in rows_with_left(rows, name):
            ex = [v for v in ("있음", "없음") if v in txt(r, 150, 172)]
            if ex:
                dist[name] = {"존재": ex[0], "이격": txt(r, 180, 205)}
                break
    d["자기교란"] = dist

    # 방위표지 가능여부 — x[150,172] 의 가능/불가 (라벨 detail '설치 가능 여부' 회피)
    for r in rows_with_left(rows, "방위표지"):
        v = txt(r, 150, 172)
        if "가능" in v:
            d["방위표지"] = "가능"
        elif "불가" in v:
            d["방위표지"] = "불가"

    # 판정 집계
    for r in rows:
        m = re.search(r"적합\s*(\d+)건\s*/\s*부적합\s*(\d+)건\s*/\s*미입력\s*(\d+)건",
                      " ".join(w[4] for w in r))
        if m:
            d["집계"] = f"적합 {m.group(1)} / 부적합 {m.group(2)} / 미입력 {m.group(3)}"
            break

    # 종합 판정 — '판정 결과' 행
    for r in rows_with_left(rows, "결과"):
        v = txt(r, 150, 320)
        for cand in ("조건부 적합", "부적합", "적합"):
            if cand in v:
                d["종합판정"] = cand
                break
        if "종합판정" in d:
            break

    # 향후 조치
    for r in rows_with_left(rows, "향후"):
        v = txt(r, 140, 320)
        if v:
            d["향후조치"] = v
            break

    # 판정 의견(조사자 자유서술) — '판정 의견' 라벨행 + 다음행
    for i, r in enumerate(rows):
        left = "".join(w[4] for w in r if w[0] < 85)
        if "판정" in left and "의견" in left:
            op = [txt(r, 85, 330)]
            if i + 1 < len(rows):
                op.append(txt(rows[i + 1], 85, 330))
            d["판정의견"] = " ".join(x for x in op if x).strip()
            break

    return d


def parse_pdf(path):
    doc = fitz.open(path)
    out = []
    for i in range(doc.page_count):
        if is_card(doc[i]):
            d = parse_card(doc[i])
            if d.get("관리번호"):
                d["_page"] = i + 1
                out.append(d)
    return out


# ── 총괄 양식 조인 ────────────────────────────────────────────────────────────
def load_master(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["총괄 목록"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[1])
    idx = {h: hdr.index(h) for h in hdr if h}
    out = {}
    for r in rows[2:]:
        if r[0]:
            out[str(r[0]).strip()] = {h: r[idx[h]] for h in idx}
    return out


# ── 엑셀 출력 ─────────────────────────────────────────────────────────────────
COLS = ["지역(파일)", "관할 본부", "관리번호", "후보지명", "도엽번호", "도엽명",
        "지번 주소", "위도", "경도", "표고(m)", "유형", "연계 기존점", "기존점거리(km)",
        "최근접 송전탑(m)", "최근접 철도(km)", "최근접 주거지(km)",
        "소유권", "조사일", "조사자", "기상",
        "차량영향", "전력시설", "통신시설", "금속류", "매설물", "건축물", "자연환경",
        "방위표지", "판정 집계", "종합 판정", "향후 조치", "판정 의견", "카드페이지"]

F_TITLE = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")
F_HDR = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
F_VAL = Font(name="맑은 고딕", size=9)
FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_HDR = PatternFill("solid", fgColor="4472C4")
FILL_BAD = PatternFill("solid", fgColor="FDE0DC")   # 있음/부적합
FILL_COND = PatternFill("solid", fgColor="FFF2CC")  # 조건부
FILL_OK = PatternFill("solid", fgColor="E2EFDA")    # 적합
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def dist_cell(item):
    if not item:
        return ""
    if item["존재"] == "있음":
        gap = item.get("이격", "").strip()
        return f"있음 ({gap}m)" if gap and gap != "-" else "있음"
    return item["존재"]


def build_row(region, card, m):
    dist = card.get("자기교란", {})
    g = lambda k: (m.get(k) if m else "") or ""
    return [
        region, g("관할 본부"), card.get("관리번호", ""), g("후보지명") or card.get("후보지명", ""),
        g("도엽번호"), g("도엽명"), g("지번 주소"), g("위도"), g("경도"), g("표고(m)"),
        g("유형"), g("연계 기존점"), g("기존점거리(km)"),
        g("최근접 송전탑(m)"), g("최근접 철도(km)"), g("최근접 주거지(km)"),
        card.get("소유권", ""), card.get("조사일", ""), card.get("조사자", ""), card.get("기상", ""),
        dist_cell(dist.get("차량")), dist_cell(dist.get("전력시설")), dist_cell(dist.get("통신시설")),
        dist_cell(dist.get("금속류")), dist_cell(dist.get("매설물")), dist_cell(dist.get("건축물")),
        dist_cell(dist.get("자연환경")), card.get("방위표지", ""), card.get("집계", ""),
        card.get("종합판정", ""), card.get("향후조치", ""), card.get("판정의견", ""),
        card.get("_page", ""),
    ]


def write_excel(records, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "현장조사 취합"
    ncol = len(COLS)
    ws.merge_cells(f"A1:{get_column_letter(ncol)}1")
    c = ws["A1"]
    c.value = f"지자기 도상선점 — 현장 조사서 취합 ({len(records)}건)   ·   {datetime.now():%Y-%m-%d}"
    c.font = F_TITLE
    c.fill = FILL_TITLE
    c.alignment = AL_L
    ws.row_dimensions[1].height = 26

    for j, h in enumerate(COLS, 1):
        cc = ws.cell(2, j, h)
        cc.font = F_HDR
        cc.fill = FILL_HDR
        cc.alignment = AL_C
        cc.border = BORDER

    dist_start = COLS.index("차량영향") + 1
    dist_end = COLS.index("자연환경") + 1
    bang_col = COLS.index("방위표지") + 1
    judge_col = COLS.index("종합 판정") + 1

    for ri, row in enumerate(records, 3):
        for j, v in enumerate(row, 1):
            cc = ws.cell(ri, j, v)
            cc.font = F_VAL
            cc.border = BORDER
            cc.alignment = AL_L if j in (7, ncol - 1) else AL_C
            if dist_start <= j <= dist_end and isinstance(v, str) and v.startswith("있음"):
                cc.fill = FILL_BAD
            if j == bang_col and v == "불가":
                cc.fill = FILL_BAD
            if j == judge_col:
                if v == "적합":
                    cc.fill = FILL_OK
                elif v == "조건부 적합":
                    cc.fill = FILL_COND
                elif v == "부적합":
                    cc.fill = FILL_BAD

    widths = [13, 14, 8, 12, 11, 8, 24, 9, 9, 7, 8, 9, 8,
              12, 11, 12, 8, 11, 12, 7,
              9, 8, 8, 8, 8, 8, 8, 8, 16, 10, 16, 34, 7]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{len(records) + 2}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", help="단일 PDF 파일")
    ap.add_argument("--pdf-dir", default=str(DEF_PDF_DIR), help="PDF 폴더")
    ap.add_argument("--master", default=str(DEF_MASTER), help="총괄 양식 xlsx")
    ap.add_argument("--out", help="출력 xlsx 경로")
    a = ap.parse_args()

    master = load_master(a.master)
    print(f"총괄 양식 로딩: {len(master)}건")

    if a.pdf:
        pdfs = [Path(a.pdf) if Path(a.pdf).is_absolute() else DEF_PDF_DIR / a.pdf]
    else:
        pdfs = sorted(Path(a.pdf_dir).glob("*.pdf"))
    if not pdfs:
        print("PDF 를 찾지 못했습니다.")
        return

    records, miss = [], []
    for pdf in pdfs:
        region = re.sub(r"^\d+\.\s*", "", pdf.stem)
        cards = parse_pdf(pdf)
        print(f"  {pdf.name}: 카드 {len(cards)}건")
        for card in cards:
            mid = card["관리번호"]
            m = master.get(mid)
            if not m:
                miss.append((pdf.name, mid))
            records.append(build_row(region, card, m))

    records.sort(key=lambda r: (r[1] or "", r[2] or ""))  # 본부, 관리번호

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(a.out) if a.out else OUT_DIR / f"{ts}_현장조사_취합_{len(records)}건.xlsx"
    write_excel(records, out)
    print(f"\n취합 완료: {out}  ({len(records)}건)")
    if miss:
        print(f"[주의] 총괄 미매칭 {len(miss)}건: {miss}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
