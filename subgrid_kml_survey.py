#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
후보지 #16 · #18  1km 서브격자 현장 답사 세부 선정
출력: KML 파일 (Google Earth / 네이버·카카오 지도 임포트용)
       Excel (네이버·카카오 지도 링크 포함)

흐름:
  1. 1km 서브격자 생성 (각 후보지 중심 ±5km)
  2. Open-Elevation API 경사도 산정
  3. zone_cache WKB 재적용 → 불적합점 제거
  4. 도로 접근성 필터 (secondary+ ≤ 2km)
  5. s2~s6 점수 산정 → 상위 10개 선정
  6. KML 저장 (두 후보지 합본 + 개별)
  7. Excel 저장 (네이버·카카오 지도 링크)
"""
import sys, json, warnings
import numpy as np
import pandas as pd
import geopandas as gpd
import shapely
from shapely.geometry import Point, LineString, box
from shapely.prepared import prep
from pathlib import Path
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
warnings.filterwarnings("ignore")

# ─── 경로 ─────────────────────────────────────────────────────────
MAIN_DIR  = Path("C:/LG_gram_backup_users/LX/2026_geomag")
DATA_DIR  = MAIN_DIR / "data"
CACHE_DIR = DATA_DIR / "zone_cache"
OUT_DIR   = MAIN_DIR / "docs" / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)

UTM_CRS   = "EPSG:5179"
WGS84_CRS = "EPSG:4326"
TS        = datetime.now().strftime("%Y%m%d_%H%M%S")

# ─── 분석 파라미터 ────────────────────────────────────────────────
RADIUS_M   = 5_000   # 탐색 반경 (m)
GRID_STEP  = 1_000   # 서브격자 간격 (m)
ROAD_MAX_M = 2_000   # 차량 접근 최대 거리 (m)
TOP_N      = 10      # 사이트당 상위 후보 수

SITES = [
    {"id": 16, "name": "후보지 #16 (전주 완주)",
     "lat": 35.59196, "lon": 127.15795,
     "score_national": 60.1, "priority": 1},
    {"id": 18, "name": "후보지 #18 (진안 무주)",
     "lat": 35.86272, "lon": 127.26755,
     "score_national": 58.8, "priority": 1},
]

ZONE_NAMES = ["power", "railway", "urban_dense", "urban_resid",
              "pipeline", "comm", "wind", "quarry",
              "water", "geology", "fault", "anomaly"]

print("=" * 64)
print("  측정 후보지 #16·#18 — 1km 서브격자 현장 답사 세부 선정")
print(f"  반경 {RADIUS_M//1000}km / 격자간격 {GRID_STEP//1000}km / 도로≤{ROAD_MAX_M//1000}km")
print("=" * 64)


# ════════════════════════════════════════════════════════════════
# 유틸리티
# ════════════════════════════════════════════════════════════════
def load_zone(name, local_box):
    """WKB 캐시 → 로컬 영역 클립"""
    p = CACHE_DIR / f"{name}.wkb"
    if not p.exists():
        return None
    g = shapely.from_wkb(p.read_bytes())
    try:
        local = g.intersection(local_box)
        return None if local.is_empty else local
    except Exception:
        return g


def zone_dist(pts_list, zone_geom, simplify_m=300):
    """zone geometry 까지 최소 거리 배열 반환"""
    if zone_geom is None or zone_geom.is_empty:
        return np.full(len(pts_list), 50_000.0)
    zs = zone_geom.simplify(simplify_m)
    return np.array([pt.distance(zs) for pt in pts_list])


def fetch_elev_batch(latlons):
    """Open-Elevation API 배치 요청"""
    url = "https://api.open-elevation.com/api/v1/lookup"
    results = [None] * len(latlons)
    for start in range(0, len(latlons), 200):
        batch = latlons[start:start + 200]
        payload = {"locations": [{"latitude": la, "longitude": lo}
                                  for la, lo in batch]}
        try:
            r = requests.post(url, json=payload, timeout=90)
            if r.status_code == 200:
                for j, item in enumerate(r.json()["results"]):
                    results[start + j] = item.get("elevation")
        except Exception as e:
            print(f"    ⚠ API 오류: {e}")
    return results


def kigam_score(lat, lon, klat, klon, kanom, radius=0.05):
    """KIGAM P90-P10 → 모델기여도 점수"""
    mask = (np.abs(klat - lat) <= radius) & (np.abs(klon - lon) <= radius)
    pts_k = kanom[mask]
    if len(pts_k) < 3:
        return 6.0, np.nan
    v = float(np.percentile(pts_k, 90) - np.percentile(pts_k, 10))
    if   v <  30:  return 5.0, round(v, 1)
    elif v <= 150: return 8.0, round(v, 1)
    elif v <= 400: return 10.0, round(v, 1)
    elif v <= 800: return 7.0, round(v, 1)
    else:          return 0.0, round(v, 1)


# KIGAM 데이터 로드 (공통)
kigam_path = DATA_DIR / "mag_1982-2018_1.5min_ed.dat"
klat_arr = klon_arr = kanom_arr = None
if kigam_path.exists():
    kdf = pd.read_csv(kigam_path, sep=r"\s+", skiprows=9, header=None,
                      names=["lon", "lat", "anomaly_nT"],
                      na_values=["99999", "-99999"], on_bad_lines="skip")
    kdf = kdf.apply(pd.to_numeric, errors="coerce").dropna()
    klat_arr  = kdf.lat.values
    klon_arr  = kdf.lon.values
    kanom_arr = kdf.anomaly_nT.values
    print(f"  KIGAM 로드: {len(kdf)}개 격자점")

# 도로 데이터 로드 (공통)
road_json = DATA_DIR / "access_roads.json"
road_gdf_wgs = None
if road_json.exists():
    with open(road_json, encoding="utf-8") as f:
        rd = json.load(f)
    nodes_rd = {e["id"]: (e["lon"], e["lat"]) for e in rd.get("elements", [])
                if e["type"] == "node"}
    road_lines = []
    for e in rd.get("elements", []):
        if e["type"] == "way":
            coords = [nodes_rd[n] for n in e.get("nodes", []) if n in nodes_rd]
            if len(coords) >= 2:
                road_lines.append(LineString(coords))
    if road_lines:
        road_gdf_wgs = gpd.GeoDataFrame(geometry=road_lines, crs=WGS84_CRS)
        print(f"  도로 세그먼트: {len(road_gdf_wgs)}개")
else:
    print("  ⚠ access_roads.json 없음 — 도로 접근성 필터 생략")


# ════════════════════════════════════════════════════════════════
# 사이트별 분석
# ════════════════════════════════════════════════════════════════
all_results = []   # 두 사이트 결과 합산

for site in SITES:
    sid    = site["id"]
    sname  = site["name"]
    clat   = site["lat"]
    clon   = site["lon"]
    snatio = site["score_national"]

    print(f"\n{'═'*64}")
    print(f"  분석: {sname}  ({clat}°N, {clon}°E, 국가격자 {snatio}점)")
    print(f"{'═'*64}")

    # ── 1. 서브격자 생성 ───────────────────────────────────────
    print("\n[1] 서브격자 생성...")
    center_gdf = gpd.GeoDataFrame(geometry=[Point(clon, clat)],
                                   crs=WGS84_CRS).to_crs(UTM_CRS)
    cx = center_gdf.geometry.iloc[0].x
    cy = center_gdf.geometry.iloc[0].y

    xs = np.arange(cx - RADIUS_M, cx + RADIUS_M + 1, GRID_STEP)
    ys = np.arange(cy - RADIUS_M, cy + RADIUS_M + 1, GRID_STEP)
    center_pt = Point(cx, cy)
    pts_utm = [Point(x, y) for x in xs for y in ys
               if Point(x, y).distance(center_pt) <= RADIUS_M]
    n_all = len(pts_utm)
    LOCAL_BOX = box(cx - RADIUS_M * 1.2, cy - RADIUS_M * 1.2,
                    cx + RADIUS_M * 1.2, cy + RADIUS_M * 1.2)

    gdf_all = gpd.GeoDataFrame({"gid": range(n_all)},
                                geometry=pts_utm, crs=UTM_CRS)
    gdf_wgs = gdf_all.to_crs(WGS84_CRS)
    gdf_all["lat"] = gdf_wgs.geometry.y.round(6)
    gdf_all["lon"] = gdf_wgs.geometry.x.round(6)
    print(f"  격자점: {n_all}개")

    # ── 2. 지형 경사도 ─────────────────────────────────────────
    print("\n[2] 지형 경사도 (Open-Elevation API)...")
    ELEV_OFFSET = 500
    DIRS = {"C": (0, 0), "N": (0, ELEV_OFFSET),
            "S": (0, -ELEV_OFFSET), "E": (ELEV_OFFSET, 0), "W": (-ELEV_OFFSET, 0)}
    ELEV_CACHE = OUT_DIR / f"site{sid}_elev_cache.json"

    if ELEV_CACHE.exists():
        print(f"  캐시 로드: {ELEV_CACHE.name}")
        with open(ELEV_CACHE) as f:
            elev_dict = {int(k): v for k, v in json.load(f).items()}
    else:
        all_pts_utm, gi_list, dir_list = [], [], []
        for gi in range(n_all):
            xc = gdf_all.geometry.iloc[gi].x
            yc = gdf_all.geometry.iloc[gi].y
            for dname, (dx, dy) in DIRS.items():
                all_pts_utm.append(Point(xc + dx, yc + dy))
                gi_list.append(gi)
                dir_list.append(dname)
        pts_conv = gpd.GeoDataFrame(geometry=all_pts_utm,
                                     crs=UTM_CRS).to_crs(WGS84_CRS)
        latlons_batch = [(row.geometry.y, row.geometry.x)
                         for _, row in pts_conv.iterrows()]
        print(f"  API 요청: {len(latlons_batch)}개 지점...")
        elevs = fetch_elev_batch(latlons_batch)
        elev_dict = {}
        for idx2, (gi, dname) in enumerate(zip(gi_list, dir_list)):
            if gi not in elev_dict:
                elev_dict[gi] = {}
            elev_dict[gi][dname] = elevs[idx2]
        with open(ELEV_CACHE, "w") as f:
            json.dump(elev_dict, f)
        print(f"  캐시 저장: {ELEV_CACHE.name}")

    slopes = np.full(n_all, np.nan)
    for gi in range(n_all):
        t = elev_dict.get(gi, {})
        ec = t.get("C")
        grads = []
        for dname in ["N", "S", "E", "W"]:
            ev = t.get(dname)
            if ec is not None and ev is not None:
                grads.append(np.degrees(np.arctan(abs(ev - ec) / ELEV_OFFSET)))
        if grads:
            slopes[gi] = np.mean(grads)
    gdf_all["slope_deg"] = np.round(slopes, 2)
    print(f"  경사도: 평균 {np.nanmean(slopes):.1f}°")

    # ── 3. 제외구역 필터링 ─────────────────────────────────────
    print("\n[3] 제외구역 필터링 (WKB 캐시)...")
    zones = {nm: load_zone(nm, LOCAL_BOX) for nm in ZONE_NAMES}
    for nm, z in zones.items():
        print(f"  [{nm:12s}] {'로드' if z is not None else '없음'}")

    excl_mask   = np.zeros(n_all, dtype=bool)
    excl_reason = [""] * n_all
    for nm in ZONE_NAMES:
        z = zones[nm]
        if z is None or z.is_empty:
            continue
        pz = prep(z)
        cnt = 0
        for i, pt in enumerate(pts_utm):
            if not excl_mask[i] and pz.contains(pt):
                excl_mask[i] = True
                excl_reason[i] = nm
                cnt += 1
        if cnt:
            print(f"  → [{nm}] {cnt}개 제외")

    gdf_all["excl"]   = excl_mask
    gdf_all["excl_r"] = excl_reason
    candidates = gdf_all[~excl_mask].copy().reset_index(drop=True)
    print(f"\n  필터 후 후보점: {len(candidates)}개 / {n_all}개")

    # ── 4. 도로 접근성 ─────────────────────────────────────────
    print("\n[4] 도로 접근성 계산...")
    road_dist = np.full(len(candidates), np.nan)

    if road_gdf_wgs is not None:
        road_utm = road_gdf_wgs.to_crs(UTM_CRS)
        local_roads = road_utm[road_utm.geometry.intersects(LOCAL_BOX)]
        print(f"  로컬 도로 세그먼트: {len(local_roads)}개")
        if len(local_roads) > 0:
            road_union = local_roads.geometry.unary_union
            pts_c = candidates.geometry.tolist()
            road_dist = np.array([pt.distance(road_union) for pt in pts_c])
        else:
            print("  ⚠ 로컬 도로 없음")
    else:
        print("  ⚠ 도로 데이터 없음 — 접근성 필터 생략")

    candidates["road_dist_m"] = np.round(road_dist, 0)

    if not np.all(np.isnan(road_dist)):
        before = len(candidates)
        candidates = candidates[
            candidates["road_dist_m"].isna() |
            (candidates["road_dist_m"] <= ROAD_MAX_M)
        ].copy().reset_index(drop=True)
        print(f"  도로 ≤{ROAD_MAX_M}m 필터: {before}개 → {len(candidates)}개")

    # ── 5. 점수 산정 ───────────────────────────────────────────
    print("\n[5] 점수 산정...")
    nc    = len(candidates)
    pts_c = candidates.geometry.tolist()

    # ② 지형 경사도 (15점)
    sl = candidates["slope_deg"].values.astype(float)
    sl_fill = np.where(np.isnan(sl),
                       np.nanmedian(sl) if not np.all(np.isnan(sl)) else 5.0, sl)
    s2 = np.clip((1.0 - sl_fill / 30.0) * 15.0, 0.0, 15.0)
    candidates["s2_지형"] = np.round(s2, 1)

    # ③ 전력/철도 이격도 (15점)
    d_power   = zone_dist(pts_c, zones["power"])
    d_railway = zone_dist(pts_c, zones["railway"])
    lp = np.log1p(d_power);  lr3 = np.log1p(d_railway)
    s3 = (lp / (lp.max() or 1) * 0.5 + lr3 / (lr3.max() or 1) * 0.5) * 15
    candidates["s3_전력철도"]  = np.round(s3, 1)
    candidates["d_power_km"]   = np.round(d_power / 1000, 1)
    candidates["d_railway_km"] = np.round(d_railway / 1000, 1)

    # ④ 인구 밀집 이격도 (15점)
    d_urban = zone_dist(pts_c, zones["urban_resid"])
    lu = np.log1p(d_urban)
    s4 = (lu / (lu.max() or 1)) * 15
    candidates["s4_인구이격"] = np.round(s4, 1)
    candidates["d_urban_km"]  = np.round(d_urban / 1000, 1)

    # ⑤ KIGAM 자기이상 모델기여도 (10점)
    s5_arr = np.full(nc, 6.0)
    p90_arr = np.full(nc, np.nan)
    cands_wgs = candidates.to_crs(WGS84_CRS)
    if klat_arr is not None:
        for i in range(nc):
            glat = cands_wgs.geometry.iloc[i].y
            glon = cands_wgs.geometry.iloc[i].x
            sc5, pp = kigam_score(glat, glon, klat_arr, klon_arr, kanom_arr)
            s5_arr[i] = sc5
            p90_arr[i] = pp
    candidates["s5_모델기여도"] = np.round(s5_arr, 1)
    candidates["mag_p90p10_nT"] = np.round(p90_arr, 1)

    # ⑥ 암상 적합성 (5점)
    gz = zones.get("geology");  fz = zones.get("fault")
    have_rock  = gz is not None and not gz.is_empty
    have_fault = fz is not None and not fz.is_empty
    s6_arr = np.full(nc, 2.5)

    lr_rock  = np.ones(nc)
    lr_fault = np.ones(nc)
    d_rock_arr  = np.full(nc, np.nan)
    d_fault_arr = np.full(nc, np.nan)

    if have_rock:
        d_rock = zone_dist(pts_c, gz.simplify(300))
        d_rock_arr = d_rock
        lr_rock = np.log1p(d_rock)
        candidates["d_geo_rock_km"] = np.round(d_rock / 1000, 1)
    else:
        candidates["d_geo_rock_km"] = np.nan

    if have_fault:
        d_fault = zone_dist(pts_c, fz.simplify(300))
        d_fault_arr = d_fault
        lr_fault = np.log1p(d_fault)
        candidates["d_geo_fault_km"] = np.round(d_fault / 1000, 1)
    else:
        candidates["d_geo_fault_km"] = np.nan

    if have_rock and have_fault:
        s6_arr = (lr_rock  / (lr_rock.max()  or 1) * 0.5 +
                  lr_fault / (lr_fault.max() or 1) * 0.5) * 5
    elif have_rock:
        s6_arr = (lr_rock  / (lr_rock.max()  or 1)) * 5
    elif have_fault:
        s6_arr = (lr_fault / (lr_fault.max() or 1)) * 5
    candidates["s6_암상"] = np.round(s6_arr, 1)

    print(f"  ② 지형 {s2.mean():.1f} / ③ 전력철도 {s3.mean():.1f} "
          f"/ ④ 인구 {s4.mean():.1f} / ⑤ 모델기여 {s5_arr.mean():.1f} "
          f"/ ⑥ 암상 {s6_arr.mean():.1f}")

    # 종합 점수 (60점 만점 → 100점)
    raw = s2 + s3 + s4 + s5_arr + s6_arr
    candidates["score_raw"] = np.round(raw, 1)
    candidates["score"]     = np.round(raw / 60 * 100, 1)
    candidates = candidates.sort_values("score", ascending=False).reset_index(drop=True)
    candidates["rank"]      = candidates.index + 1
    candidates["site_id"]   = sid
    candidates["site_name"] = sname
    candidates["center_lat"] = clat
    candidates["center_lon"] = clon

    top = candidates.head(TOP_N).copy()
    top_wgs = top.to_crs(WGS84_CRS)
    top["lat"] = top_wgs.geometry.y.round(6)
    top["lon"] = top_wgs.geometry.x.round(6)

    print(f"\n  ── 상위 {TOP_N}개 ─────────────────────────────────")
    cols_show = ["rank","lat","lon","score",
                 "s2_지형","s3_전력철도","s4_인구이격",
                 "s5_모델기여도","s6_암상","slope_deg",
                 "d_power_km","d_railway_km","d_urban_km","road_dist_m"]
    print(top[cols_show].to_string(index=False))

    all_results.append({
        "site": site,
        "top":  top,
        "gdf_all": gdf_all,
        "excl_mask": excl_mask,
        "zones": zones,
        "cx": cx, "cy": cy,
        "clat": clat, "clon": clon,
    })


# ════════════════════════════════════════════════════════════════
# 6. KML 출력
# ════════════════════════════════════════════════════════════════
print("\n\n[6] KML 파일 생성...")

def make_kml_placemark(rank, lat, lon, score, site_name, row, site_id):
    """KML <Placemark> 블록 생성"""
    def _v(val, fmt=".1f", unit=""):
        try:
            if np.isnan(float(val)):
                return "-"
            return f"{float(val):{fmt}}{unit}"
        except Exception:
            return str(val)

    desc = (
        f"사이트: {site_name}\n"
        f"순위: #{rank} / 서브격자 종합점수: {score:.1f}점\n"
        f"위도: {lat:.6f}°N  경도: {lon:.6f}°E\n\n"
        f"── 점수 내역 ──\n"
        f"② 지형 경사도: {_v(row.get('s2_지형',''))} / 15점 "
        f"(경사 {_v(row.get('slope_deg',''), '.1f', '°')})\n"
        f"③ 전력·철도:   {_v(row.get('s3_전력철도',''))} / 15점 "
        f"(전력 {_v(row.get('d_power_km',''), '.1f', 'km')} / "
        f"철도 {_v(row.get('d_railway_km',''), '.1f', 'km')})\n"
        f"④ 인구 이격:   {_v(row.get('s4_인구이격',''))} / 15점 "
        f"(도시 {_v(row.get('d_urban_km',''), '.1f', 'km')})\n"
        f"⑤ 모델기여도:  {_v(row.get('s5_모델기여도',''))} / 10점 "
        f"(P90-P10 {_v(row.get('mag_p90p10_nT',''), '.0f', 'nT')})\n"
        f"⑥ 암상 적합성: {_v(row.get('s6_암상',''))} / 5점 "
        f"(암상 {_v(row.get('d_geo_rock_km',''), '.1f', 'km')} / "
        f"단층 {_v(row.get('d_geo_fault_km',''), '.1f', 'km')})\n\n"
        f"도로 접근: {_v(row.get('road_dist_m',''), '.0f', 'm')}\n\n"
        f"네이버 지도: https://map.naver.com/v5/?c={lon},{lat},16,0,0,0,dh\n"
        f"카카오 지도: https://map.kakao.com/link/map/#{rank}_{site_id},{lat},{lon}"
    )
    # 순위별 스타일 ID
    style_id = "top3" if rank <= 3 else ("top10" if rank <= TOP_N else "normal")
    label = f"#{rank} {score:.1f}점"
    return f"""    <Placemark>
      <name>{label}</name>
      <description><![CDATA[{desc}]]></description>
      <styleUrl>#{style_id}</styleUrl>
      <Point>
        <coordinates>{lon},{lat},0</coordinates>
      </Point>
    </Placemark>
"""

def make_kml(results, filepath):
    """두 사이트 통합 KML"""
    kml_lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        '<Document>',
        '  <name>지구자기장 측정 후보지 서브격자 분석</name>',
        '  <description>후보지 #16·#18 1km 서브격자 현장 답사 세부 선정</description>',
        '',
        '  <!-- 스타일 정의 -->',
        '  <Style id="top3">',
        '    <IconStyle>',
        '      <color>ff0000ff</color>',   # 빨강 (최상위 1~3위)
        '      <scale>1.3</scale>',
        '      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-stars.png</href></Icon>',
        '    </IconStyle>',
        '    <LabelStyle><scale>1.1</scale></LabelStyle>',
        '  </Style>',
        '  <Style id="top10">',
        '    <IconStyle>',
        '      <color>ff00aaff</color>',   # 주황
        '      <scale>1.1</scale>',
        '      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/orange-circle.png</href></Icon>',
        '    </IconStyle>',
        '    <LabelStyle><scale>0.9</scale></LabelStyle>',
        '  </Style>',
        '  <Style id="center">',
        '    <IconStyle>',
        '      <color>ffff0000</color>',   # 파랑 (원점)
        '      <scale>1.4</scale>',
        '      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-circle.png</href></Icon>',
        '    </IconStyle>',
        '  </Style>',
        '',
    ]

    for res in results:
        site  = res["site"]
        top   = res["top"]
        clat  = res["clat"]
        clon  = res["clon"]
        sid   = site["id"]
        sname = site["name"]

        kml_lines.append(f'  <Folder>')
        kml_lines.append(f'    <name>{sname}</name>')
        kml_lines.append(f'    <description>국가격자 점수: {site["score_national"]}점 / 우선순위 {site["priority"]}등급</description>')
        kml_lines.append('')

        # 원점 마커
        kml_lines.append(f'    <Placemark>')
        kml_lines.append(f'      <name>▶ {sname} 중심</name>')
        kml_lines.append(f'      <description><![CDATA['
                         f'국가 10km격자 후보지<br>'
                         f'도상 점수: {site["score_national"]}점<br>'
                         f'우선순위: {site["priority"]}등급<br>'
                         f'위도: {clat}°N  경도: {clon}°E<br><br>'
                         f'<a href="https://map.naver.com/v5/?c={clon},{clat},14,0,0,0,dh">네이버 지도</a>  '
                         f'<a href="https://map.kakao.com/link/map/{sname},{clat},{clon}">카카오 지도</a>'
                         f']]></description>')
        kml_lines.append(f'      <styleUrl>#center</styleUrl>')
        kml_lines.append(f'      <Point><coordinates>{clon},{clat},0</coordinates></Point>')
        kml_lines.append(f'    </Placemark>')
        kml_lines.append('')

        # 상위 후보점
        for _, row in top.iterrows():
            lat  = row["lat"]
            lon  = row["lon"]
            rank = int(row["rank"])
            sc   = row["score"]
            kml_lines.append(
                make_kml_placemark(rank, lat, lon, sc, sname, row, sid)
            )

        kml_lines.append(f'  </Folder>')
        kml_lines.append('')

    kml_lines += ['</Document>', '</kml>']

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(kml_lines))
    print(f"  ✅ KML 저장: {filepath.name}")


# 합본 KML
kml_path = OUT_DIR / f"{TS}_subgrid_site16_18.kml"
make_kml(all_results, kml_path)

# 개별 KML (사이트별)
for res in all_results:
    sid = res["site"]["id"]
    p = OUT_DIR / f"{TS}_subgrid_site{sid}.kml"
    make_kml([res], p)


# ════════════════════════════════════════════════════════════════
# 7. Excel (네이버·카카오 지도 링크 포함)
# ════════════════════════════════════════════════════════════════
print("\n[7] Excel 파일 생성...")

wb = openpyxl.Workbook()
wb.remove(wb.active)   # 기본 시트 제거

# 색상 팔레트
C_HEADER  = "1F4E79"   # 진파랑
C_SITE16  = "2F75B6"   # 파랑
C_SITE18  = "375623"   # 초록
C_TOP3    = "FFE699"   # 연노랑
C_TOP5    = "EBF3FB"   # 연파랑
C_WHITE   = "FFFFFF"
C_LABEL   = "D6DCE4"   # 연회색

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_site_sheet(wb, res, sheet_name, header_color):
    ws = wb.create_sheet(sheet_name)
    site  = res["site"]
    top   = res["top"]
    sid   = site["id"]
    sname = site["name"]
    clat  = res["clat"] if "clat" in res else site["lat"]
    clon  = res["clon"] if "clon" in res else site["lon"]

    # 타이틀
    ws.merge_cells("A1:N1")
    ws["A1"] = f"측정 후보지 #{sid} — 1km 서브격자 현장 답사 세부 선정"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hdr_fill(header_color)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 메타 정보
    meta = [
        ("사이트명", sname),
        ("10km격자 중심", f"{clat}°N, {clon}°E"),
        ("국가격자 점수", f"{site['score_national']}점 (우선순위 {site['priority']}등급)"),
        ("분석 파라미터",
         f"반경 {RADIUS_M//1000}km / 격자 {GRID_STEP//1000}km / 도로≤{ROAD_MAX_M//1000}km"),
        ("생성 일시", datetime.now().strftime("%Y년 %m월 %d일 %H:%M")),
    ]
    for r_off, (k, v) in enumerate(meta, start=2):
        ws[f"A{r_off}"] = k
        ws[f"A{r_off}"].font = Font(bold=True, name="맑은 고딕", size=10)
        ws[f"A{r_off}"].fill = hdr_fill(C_LABEL)
        ws[f"A{r_off}"].alignment = Alignment(horizontal="right")
        ws.merge_cells(f"B{r_off}:N{r_off}")
        ws[f"B{r_off}"] = v
        ws[f"B{r_off}"].font = Font(name="맑은 고딕", size=10)

    # 빈 줄
    ROW_H = len(meta) + 2 + 1  # = 8

    # 컬럼 헤더
    headers = [
        ("순위", 5), ("위도(°N)", 12), ("경도(°E)", 12),
        ("종합점수", 9), ("②지형\n/15", 8), ("③전력철도\n/15", 9),
        ("④인구이격\n/15", 9), ("⑤모델기여\n/10", 9), ("⑥암상\n/5", 7),
        ("경사도(°)", 9), ("전력(km)", 8), ("철도(km)", 8),
        ("네이버 지도", 18), ("카카오 지도", 18),
    ]
    for col_idx, (h_text, w) in enumerate(headers, start=1):
        cell = ws.cell(row=ROW_H, column=col_idx, value=h_text)
        cell.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        cell.fill = hdr_fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[ROW_H].height = 32

    # 데이터 행
    top_wgs = top.to_crs(WGS84_CRS)
    for row_off, (_, row) in enumerate(top.iterrows(), start=1):
        r = ROW_H + row_off
        rank  = int(row["rank"])
        lat   = round(float(top_wgs.geometry.iloc[row_off - 1].y), 6)
        lon   = round(float(top_wgs.geometry.iloc[row_off - 1].x), 6)
        score = float(row["score"])

        # 배경색
        if rank <= 3:
            bg = C_TOP3
        elif rank <= 5:
            bg = C_TOP5
        else:
            bg = C_WHITE

        def _val(key, fmt=".1f"):
            try:
                v = float(row.get(key, np.nan))
                return "-" if np.isnan(v) else f"{v:{fmt}}"
            except Exception:
                return "-"

        naver_url = f"https://map.naver.com/v5/?c={lon},{lat},16,0,0,0,dh"
        kakao_url = f"https://map.kakao.com/link/map/#{rank}_{sid},{lat},{lon}"

        row_data = [
            rank, lat, lon, score,
            _val("s2_지형"), _val("s3_전력철도"), _val("s4_인구이격"),
            _val("s5_모델기여도"), _val("s6_암상"),
            _val("slope_deg"), _val("d_power_km"), _val("d_railway_km"),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(name="맑은 고딕", size=10,
                             bold=(col_idx in [1, 4]))
            cell.fill = hdr_fill(bg)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

        # 네이버 지도 링크
        nc = ws.cell(row=r, column=13,
                     value=f"🗺 네이버 #{rank}")
        nc.hyperlink = naver_url
        nc.font  = Font(name="맑은 고딕", size=10, color="0563C1",
                        underline="single")
        nc.fill  = hdr_fill(bg)
        nc.alignment = Alignment(horizontal="center")
        nc.border = thin_border()

        # 카카오 지도 링크
        kc = ws.cell(row=r, column=14,
                     value=f"🗺 카카오 #{rank}")
        kc.hyperlink = kakao_url
        kc.font  = Font(name="맑은 고딕", size=10, color="0563C1",
                        underline="single")
        kc.fill  = hdr_fill(bg)
        kc.alignment = Alignment(horizontal="center")
        kc.border = thin_border()

    # 범례
    legend_row = ROW_H + TOP_N + 2
    ws.merge_cells(f"A{legend_row}:N{legend_row}")
    ws[f"A{legend_row}"] = (
        "★ 상위 1~3위(노랑) — 최우선 현장 확인 대상  "
        "● 4~5위(하늘) — 대안 후보  "
        "○ 6~10위(흰색) — 참고 후보  "
        "│ 종합점수 = (②+③+④+⑤+⑥) / 60 × 100"
    )
    ws[f"A{legend_row}"].font = Font(name="맑은 고딕", size=9, italic=True,
                                     color="595959")
    ws[f"A{legend_row}"].fill = hdr_fill(C_LABEL)

    # 창 고정 (헤더 행)
    ws.freeze_panes = f"A{ROW_H + 1}"


# 시트별 색상
SITE_COLORS = {16: C_SITE16, 18: C_SITE18}
for res in all_results:
    sid = res["site"]["id"]
    make_site_sheet(wb, res, f"후보지 #{sid}", SITE_COLORS.get(sid, C_HEADER))

# ── 합본 요약 시트 ──────────────────────────────────────────────
ws_sum = wb.create_sheet("📋 전체 요약", 0)
ws_sum.merge_cells("A1:L1")
ws_sum["A1"] = "지구자기장 측정 후보지 서브격자 분석 — 전체 요약"
ws_sum["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
ws_sum["A1"].fill = hdr_fill(C_HEADER)
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 28

ws_sum["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  /  " \
               f"분석 조건: 반경{RADIUS_M//1000}km / 격자{GRID_STEP//1000}km / 도로≤{ROAD_MAX_M//1000}km"
ws_sum["A2"].font = Font(name="맑은 고딕", size=9, italic=True)
ws_sum.merge_cells("A2:L2")

headers_sum = [
    ("사이트", 14), ("순위", 5), ("위도(°N)", 12), ("경도(°E)", 12),
    ("종합점수", 9), ("②지형", 7), ("③전력철도", 8), ("④인구", 7),
    ("⑤모델기여", 8), ("⑥암상", 6), ("네이버 지도", 18), ("카카오 지도", 18),
]
for ci, (h, w) in enumerate(headers_sum, start=1):
    c = ws_sum.cell(row=3, column=ci, value=h)
    c.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    c.fill = hdr_fill(C_HEADER)
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()
    ws_sum.column_dimensions[get_column_letter(ci)].width = w
ws_sum.row_dimensions[3].height = 22

row_sum = 4
SITE_HDR_COLORS = {16: C_SITE16, 18: C_SITE18}

for res in all_results:
    site = res["site"]
    sid  = site["id"]
    top  = res["top"]
    top_wgs = top.to_crs(WGS84_CRS)
    hcol = SITE_HDR_COLORS.get(sid, C_HEADER)

    # 사이트 소제목 행
    ws_sum.merge_cells(f"A{row_sum}:L{row_sum}")
    c = ws_sum[f"A{row_sum}"]
    c.value = f"▶ {site['name']}  (국가격자 {site['score_national']}점, {site['priority']}등급)"
    c.font  = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    c.fill  = hdr_fill(hcol)
    c.alignment = Alignment(horizontal="left")
    ws_sum.row_dimensions[row_sum].height = 20
    row_sum += 1

    for idx2, (_, row) in enumerate(top.iterrows(), start=1):
        rank  = int(row["rank"])
        lat   = round(float(top_wgs.geometry.iloc[idx2 - 1].y), 6)
        lon   = round(float(top_wgs.geometry.iloc[idx2 - 1].x), 6)
        score = float(row["score"])

        bg = C_TOP3 if rank <= 3 else (C_TOP5 if rank <= 5 else C_WHITE)

        def _v(key, fmt=".1f"):
            try:
                v = float(row.get(key, np.nan))
                return "-" if np.isnan(v) else f"{v:{fmt}}"
            except Exception:
                return "-"

        row_vals = [
            site["name"], rank, lat, lon, score,
            _v("s2_지형"), _v("s3_전력철도"), _v("s4_인구이격"),
            _v("s5_모델기여도"), _v("s6_암상"),
        ]
        for ci, val in enumerate(row_vals, start=1):
            c = ws_sum.cell(row=row_sum, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=(ci in [5]))
            c.fill = hdr_fill(bg)
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()

        naver_url = f"https://map.naver.com/v5/?c={lon},{lat},16,0,0,0,dh"
        kakao_url = f"https://map.kakao.com/link/map/#{rank}_{sid},{lat},{lon}"

        nc = ws_sum.cell(row=row_sum, column=11, value=f"🗺 네이버 #{rank}")
        nc.hyperlink = naver_url
        nc.font  = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
        nc.fill  = hdr_fill(bg)
        nc.alignment = Alignment(horizontal="center")
        nc.border = thin_border()

        kc = ws_sum.cell(row=row_sum, column=12, value=f"🗺 카카오 #{rank}")
        kc.hyperlink = kakao_url
        kc.font  = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
        kc.fill  = hdr_fill(bg)
        kc.alignment = Alignment(horizontal="center")
        kc.border = thin_border()

        row_sum += 1

    row_sum += 1   # 사이트 구분 빈 줄

ws_sum.freeze_panes = "A4"

# Excel 저장
xl_path = OUT_DIR / f"{TS}_subgrid_site16_18_survey.xlsx"
wb.save(xl_path)
print(f"  ✅ Excel 저장: {xl_path.name}")


# ════════════════════════════════════════════════════════════════
# 완료 요약
# ════════════════════════════════════════════════════════════════
print("\n" + "=" * 64)
print("  ✅ 완료!")
print(f"  KML  : {kml_path.name}")
for res in all_results:
    sid = res["site"]["id"]
    print(f"         (개별) {TS}_subgrid_site{sid}.kml")
print(f"  Excel: {xl_path.name}")
print(f"  저장 위치: {OUT_DIR}")
print()
print("  ── KML 사용법 ───────────────────────────────────────")
print("  Google Earth: 파일 → KML 열기")
print("  네이버 지도: 내 장소 → KML 가져오기")
print("  카카오맵: 장소 저장 후 공유 기능 활용")
print("  스마트폰: KML 앱(Maps.me, OsmAnd 등)으로 직접 열기")
print("=" * 64)
