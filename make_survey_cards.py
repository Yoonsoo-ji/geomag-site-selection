# -*- coding: utf-8 -*-
"""
도상선점 조사 카드 + 마스터 목록 생성기
- 목적: ① 현장 작업자 배포용 정보 ② 도상선점 위치 정리·보관용
- 구성: 마스터 목록 시트 + 후보지별 카드(지도 4종 포함)
- 섹션: ① 선점 정보 / ② 주변 간섭요인(자동계산) / ③ 도상 판정 / ④ 현장조사 결과(공란)
"""
import json, re, math, os, time
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET
import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from shapely.geometry import shape, Point

from capture_map_images import MapCapturer, IMG_DIR
from interference import Interference

ROOT     = Path(__file__).parent
KML_PATH = ROOT / "docs/output/20260617_kang_site_v4.kml"
EXISTING = ROOT / "docs/data/existing_sites.geojson"
TOPO     = ROOT / "docs/data/topo_sheets.geojson"
OUT_DIR  = ROOT / "docs/output"
N_CARDS  = 103
EMBED_PX = 300
FONT_KR  = "C:/Windows/Fonts/malgun.ttf"
KAKAO_KEY = os.environ.get("KAKAO_REST_KEY", "")   # 역지오코딩 키 (환경변수)

NS = {'k': 'http://www.opengis.net/kml/2.2'}


# ── Kakao 역지오코딩 (좌표 → 지번/도로명 주소) ──────────────────────────────
_geo_cache = {}

def reverse_geocode(lat, lon):
    """returns (지번주소, 도로명주소). 키 없거나 실패 시 ('',''))"""
    if not KAKAO_KEY:
        return "", ""
    key = (round(lon, 6), round(lat, 6))
    if key in _geo_cache:
        return _geo_cache[key]
    jibun = road = ""
    try:
        r = requests.get("https://dapi.kakao.com/v2/local/geo/coord2address.json",
                         params={"x": lon, "y": lat},
                         headers={"Authorization": f"KakaoAK {KAKAO_KEY}"}, timeout=10)
        docs = r.json().get("documents", [])
        if docs:
            jibun = (docs[0].get("address") or {}).get("address_name", "") or ""
            rd = docs[0].get("road_address")
            road = (rd or {}).get("address_name", "") if rd else ""
    except Exception as e:
        print("  [geocode 실패]", lat, lon, e)
    _geo_cache[key] = (jibun, road)
    time.sleep(0.04)
    return jibun, road

# ── 스타일 ──────────────────────────────────────────────────────────────────
F_TITLE  = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
F_SEC    = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
F_LBL    = Font(name='맑은 고딕', size=10, bold=True)
F_VAL    = Font(name='맑은 고딕', size=10)
F_CAP    = Font(name='맑은 고딕', size=9, italic=True, color='555555')
FILL_TITLE = PatternFill('solid', fgColor='1F3864')
FILL_SEC   = PatternFill('solid', fgColor='4472C4')
FILL_LBL   = PatternFill('solid', fgColor='F2F2F2')
FILL_FIELD = PatternFill('solid', fgColor='FFF8E1')   # 현장 기입란 강조
AL_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
_thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def clean_html(s):
    if not s:
        return ""
    s = re.sub(r'<br\s*/?>', '\n', s, flags=re.I)
    s = re.sub(r'</div>', '\n', s, flags=re.I)
    s = re.sub(r'<[^>]+>', '', s)
    return re.sub(r'\n{2,}', '\n', s).strip()


def parse_kml():
    root = ET.parse(KML_PATH).getroot()
    out = []
    for pm in root.findall('.//k:Placemark', NS):
        nm = pm.find('k:name', NS); ds = pm.find('k:description', NS)
        co = pm.find('.//k:coordinates', NS)
        if co is None:
            continue
        lon, lat, *rest = [float(x) for x in co.text.strip().split(',')]
        out.append({"name": nm.text.strip() if nm is not None else "",
                    "desc": clean_html(ds.text if ds is not None else ""),
                    "lon": lon, "lat": lat, "elev": rest[0] if rest else None})
    return out


def classify(name, desc):
    c = (name + " " + desc).lower()
    if "간선임도" in c:                                      return "간선임도"
    if "작업임도" in c:                                      return "작업임도"
    if "지선임도" in c:                                      return "지선임도"
    if "임도 아님" in c or "임도아님" in c or "임도 없음" in c or "임도없음" in c:
        return "임도 외"
    if "임도" in c:                                          return "임도(종류미상)"
    return "기타"


def dms_str(dec, hp, hn):
    h = hp if dec >= 0 else hn
    dec = abs(dec); d = int(dec); mf = (dec - d) * 60; m = int(mf); s = (mf - m) * 60
    return f"{d}°{m:02d}′{s:05.2f}″{h}"


def haversine(a, b, c, d):
    R = 6371.0
    p1, p2 = math.radians(a), math.radians(c)
    dp, dl = math.radians(c - a), math.radians(d - b)
    h = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(h))


def nearest_existing(lat, lon, ex):
    best, bd = None, 1e9
    for e in ex:
        dd = haversine(lat, lon, e['lat'], e['lon'])
        if dd < bd:
            bd, best = dd, e
    return best, bd


def sheet_of(lon, lat, polys):
    pt = Point(lon, lat)
    for prop, g in polys:
        if g.contains(pt):
            return prop
    return {}


def vehicle_access(desc):
    d = desc.lower()
    if "진입 불가" in d or "진입불가" in d or "접근 불가" in d:
        return "불가 (도보)"
    if "차량" in d and ("불가" in d or "힘들" in d):
        return "제한적"
    if "가능" in d:
        return "가능(추정)"
    return "현장확인"


def cautions(desc):
    d = desc; tags = []
    if "차단기" in d:                       tags.append("임도 차단기")
    if "사유지" in d or "사유" in d:          tags.append("사유지 가능")
    if "급경사" in d or "경사" in d:          tags.append("경사 주의")
    if "전신주" in d or "전주" in d:          tags.append("전주 인접")
    return ", ".join(tags) if tags else "특이사항 없음(현장확인)"


def judge(inf):
    """간섭 거리로 도상 1차 판정"""
    issues = []
    if inf['tower'] < 1.0: issues.append(f"송전탑 {inf['tower']*1000:.0f}m(<1km)")
    if inf['rail']  < 5.0: issues.append(f"철도 {inf['rail']:.1f}km(<5km)")
    if not issues:
        return "적합", "주요 간섭요인(송전탑·철도) 이격기준 충족."
    return "조건부 적합", "기준 미달 항목: " + ", ".join(issues) + " → 현장 정밀확인 필요."


# ── 지도 삽입 (캡션 burn-in) ────────────────────────────────────────────────
def embed_map(ws, png, anchor, caption):
    if not png or not Path(png).exists():
        return
    img = Image.open(png).convert('RGB')
    d = ImageDraw.Draw(img)
    try:
        fnt = ImageFont.truetype(FONT_KR, 26)
    except Exception:
        fnt = ImageFont.load_default()
    bar = 38
    d.rectangle([0, 0, img.width, bar], fill=(31, 56, 100))
    d.text((10, 6), caption, fill=(255, 255, 255), font=fnt)
    if img.width != EMBED_PX:
        h = int(img.height * EMBED_PX / img.width)
        img = img.resize((EMBED_PX, h), Image.LANCZOS)
    small = Path(png).with_name(Path(png).stem + "_e.png")
    img.save(small)
    ws.add_image(XLImage(str(small)), anchor)


# ── 카드 작성 ───────────────────────────────────────────────────────────────
def lbl(ws, ref, text):
    c = ws[ref]; c.value = text; c.font = F_LBL; c.fill = FILL_LBL
    c.alignment = AL_C; c.border = BORDER


def val(ws, ref, text, field=False):
    c = ws[ref]; c.value = text; c.font = F_VAL
    c.alignment = AL_L; c.border = BORDER
    if field:
        c.fill = FILL_FIELD


def sec(ws, row, text):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]; c.value = text; c.font = F_SEC; c.fill = FILL_SEC
    c.alignment = AL_L


def merge_val(ws, row, lab, span_lab, span_val, text, field=False):
    """라벨 + 병합 값"""
    lbl(ws, f"{span_lab}{row}", lab)
    a, b = span_val
    ws.merge_cells(f"{a}{row}:{b}{row}")
    val(ws, f"{a}{row}", text, field)


def build_card(wb, idx, p, exi, dist, sh, inf, imgs):
    sid = f"DS-{idx:03d}"
    safe = re.sub(r'[\\/*?:\[\]]', '', p['name'])[:22]
    ws = wb.create_sheet(f"{idx:02d}_{safe}")

    for col, w in zip("ABCDEFG", [13, 11, 11, 11, 11, 11, 11]):
        ws.column_dimensions[col].width = w
    for col in "HIJK":
        ws.column_dimensions[col].width = 21

    cat = classify(p['name'], p['desc'])
    jd, jr = judge(inf)

    # 제목
    ws.merge_cells("A1:K1")
    t = ws["A1"]; t.value = f"도상 선점 조사 카드   |   {sid}   {p['name']}"
    t.font = F_TITLE; t.fill = FILL_TITLE; t.alignment = AL_L
    ws.row_dimensions[1].height = 30

    # ① 선점 정보
    sec(ws, 2, "① 선점 정보")
    merge_val(ws, 3, "관리번호", "A", ("B", "C"), sid)
    lbl(ws, "D3", "후보지명"); ws.merge_cells("E3:G3"); val(ws, "E3", p['name'])
    merge_val(ws, 4, "도엽번호", "A", ("B", "C"), sh.get('sheet_code', '-'))
    lbl(ws, "D4", "도엽명"); ws.merge_cells("E4:G4"); val(ws, "E4", sh.get('sheet_name', '-'))
    merge_val(ws, 5, "위도(십진)", "A", ("B", "C"), f"{p['lat']:.6f}")
    lbl(ws, "D5", "경도(십진)"); ws.merge_cells("E5:G5"); val(ws, "E5", f"{p['lon']:.6f}")
    merge_val(ws, 6, "좌표(도분초)", "A", ("B", "G"),
              f"{dms_str(p['lat'],'N','S')}  /  {dms_str(p['lon'],'E','W')}")
    merge_val(ws, 7, "표고", "A", ("B", "C"),
              f"{p['elev']:.0f} m" if p['elev'] is not None else "-")
    lbl(ws, "D7", "후보지 유형"); ws.merge_cells("E7:G7"); val(ws, "E7", cat)
    merge_val(ws, 8, "지번 주소", "A", ("B", "G"), p.get('jibun') or "-")
    merge_val(ws, 9, "도로명 주소", "A", ("B", "G"), p.get('road') or "-")
    merge_val(ws, 10, "연계 기존점", "A", ("B", "G"),
              f"{exi['name']} (직선거리 약 {dist:.1f} km)")
    merge_val(ws, 11, "접근 경로", "A", ("B", "G"), p['desc'] or "현장확인")
    ws.row_dimensions[11].height = 58
    merge_val(ws, 12, "차량 진입", "A", ("B", "C"), vehicle_access(p['desc']))
    lbl(ws, "D12", "주차→도보"); ws.merge_cells("E12:G12"); val(ws, "E12", "현장확인")
    merge_val(ws, 13, "유의사항", "A", ("B", "G"), cautions(p['desc']))
    # 내비 링크
    lbl(ws, "A14", "내비 바로가기")
    kakao = f"https://map.kakao.com/link/map/{p['name']},{p['lat']:.6f},{p['lon']:.6f}"
    naver = f"https://map.naver.com/v5/search/{p['lat']:.6f},{p['lon']:.6f}"
    google = f"https://www.google.com/maps/search/?api=1&query={p['lat']:.6f},{p['lon']:.6f}"
    for ref, span, txt, url in [("B", "C", "카카오맵", kakao),
                                 ("D", "E", "네이버지도", naver),
                                 ("F", "G", "구글지도", google)]:
        ws.merge_cells(f"{ref}14:{chr(ord(ref)+1)}14")
        c = ws[f"{ref}14"]; c.value = txt; c.hyperlink = url
        c.font = Font(name='맑은 고딕', size=10, color='1A73E8', underline='single')
        c.alignment = AL_C; c.border = BORDER

    # ② 주변 간섭요인
    sec(ws, 15, "② 주변 간섭요인 (자동 계산)")
    ok = lambda b: "✓ 충족" if b else "✗ 미달"
    merge_val(ws, 16, "최근접 송전탑", "A", ("B", "C"), f"{inf['tower']*1000:.0f} m")
    lbl(ws, "D16", "기준 1km"); ws.merge_cells("E16:G16"); val(ws, "E16", ok(inf['tower'] >= 1.0))
    merge_val(ws, 17, "최근접 철도", "A", ("B", "C"), f"{inf['rail']:.1f} km")
    lbl(ws, "D17", "기준 5km"); ws.merge_cells("E17:G17"); val(ws, "E17", ok(inf['rail'] >= 5.0))
    merge_val(ws, 18, "최근접 통신탑", "A", ("B", "C"), f"{inf['comm']:.1f} km")
    lbl(ws, "D18", "최근접 주거지"); ws.merge_cells("E18:G18"); val(ws, "E18", f"{inf['resid']:.2f} km")

    # ③ 도상 판정
    sec(ws, 19, "③ 도상 판정")
    mark = {"적합": "■ 적합  □ 조건부 적합  □ 부적합",
            "조건부 적합": "□ 적합  ■ 조건부 적합  □ 부적합"}.get(jd)
    merge_val(ws, 20, "종합판정", "A", ("B", "G"), mark)
    merge_val(ws, 21, "종합의견", "A", ("B", "G"), jr +
              " 현장답사를 통해 접근성·평탄성·방위각 측정 가능 여부 확인 권장.")
    ws.row_dimensions[21].height = 56

    # ④ 현장조사 결과 (작업자 기입)
    sec(ws, 22, "④ 현장조사 결과 (현장 작업자 기입)")
    merge_val(ws, 23, "방문일", "A", ("B", "C"), "", field=True)
    lbl(ws, "D23", "조사자"); ws.merge_cells("E23:G23"); val(ws, "E23", "", field=True)
    merge_val(ws, 24, "실측 GPS", "A", ("B", "G"), "", field=True)
    merge_val(ws, 25, "실제 접근성", "A", ("B", "C"), "", field=True)
    lbl(ws, "D25", "장비설치 가능"); ws.merge_cells("E25:G25"); val(ws, "E25", "", field=True)
    merge_val(ws, 26, "측정 적합성", "A", ("B", "C"), "", field=True)
    lbl(ws, "D26", "사진번호"); ws.merge_cells("E26:G26"); val(ws, "E26", "", field=True)
    merge_val(ws, 27, "최종판정", "A", ("B", "C"), "", field=True)
    lbl(ws, "D27", "비고"); ws.merge_cells("E27:G27"); val(ws, "E27", "", field=True)

    # 지도 (우측 2x2)
    embed_map(ws, imgs['sat'],  'H2',  "위성 근접 (현장 지형·임도)")
    embed_map(ws, imgs['sur'],  'J2',  "주변 (도로·마을)")
    embed_map(ws, imgs['int'],  'H17', "간섭요인 (송전탑·철도 버퍼)")
    embed_map(ws, imgs['wide'], 'J17', "광역 위치")

    return ws, sid, cat, jd


# ── 마스터 목록 ─────────────────────────────────────────────────────────────
MASTER_COLS = ["관리번호", "후보지명", "지번 주소", "도엽번호", "도엽명", "위도", "경도", "표고(m)",
               "유형", "연계 기존점", "기존점거리(km)", "최근접송전탑(m)", "송전탑(1km)",
               "최근접철도(km)", "철도(5km)", "최근접주거지(km)", "도상판정", "진행상태", "카드"]


def build_master(wb, rows):
    ws = wb.create_sheet("전체 목록", 0)
    widths = [9, 14, 24, 11, 9, 10, 10, 8, 11, 12, 12, 13, 10, 12, 9, 13, 12, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells(f"A1:{get_column_letter(len(MASTER_COLS))}1")
    t = ws["A1"]; t.value = "도상 선점 후보지 전체 목록"
    t.font = F_TITLE; t.fill = FILL_TITLE; t.alignment = AL_L
    ws.row_dimensions[1].height = 26
    for j, h in enumerate(MASTER_COLS, 1):
        c = ws.cell(2, j, h); c.font = F_LBL; c.fill = FILL_SEC
        c.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
        c.alignment = AL_C; c.border = BORDER
    for r, row in enumerate(rows, 3):
        for j, v in enumerate(row[:-1], 1):
            c = ws.cell(r, j, v); c.font = F_VAL; c.alignment = AL_C; c.border = BORDER
        link_cell = ws.cell(r, len(MASTER_COLS), "이동")
        link_cell.hyperlink = f"#'{row[-1]}'!A1"
        link_cell.font = Font(name='맑은 고딕', size=10, color='1A73E8', underline='single')
        link_cell.alignment = AL_C; link_cell.border = BORDER
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(MASTER_COLS))}{len(rows)+2}"


def main():
    pts = parse_kml()[:N_CARDS]
    ex = [f['properties'] for f in json.load(open(EXISTING, encoding='utf-8'))['features']]
    polys = [(f['properties'], shape(f['geometry']))
             for f in json.load(open(TOPO, encoding='utf-8'))['features']]
    print("간섭요인 데이터 로딩...")
    inf_calc = Interference()
    print("지도 캡처 시작...")
    cap = MapCapturer()

    wb = Workbook(); wb.remove(wb.active)
    master_rows = []
    for i, p in enumerate(pts, 1):
        exi, dist = nearest_existing(p['lat'], p['lon'], ex)
        # 도엽은 **후보지 자신의 좌표**로 판정한다. 기존점(exi) 좌표를 쓰면
        # 연계 기존점의 도엽이 찍혀 실제 도엽과 어긋난다(2026-08 정정).
        sh = sheet_of(p['lon'], p['lat'], polys)
        inf = inf_calc.all(p['lat'], p['lon'])
        p['jibun'], p['road'] = reverse_geocode(p['lat'], p['lon'])
        imgs = cap.capture_card(f"{i:02d}", p['lat'], p['lon'])
        ws, sid, cat, jd = build_card(wb, i, p, exi, dist, sh, inf, imgs)
        master_rows.append([
            sid, p['name'], p.get('jibun') or "-",
            sh.get('sheet_code', '-'), sh.get('sheet_name', '-'),
            f"{p['lat']:.5f}", f"{p['lon']:.5f}",
            f"{p['elev']:.0f}" if p['elev'] is not None else "-",
            cat, exi['name'], f"{dist:.1f}",
            f"{inf['tower']*1000:.0f}", "✓" if inf['tower'] >= 1 else "✗",
            f"{inf['rail']:.1f}", "✓" if inf['rail'] >= 5 else "✗",
            f"{inf['resid']:.2f}", jd, "도상선점", ws.title,
        ])
        print(f"  [{i}/{len(pts)}] {p['name']}  판정:{jd}")
    cap.close()

    build_master(wb, master_rows)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUT_DIR / f"{ts}_도상선점_조사카드_{N_CARDS}건.xlsx"
    wb.save(out)
    print(f"저장 완료: {out}")


if __name__ == "__main__":
    main()
